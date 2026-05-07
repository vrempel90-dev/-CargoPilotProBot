from __future__ import annotations

import asyncio
from contextlib import asynccontextmanager
import html
import logging
import os
import re
from io import BytesIO
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Iterable, Optional

import aiosqlite
import qrcode
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
from aiohttp import web
from aiogram import Bot, Dispatcher, F, Router
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.exceptions import TelegramBadRequest
from aiogram.filters import Command, CommandStart
from aiogram.filters.command import CommandObject
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    CallbackQuery,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    KeyboardButton,
    Message,
    ReplyKeyboardMarkup,
    FSInputFile,
    BufferedInputFile,
)
from dotenv import load_dotenv

load_dotenv()

# =========================
# CONFIG
# =========================
BOT_TOKEN = os.getenv("BOT_TOKEN", "").strip()
if not BOT_TOKEN:
    raise RuntimeError("BOT_TOKEN is empty. Add it to .env or hosting environment variables.")

COMPANY_NAME = os.getenv("COMPANY_NAME", "CargoAI CRM").strip()
DATA_DIR = Path(os.getenv("DATA_DIR", ".")).resolve()
DATA_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH = os.getenv("DB_PATH", str(DATA_DIR / "cargo_ai_crm.sqlite3"))
EXPORT_DIR = DATA_DIR / "exports"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)
RECEIPT_DIR = DATA_DIR / "receipts"
RECEIPT_DIR.mkdir(parents=True, exist_ok=True)
PORT = int(os.getenv("PORT", "8080"))
CURRENCY = os.getenv("CURRENCY", "$").strip() or "$"

# Публичная ссылка для веб-трекинга.
# Для Railway можно оставить дефолт или задать свою ссылку в Variables:
# PUBLIC_BASE_URL=https://your-service.up.railway.app
PUBLIC_BASE_URL = (
    os.getenv("PUBLIC_BASE_URL")
    or os.getenv("PUBLIC_WEB_URL")
    or os.getenv("PUBLIC_TRACK_BASE_URL")
    or "https://cargopilotprobot-production.up.railway.app"
).strip().rstrip("/")

# CargoPilot SmartFlow: уникальные функции контроля грузов.
ISSUE_STALE_HOURS = int(os.getenv("ISSUE_STALE_HOURS", "24"))
OWNER_REPORT_ENABLED = os.getenv("OWNER_REPORT_ENABLED", "false").strip().lower() in {"1", "true", "yes", "on"}
OWNER_REPORT_INTERVAL_SECONDS = int(os.getenv("OWNER_REPORT_INTERVAL_SECONDS", "86400"))

# CargoPilot TrustFlow: не просто статус, а понятное объяснение для клиента.
# Показывает: всё по плану / возможна задержка / нужна проверка.
TRUSTFLOW_ENABLED = os.getenv("TRUSTFLOW_ENABLED", "true").strip().lower() in {"1", "true", "yes", "on"}
TRUSTFLOW_GRACE_HOURS = int(os.getenv("TRUSTFLOW_GRACE_HOURS", "12"))
BOT_USERNAME = (os.getenv("BOT_USERNAME") or "CargoPilotProBot").strip().lstrip("@")

# CargoPromise OS: контроль обещаний, рисков и доверия.
PROMISE_OS_ENABLED = os.getenv("PROMISE_OS_ENABLED", "true").strip().lower() in {"1", "true", "yes", "on"}
PROMISE_STALE_HOURS = int(os.getenv("PROMISE_STALE_HOURS", "24"))
PROMISE_HIGH_TRACK_VIEWS = int(os.getenv("PROMISE_HIGH_TRACK_VIEWS", "5"))

# Demo Bank Payments: виртуальные банки для показа клиенту.
# Это НЕ реальная оплата и НЕ списывает деньги.
DEMO_PAYMENT_ENABLED = os.getenv("DEMO_PAYMENT_ENABLED", "true").strip().lower() in {"1", "true", "yes", "on"}
DEMO_PAYMENT_AMOUNT = float(os.getenv("DEMO_PAYMENT_AMOUNT", "10000"))

DEMO_PAYMENT_PROVIDERS = {
    "kaspi": {
        "title": "Kaspi Pay Demo",
        "label": "Kaspi",
        "subtitle": "Оплата через Kaspi QR / Kaspi Pay",
        "color": "#e31e24",
    },
    "halyk": {
        "title": "Halyk Bank Demo",
        "label": "Halyk",
        "subtitle": "Оплата картой / Halyk QR",
        "color": "#00a651",
    },
    "bcc": {
        "title": "ЦентрКредит Demo",
        "label": "BCC",
        "subtitle": "Оплата через Bank CenterCredit",
        "color": "#0060a9",
    },
    "freedom": {
        "title": "Freedom Bank Demo",
        "label": "Freedom",
        "subtitle": "Оплата через Freedom Bank",
        "color": "#ffb000",
    },
}

# Бесплатные автостатусы: бот сам меняет статус по срокам маршрута.
# Не требует WhatsApp API, OpenAI, Gemini, Kaspi API или других платных сервисов.
AUTO_STATUS_ENABLED = os.getenv("AUTO_STATUS_ENABLED", "true").strip().lower() in {"1", "true", "yes", "on"}
AUTO_STATUS_INTERVAL_SECONDS = int(os.getenv("AUTO_STATUS_INTERVAL_SECONDS", "300"))
# days = реальная работа, minutes = демо-режим для быстрых показов клиенту. Оба режима бесплатные.
AUTO_STATUS_TIME_UNIT = os.getenv("AUTO_STATUS_TIME_UNIT", "days").strip().lower()
# По новой логике автостатусы НЕ включаются сразу. Сначала клиент оплачивает, потом админ включает их кнопкой.
AUTO_STATUS_ON_NEW_ORDERS = os.getenv("AUTO_STATUS_ON_NEW_ORDERS", "false").strip().lower() in {"1", "true", "yes", "on"}
AUTO_STATUS_AFTER_PAYMENT_ONLY = os.getenv("AUTO_STATUS_AFTER_PAYMENT_ONLY", "true").strip().lower() in {"1", "true", "yes", "on"}
REGION_MODE = os.getenv("REGION_MODE", "CIS").strip() or "CIS"

ADMIN_IDS = {
    int(x.strip())
    for x in os.getenv("ADMIN_IDS", "").split(",")
    if x.strip().lstrip("-").isdigit()
}
WAREHOUSE_IDS = {
    int(x.strip())
    for x in os.getenv("WAREHOUSE_IDS", "").split(",")
    if x.strip().lstrip("-").isdigit()
}
COURIER_IDS = {
    int(x.strip())
    for x in os.getenv("COURIER_IDS", "").split(",")
    if x.strip().lstrip("-").isdigit()
}

DEFAULT_RATES = {
    # Основные направления карго
    "китай": 3.5,
    "china": 3.5,
    "турция": 4.2,
    "turkey": 4.2,
    "дубай": 5.0,
    "uae": 5.0,
    "оаэ": 5.0,

    # СНГ / локальная логистика
    "казахстан": 2.0,
    "kazakhstan": 2.0,
    "россия": 2.5,
    "russia": 2.5,
    "узбекистан": 2.8,
    "uzbekistan": 2.8,
    "кыргызстан": 2.6,
    "киргизия": 2.6,
    "kyrgyzstan": 2.6,
    "беларусь": 3.0,
    "belarus": 3.0,
    "армения": 3.2,
    "armenia": 3.2,
    "азербайджан": 3.2,
    "azerbaijan": 3.2,
    "таджикистан": 3.0,
    "tajikistan": 3.0,
    "молдова": 3.1,
    "moldova": 3.1,
}
DEFAULT_COMMISSION_PERCENT = float(os.getenv("PARTNER_COMMISSION_PERCENT", "5"))

STATUSES = [
    "новая заявка",
    "принят на склад",
    "ожидает отправки",
    "отправлен",
    "в пути",
    "на таможне",
    "прибыл в страну",
    "прибыл в Казахстан",
    "прибыл в город",
    "готов к выдаче",
    "передан курьеру",
    "доставлен",
    "проблема",
    "отменён",
]

# Бесплатные шаблоны автостатусов. День считается от даты создания заявки.
# Это не интеграция с реальным складом/API, а автоматизация по типовым срокам маршрута.
AUTO_STATUS_ROUTES = {
    "china_cis": {
        "name": "Китай → СНГ",
        "stages": [
            (0, "новая заявка", "Заявка создана"),
            (1, "принят на склад", "Груз принят на склад отправителя"),
            (3, "отправлен", "Груз отправлен по маршруту"),
            (5, "в пути", "Груз находится в пути"),
            (8, "на таможне", "Груз проходит таможенный этап"),
            (12, "прибыл в страну", "Груз прибыл в страну назначения"),
            (14, "прибыл в город", "Груз прибыл в город получателя"),
            (15, "готов к выдаче", "Груз готов к выдаче"),
        ],
    },
    "turkey_cis": {
        "name": "Турция → СНГ",
        "stages": [
            (0, "новая заявка", "Заявка создана"),
            (1, "принят на склад", "Груз принят на склад отправителя"),
            (2, "отправлен", "Груз отправлен по маршруту"),
            (4, "в пути", "Груз находится в пути"),
            (7, "на таможне", "Груз проходит таможенный этап"),
            (10, "прибыл в страну", "Груз прибыл в страну назначения"),
            (12, "прибыл в город", "Груз прибыл в город получателя"),
            (13, "готов к выдаче", "Груз готов к выдаче"),
        ],
    },
    "uae_cis": {
        "name": "Дубай/ОАЭ → СНГ",
        "stages": [
            (0, "новая заявка", "Заявка создана"),
            (1, "принят на склад", "Груз принят на склад отправителя"),
            (2, "отправлен", "Груз отправлен по маршруту"),
            (4, "в пути", "Груз находится в пути"),
            (6, "на таможне", "Груз проходит таможенный этап"),
            (9, "прибыл в страну", "Груз прибыл в страну назначения"),
            (11, "прибыл в город", "Груз прибыл в город получателя"),
            (12, "готов к выдаче", "Груз готов к выдаче"),
        ],
    },
    "cis_local": {
        "name": "СНГ → СНГ",
        "stages": [
            (0, "новая заявка", "Заявка создана"),
            (1, "принят на склад", "Груз принят на склад"),
            (2, "отправлен", "Груз отправлен"),
            (3, "в пути", "Груз находится в пути"),
            (5, "прибыл в город", "Груз прибыл в город получателя"),
            (6, "готов к выдаче", "Груз готов к выдаче"),
        ],
    },
}

TERMINAL_STATUSES = {"доставлен", "отменён", "проблема", "передан курьеру"}

SUPPORT_REPLY_TEMPLATES = {
    "check": "Здравствуйте! Мы проверим информацию по вашему обращению и вернёмся с ответом.",
    "route": "Здравствуйте! Ваш груз находится в пути. Как только статус изменится, бот автоматически отправит уведомление.",
    "payment": "Здравствуйте! Пожалуйста, проверьте оплату по вашему грузу. Если уже оплатили — отправьте подтверждение менеджеру.",
    "ready": "Здравствуйте! Ваш груз готов к выдаче. Можете связаться с менеджером для уточнения времени получения.",
}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
)
logger = logging.getLogger("cargo_ai_crm")
router = Router()


# =========================
# HELPERS
# =========================
def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def safe(text: object) -> str:
    return html.escape(str(text or ""))


def normalize(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").lower().strip())


def short_username(message: Message) -> str:
    user = message.from_user
    if not user:
        return ""
    return user.username or ""


def full_name(message: Message) -> str:
    user = message.from_user
    if not user:
        return ""
    return " ".join(filter(None, [user.first_name, user.last_name])).strip()


def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS


def is_warehouse(user_id: int) -> bool:
    return user_id in WAREHOUSE_IDS or is_admin(user_id)


def is_courier_static(user_id: int) -> bool:
    return user_id in COURIER_IDS or is_admin(user_id)


async def has_role(user_id: int, *roles: str) -> bool:
    if is_admin(user_id):
        return True
    user = await get_user(user_id)
    return bool(user and user["role"] in roles)


def partner_code(user_id: int) -> str:
    return f"P{abs(user_id)}"


def parse_tracking_code(text: str) -> Optional[str]:
    match = re.search(r"\bCG\d{6}\d{4,}\b", (text or "").upper())
    if match:
        return match.group(0)
    text = (text or "").strip().upper()
    if text.startswith("CG"):
        return text
    return None


def track_url(code: str | None = None) -> str:
    """Публичная ссылка для клиента.

    Клиенту не нужно вводить команды в Telegram: он получает готовую
    ссылку и открывает статус груза в браузере.
    """
    if code:
        return f"{PUBLIC_BASE_URL}/track/{str(code).strip().upper()}"
    return f"{PUBLIC_BASE_URL}/track"


def track_button(code: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="🔎 Отследить груз", url=track_url(code))],
        ]
    )


def payment_page_url(code: str) -> str:
    return f"{PUBLIC_BASE_URL}/demo-pay/{str(code).strip().upper()}"


def track_pay_button(code: str, debt: float = 0) -> InlineKeyboardMarkup:
    buttons = [[InlineKeyboardButton(text="🔎 Отследить груз", url=track_url(code))]]
    if DEMO_PAYMENT_ENABLED and debt > 0:
        buttons.append([InlineKeyboardButton(text="💳 Оплатить", url=payment_page_url(code))])
    return InlineKeyboardMarkup(inline_keyboard=buttons)


def provider_info(provider: str) -> dict:
    return DEMO_PAYMENT_PROVIDERS.get((provider or "").strip().lower(), DEMO_PAYMENT_PROVIDERS["kaspi"])


def status_notify_text(code: str, status: str = "", comment: str = "") -> str:
    """Короткое уведомление без раскрытия статуса в Telegram.

    Подробности специально вынесены на веб-трекер:
    статус, объяснение, следующий этап, задержка и история.
    Так сайт отслеживания становится главным источником информации,
    а Telegram не дублирует весь трекер сообщениями.
    """
    return (
        f"📦 Есть обновление по грузу <b>{safe(code)}</b>.\n\n"
        "Откройте трекер, чтобы посмотреть статус, следующий этап и историю."
    )


def client_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="📦 Оформить доставку"), KeyboardButton(text="🔎 Где мой груз?")],
            [KeyboardButton(text="💳 Оплатить доставку"), KeyboardButton(text="🛠️ Техподдержка")],
            [KeyboardButton(text="📋 Мои заказы"), KeyboardButton(text="🧮 Рассчитать доставку")],
            [KeyboardButton(text="🛒 Выкуп товара"), KeyboardButton(text="🏢 Оптовая доставка")],
            [KeyboardButton(text="⚠️ Жалоба / проблема"), KeyboardButton(text="❓ FAQ")],
        ],
        resize_keyboard=True,
        input_field_placeholder="Выберите действие или напишите номер груза CG...",
    )


def admin_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="🧭 Promise OS"), KeyboardButton(text="📥 Новые заявки")],
            [KeyboardButton(text="📦 Все грузы"), KeyboardButton(text="🔎 Найти груз")],
            [KeyboardButton(text="💳 Оплаты/долги"), KeyboardButton(text="🤖 Автостатусы")],
            [KeyboardButton(text="📦 Партии"), KeyboardButton(text="🚨 Проблемные грузы")],
            [KeyboardButton(text="🔁 Изменить статус"), KeyboardButton(text="🏭 Меню склада")],
            [KeyboardButton(text="🛠️ Техподдержка"), KeyboardButton(text="💬 Жалобы")],
            [KeyboardButton(text="👥 Клиенты"), KeyboardButton(text="📢 Рассылка")],
            [KeyboardButton(text="📤 Excel экспорт"), KeyboardButton(text="📥 Обновить из Excel")],
            [KeyboardButton(text="⚙️ Тарифы"), KeyboardButton(text="⚙️ Статусы")],
            [KeyboardButton(text="📄 PDF квитанция"), KeyboardButton(text="👥 Роли")],
            [KeyboardButton(text="📊 Отчёт за день"), KeyboardButton(text="👤 Клиентское меню")],
        ],
        resize_keyboard=True,
        input_field_placeholder="CargoPromise OS: выберите действие",
    )


def warehouse_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="➕ Принять груз"), KeyboardButton(text="⚖️ Указать вес")],
            [KeyboardButton(text="📸 Добавить фото"), KeyboardButton(text="🔁 Изменить статус")],
            [KeyboardButton(text="📦 Грузы на складе"), KeyboardButton(text="🔎 Найти груз")],
            [KeyboardButton(text="🚚 Курьерское меню"), KeyboardButton(text="👤 Клиентское меню")],
        ],
        resize_keyboard=True,
        input_field_placeholder="Склад: введите CG... или выберите действие",
    )


def courier_keyboard() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="🚚 Мои доставки"), KeyboardButton(text="✅ Отметить доставлено")],
            [KeyboardButton(text="🔁 Изменить статус"), KeyboardButton(text="🔎 Найти груз")],
            [KeyboardButton(text="👤 Клиентское меню")],
        ],
        resize_keyboard=True,
        input_field_placeholder="Курьер: выберите действие",
    )



def client_menu_text(name: str = "") -> str:
    greeting = f", {safe(name)}" if name else ""
    return (
        f"<b>🚚 {safe(COMPANY_NAME)}</b>\n"
        f"Здравствуйте{greeting}!\n\n"
        "<b>Что можно сделать:</b>\n"
        "📦 оформить доставку или оптовую партию\n"
        "🔎 проверить статус груза по номеру\n"
        "🧮 рассчитать примерную стоимость\n"
        "🛒 оставить заявку на выкуп товара\n"
        "⚠️ отправить жалобу или проблему\n"
        "🛠️ создать обращение в техподдержку\n\n"
        "Напишите номер груза вида <code>CG...</code> или выберите действие ниже."
    )


def admin_menu_text() -> str:
    return (
        f"<b>👑 Админ-меню {safe(COMPANY_NAME)}</b>\n\n"
        "<b>Работаем только через Telegram:</b>\n"
        "📥 заявки и грузы — обработка клиентов\n"
        "🤖 автостатусы — бот сам ведёт груз по маршруту и уведомляет клиента\n"
        "🔎 поиск груза — быстро найти заказ по номеру\n"
        "🔁 статусы — ручная смена этапов, если нужно поправить маршрут\n"
        "💰 финансы — цена, оплаты, долги, маржа\n"
        "💬 жалобы — ответы клиентам\n"
        "🛠️ техподдержка — тикеты, ответы и закрытие обращений\n"
        "🚚 курьеры — выдача и доставка по городу\n"
        "📦 партии — менять статус сразу группе грузов\n"
        "🧭 Promise OS — обещания клиенту, риск спора, тревожность и потери\n"
        "📤 Excel — только для отчётов и массовых обновлений\n\n"
        "Для работы на компьютере откройте обычный <b>Telegram Desktop</b> и пользуйтесь этим же меню."
    )


def warehouse_menu_text() -> str:
    return (
        f"<b>🏭 Склад {safe(COMPANY_NAME)}</b>\n\n"
        "Здесь удобно работать с телефона:\n"
        "➕ принять груз\n"
        "⚖️ указать фактический вес\n"
        "📸 добавить фото коробки\n"
        "🔁 поменять статус и уведомить клиента\n\n"
        "Можно просто отправить номер груза <code>CG...</code>."
    )


def courier_menu_text() -> str:
    return (
        f"<b>🚚 Курьерское меню</b>\n\n"
        "Здесь курьер видит свои доставки, адреса и может отметить груз доставленным."
    )


async def menu_for_user(user_id: int, name: str = "") -> tuple[str, ReplyKeyboardMarkup]:
    """Возвращает правильное стартовое меню по роли пользователя.

    Админ при /start видит админку, склад — складское меню,
    курьер — курьерское меню, обычный клиент — клиентское меню.
    """
    if is_admin(user_id):
        return admin_menu_text(), admin_keyboard()

    if is_warehouse(user_id) or await has_role(user_id, "warehouse"):
        return warehouse_menu_text(), warehouse_keyboard()

    if is_courier_static(user_id) or await has_role(user_id, "courier"):
        return courier_menu_text(), courier_keyboard()

    return client_menu_text(name), client_keyboard()


def status_keyboard(order_id: int, statuses: Optional[list[str]] = None) -> InlineKeyboardMarkup:
    rows = []
    for status in statuses or STATUSES:
        rows.append([InlineKeyboardButton(text=status, callback_data=f"st:{order_id}:{status}")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def order_actions_keyboard(order_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="🤖 Включить автостатусы", callback_data=f"auto_enable:{order_id}")],
            [InlineKeyboardButton(text="🔁 Изменить статус", callback_data=f"open_status:{order_id}")],
            [InlineKeyboardButton(text="✅ Доставлен", callback_data=f"st:{order_id}:доставлен")],
            [InlineKeyboardButton(text="⚠️ Проблема", callback_data=f"st:{order_id}:проблема")],
        ]
    )


def new_order_actions_keyboard(order_id: int) -> InlineKeyboardMarkup:
    """Кнопки для новой заявки до оплаты.

    Автостатусы специально не показываем, чтобы логика была:
    заявка → оплата → уведомление админу → админ включает автостатусы.
    """
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="🔁 Изменить статус", callback_data=f"open_status:{order_id}")],
            [InlineKeyboardButton(text="⚠️ Проблема", callback_data=f"st:{order_id}:проблема")],
        ]
    )


def auto_status_keyboard(order_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="✅ Включить автостатусы", callback_data=f"auto_enable:{order_id}")],
            [InlineKeyboardButton(text="⏸ Выключить автостатусы", callback_data=f"auto_disable:{order_id}")],
            [InlineKeyboardButton(text="🔁 Изменить статус вручную", callback_data=f"open_status:{order_id}")],
        ]
    )


def auto_status_panel_keyboard(orders: list[aiosqlite.Row]) -> InlineKeyboardMarkup:
    """Главная панель автостатусов без команд.

    Админ не должен писать /autostatus или /autoroute. Он нажимает кнопку,
    а бот сам включает автостатусы и сам выбирает маршрут по направлению груза.
    """
    rows = [
        [InlineKeyboardButton(text="✅ Включить всем активным грузам", callback_data="auto_enable_all")],
        [InlineKeyboardButton(text="📋 Грузы без автостатусов", callback_data="auto_show_disabled")],
    ]
    for order in orders[:10]:
        enabled = bool(int(order["auto_status_enabled"] or 0))
        marker = "✅" if enabled else "🤖"
        code = order["tracking_code"]
        from_country = order["from_country"] or "?"
        to_city = order["to_city"] or "?"
        rows.append([
            InlineKeyboardButton(
                text=f"{marker} {code} · {from_country} → {to_city}",
                callback_data=f"open_auto:{order['id']}",
            )
        ])
    rows.append([InlineKeyboardButton(text="↩️ Назад в админ-меню", callback_data="admin_back")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def complaint_actions_keyboard(complaint_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="✅ Закрыть жалобу", callback_data=f"close_complaint:{complaint_id}")],
        ]
    )


def support_client_menu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="🆕 Создать обращение", callback_data="support_new")],
            [InlineKeyboardButton(text="📋 Мои обращения", callback_data="support_my")],
        ]
    )


def support_topic_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="📦 Вопрос по грузу", callback_data="support_topic:cargo")],
            [InlineKeyboardButton(text="💰 Оплата / долг", callback_data="support_topic:payment")],
            [InlineKeyboardButton(text="📸 Фото / повреждение", callback_data="support_topic:damage")],
            [InlineKeyboardButton(text="⚙️ Ошибка в боте", callback_data="support_topic:bot")],
            [InlineKeyboardButton(text="❓ Другое", callback_data="support_topic:other")],
        ]
    )


def support_ticket_actions_keyboard(ticket_id: int, include_client_reply: bool = False) -> InlineKeyboardMarkup:
    if include_client_reply:
        return InlineKeyboardMarkup(
            inline_keyboard=[
                [InlineKeyboardButton(text="✍️ Ответить в обращение", callback_data=f"support_client_reply:{ticket_id}")],
            ]
        )
    return InlineKeyboardMarkup(
        inline_keyboard=[
            [InlineKeyboardButton(text="💬 Ответить", callback_data=f"support_reply:{ticket_id}")],
            [InlineKeyboardButton(text="✅ Проверим", callback_data=f"support_tpl:{ticket_id}:check"), InlineKeyboardButton(text="🚚 В пути", callback_data=f"support_tpl:{ticket_id}:route")],
            [InlineKeyboardButton(text="💰 Оплата", callback_data=f"support_tpl:{ticket_id}:payment"), InlineKeyboardButton(text="📦 Готов", callback_data=f"support_tpl:{ticket_id}:ready")],
            [InlineKeyboardButton(text="✅ Закрыть", callback_data=f"support_close:{ticket_id}")],
        ]
    )




def batch_panel_keyboard(batches: list[aiosqlite.Row]) -> InlineKeyboardMarkup:
    rows = [
        [InlineKeyboardButton(text="➕ Создать партию", callback_data="batch_create")],
        [InlineKeyboardButton(text="➕ Добавить грузы в партию", callback_data="batch_add_orders")],
        [InlineKeyboardButton(text="🔁 Сменить статус партии", callback_data="batch_change_status")],
        [InlineKeyboardButton(text="📋 Активные партии", callback_data="batch_list")],
    ]
    for b in batches[:8]:
        rows.append([InlineKeyboardButton(text=f"📦 #{b['id']} · {b['name']} · {b['status']}", callback_data=f"batch_open:{b['id']}")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def batch_actions_keyboard(batch_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="➕ Добавить грузы", callback_data=f"batch_add_to:{batch_id}")],
        [InlineKeyboardButton(text="🔁 Сменить статус партии", callback_data=f"batch_status_for:{batch_id}")],
        [InlineKeyboardButton(text="📋 Грузы в партии", callback_data=f"batch_orders:{batch_id}")],
    ])


def batch_status_keyboard(batch_id: int, statuses: list[str]) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=status, callback_data=f"batch_set_status:{batch_id}:{status}")] for status in statuses
    ])


def tariffs_panel_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="➕ Добавить/изменить тариф", callback_data="tariff_add")],
        [InlineKeyboardButton(text="📋 Показать тарифы", callback_data="tariff_list")],
    ])


def statuses_panel_keyboard(custom_statuses: list[aiosqlite.Row]) -> InlineKeyboardMarkup:
    rows = [
        [InlineKeyboardButton(text="➕ Добавить статус", callback_data="status_add")],
        [InlineKeyboardButton(text="📋 Показать все статусы", callback_data="status_list")],
    ]
    for st in custom_statuses[:10]:
        rows.append([InlineKeyboardButton(text=f"🗑 Удалить: {st['name']}", callback_data=f"status_delete:{st['id']}")])
    return InlineKeyboardMarkup(inline_keyboard=rows)


def debts_keyboard(rows: list[aiosqlite.Row]) -> InlineKeyboardMarkup | None:
    buttons = []
    for o in rows[:10]:
        buttons.append([InlineKeyboardButton(text=f"📩 Напомнить {o['tracking_code']}", callback_data=f"pay_remind:{o['id']}")])
    return InlineKeyboardMarkup(inline_keyboard=buttons) if buttons else None


def broadcast_audience_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="👥 Всем клиентам", callback_data="broadcast_audience:all")],
        [InlineKeyboardButton(text="📦 Клиентам с активными грузами", callback_data="broadcast_audience:active")],
        [InlineKeyboardButton(text="💸 Только должникам", callback_data="broadcast_audience:debtors")],
        [InlineKeyboardButton(text="🤝 Партнёрам", callback_data="broadcast_audience:partners")],
    ])

@asynccontextmanager
async def get_db():
    db = await aiosqlite.connect(DB_PATH)
    db.row_factory = aiosqlite.Row
    try:
        yield db
    finally:
        await db.close()




async def db_fetchone(db: aiosqlite.Connection, query: str, params: tuple = ()):
    cur = await db.execute(query, params)
    return await cur.fetchone()


async def db_fetchall(db: aiosqlite.Connection, query: str, params: tuple = ()) -> list[aiosqlite.Row]:
    cur = await db.execute(query, params)
    return await cur.fetchall()

async def init_db() -> None:
    async with get_db() as db:
        await db.executescript(
            """
            CREATE TABLE IF NOT EXISTS users (
                telegram_id INTEGER PRIMARY KEY,
                username TEXT,
                full_name TEXT,
                role TEXT DEFAULT 'client',
                partner_code TEXT,
                ref_partner_id INTEGER,
                created_at TEXT,
                updated_at TEXT
            );

            CREATE TABLE IF NOT EXISTS orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                tracking_code TEXT UNIQUE,
                order_type TEXT DEFAULT 'cargo',
                from_country TEXT,
                to_city TEXT,
                cargo_type TEXT,
                weight REAL,
                volume REAL,
                description TEXT,
                customer_name TEXT,
                phone TEXT,
                status TEXT DEFAULT 'новая заявка',
                price REAL DEFAULT 0,
                cost REAL DEFAULT 0,
                margin REAL DEFAULT 0,
                partner_id INTEGER,
                created_at TEXT,
                updated_at TEXT
            );

            CREATE TABLE IF NOT EXISTS status_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id INTEGER,
                status TEXT,
                comment TEXT,
                created_by INTEGER,
                created_at TEXT
            );

            CREATE TABLE IF NOT EXISTS complaints (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER,
                order_id INTEGER,
                tracking_code TEXT,
                text TEXT,
                urgency TEXT,
                status TEXT DEFAULT 'open',
                admin_reply TEXT,
                created_at TEXT,
                updated_at TEXT
            );


            CREATE TABLE IF NOT EXISTS support_tickets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                tracking_code TEXT,
                topic TEXT,
                status TEXT DEFAULT 'open',
                priority TEXT DEFAULT 'normal',
                last_message TEXT,
                created_at TEXT,
                updated_at TEXT,
                closed_at TEXT
            );

            CREATE TABLE IF NOT EXISTS support_messages (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ticket_id INTEGER NOT NULL,
                sender_id INTEGER,
                sender_role TEXT,
                text TEXT,
                created_at TEXT
            );

            CREATE TABLE IF NOT EXISTS partner_leads (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                partner_id INTEGER,
                client_id INTEGER,
                created_at TEXT,
                UNIQUE(partner_id, client_id)
            );

            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT
            );

            CREATE TABLE IF NOT EXISTS tariffs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                country_key TEXT UNIQUE,
                country_name TEXT,
                rate REAL NOT NULL,
                created_at TEXT,
                updated_at TEXT
            );

            CREATE TABLE IF NOT EXISTS cargo_photos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id INTEGER,
                file_id TEXT,
                comment TEXT,
                created_by INTEGER,
                created_at TEXT
            );

            CREATE TABLE IF NOT EXISTS payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id INTEGER,
                amount REAL,
                comment TEXT,
                created_by INTEGER,
                created_at TEXT
            );

            CREATE TABLE IF NOT EXISTS batches (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                route_key TEXT,
                status TEXT DEFAULT 'новая заявка',
                created_by INTEGER,
                created_at TEXT,
                updated_at TEXT
            );

            CREATE TABLE IF NOT EXISTS batch_orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                batch_id INTEGER,
                order_id INTEGER,
                created_at TEXT,
                UNIQUE(batch_id, order_id)
            );

            CREATE TABLE IF NOT EXISTS custom_statuses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE,
                created_at TEXT
            );

            CREATE TABLE IF NOT EXISTS track_views (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tracking_code TEXT,
                ip TEXT,
                user_agent TEXT,
                created_at TEXT
            );

            CREATE TABLE IF NOT EXISTS demo_payment_requests (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                payment_id TEXT UNIQUE,
                order_id INTEGER,
                tracking_code TEXT,
                amount REAL,
                status TEXT DEFAULT 'created',
                provider TEXT,
                created_at TEXT,
                paid_at TEXT
            );
            """
        )
        for sql in [
            "ALTER TABLE orders ADD COLUMN paid_amount REAL DEFAULT 0",
            "ALTER TABLE orders ADD COLUMN payment_status TEXT DEFAULT 'не оплачено'",
            "ALTER TABLE orders ADD COLUMN courier_id INTEGER",
            "ALTER TABLE orders ADD COLUMN delivery_address TEXT",
            "ALTER TABLE orders ADD COLUMN auto_status_enabled INTEGER DEFAULT 0",
            "ALTER TABLE orders ADD COLUMN auto_status_route TEXT",
            "ALTER TABLE orders ADD COLUMN auto_status_started_at TEXT",
            "ALTER TABLE orders ADD COLUMN auto_status_last_at TEXT",
        ]:
            try:
                await db.execute(sql)
            except Exception:
                pass
        for country, rate in DEFAULT_RATES.items():
            # В тарифы показываем русские названия, английские ключи оставляем только для распознавания.
            if re.search(r"[a-z]", country):
                continue
            await db.execute(
                "INSERT OR IGNORE INTO tariffs (country_key, country_name, rate, created_at, updated_at) VALUES (?, ?, ?, ?, ?)",
                (normalize(country), country.title(), rate, now_iso(), now_iso()),
            )
        await db.commit()


async def upsert_user(message: Message, ref_partner_id: Optional[int] = None) -> None:
    if not message.from_user:
        return
    user_id = message.from_user.id
    username = short_username(message)
    name = full_name(message)
    role = "admin" if is_admin(user_id) else "warehouse" if user_id in WAREHOUSE_IDS else "client"
    code = partner_code(user_id)
    async with get_db() as db:
        existing = await db_fetchone(db, 
            "SELECT telegram_id, ref_partner_id, role FROM users WHERE telegram_id=?",
            (user_id,),
        )
        if existing:
            current_ref = existing["ref_partner_id"]
            saved_role = existing["role"] or role
            if is_admin(user_id):
                saved_role = "admin"
            elif user_id in WAREHOUSE_IDS:
                saved_role = "warehouse"
            await db.execute(
                """
                UPDATE users
                SET username=?, full_name=?, role=?, partner_code=?,
                    ref_partner_id=COALESCE(ref_partner_id, ?), updated_at=?
                WHERE telegram_id=?
                """,
                (username, name, saved_role, code, ref_partner_id, now_iso(), user_id),
            )
            final_ref = current_ref or ref_partner_id
        else:
            await db.execute(
                """
                INSERT INTO users (telegram_id, username, full_name, role, partner_code, ref_partner_id, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (user_id, username, name, role, code, ref_partner_id, now_iso(), now_iso()),
            )
            final_ref = ref_partner_id
        if final_ref and final_ref != user_id:
            await db.execute(
                "INSERT OR IGNORE INTO partner_leads (partner_id, client_id, created_at) VALUES (?, ?, ?)",
                (final_ref, user_id, now_iso()),
            )
        await db.commit()


async def get_user(user_id: int) -> Optional[aiosqlite.Row]:
    async with get_db() as db:
        return await db_fetchone(db, "SELECT * FROM users WHERE telegram_id=?", (user_id,))


async def get_order_by_code(code: str) -> Optional[aiosqlite.Row]:
    async with get_db() as db:
        return await db_fetchone(db, "SELECT * FROM orders WHERE UPPER(tracking_code)=UPPER(?)", (code,))


async def get_order(order_id: int) -> Optional[aiosqlite.Row]:
    async with get_db() as db:
        return await db_fetchone(db, "SELECT * FROM orders WHERE id=?", (order_id,))


async def order_payment_confirmed(order: aiosqlite.Row) -> bool:
    """Можно ли включать автостатусы.

    В новой логике автостатусы запускаются только после оплаты.
    Для демо достаточно paid_amount > 0 или payment_status = оплачено/частично.
    """
    if not AUTO_STATUS_AFTER_PAYMENT_ONLY:
        return True
    paid = float(order["paid_amount"] or 0)
    status = normalize(order["payment_status"] or "")
    return paid > 0 or status in {"оплачено", "частично"}


def choose_auto_route(from_country: str, to_city: str = "") -> str:
    text = normalize(f"{from_country} {to_city}")
    if any(x in text for x in ["китай", "china", "guangzhou", "гуанчжоу", "иу", "yiwu", "1688", "taobao"]):
        return "china_cis"
    if any(x in text for x in ["турция", "turkey", "стамбул", "istanbul"]):
        return "turkey_cis"
    if any(x in text for x in ["дубай", "оаэ", "uae", "dubai", "emirates"]):
        return "uae_cis"
    return "cis_local"


def auto_route_label(route_key: str) -> str:
    route = AUTO_STATUS_ROUTES.get(route_key) or AUTO_STATUS_ROUTES["cis_local"]
    return route["name"]


STATUS_EXPLANATIONS = {
    "новая заявка": "Заявка принята. Груз ещё не прошёл складскую обработку.",
    "принят на склад": "Груз принят на склад отправителя. Следующий этап — отправка по маршруту.",
    "ожидает отправки": "Груз готовится к отправке. Обычно он ждёт формирования партии.",
    "отправлен": "Груз отправлен по маршруту. Обновление появится после следующего этапа движения.",
    "в пути": "Груз находится между складами или городами. На этом этапе статус может обновляться не каждый день.",
    "на таможне": "Груз проходит таможенный этап. Возможны задержки из-за проверки документов или партии.",
    "прибыл в страну": "Груз прибыл в страну назначения и ожидает дальнейшую сортировку.",
    "прибыл в Казахстан": "Груз прибыл в Казахстан и ожидает сортировку или доставку по городу.",
    "прибыл в город": "Груз прибыл в город получателя и готовится к выдаче.",
    "готов к выдаче": "Груз готов к получению. Можно связаться с менеджером для уточнения выдачи.",
    "передан курьеру": "Груз передан курьеру для доставки.",
    "доставлен": "Груз доставлен клиенту.",
    "проблема": "По грузу нужна проверка менеджера. Возможна задержка, уточнение оплаты или данных.",
    "отменён": "Заявка отменена.",
}


def status_explanation(status: str) -> str:
    return STATUS_EXPLANATIONS.get(normalize(status), "Статус обновлён. Следующее изменение появится после обработки груза.")


def _route_stage_index(route_key: str, status: str) -> int:
    route = AUTO_STATUS_ROUTES.get(route_key) or AUTO_STATUS_ROUTES["cis_local"]
    stage_statuses = [s for _, s, _ in route["stages"]]
    status = normalize(status)
    return stage_statuses.index(status) if status in stage_statuses else -1


def _stage_due_datetime(started_at: datetime, stage_value: int) -> datetime:
    if AUTO_STATUS_TIME_UNIT == "minutes":
        return started_at + timedelta(minutes=stage_value)
    return started_at + timedelta(days=stage_value)


def trustflow_info(order: aiosqlite.Row) -> dict:
    """Клиентская логика Anti-'Где мой груз?'.

    Показывает не только статус, а:
    - что этот статус значит;
    - следующий этап;
    - когда ждать обновление;
    - нужно ли писать менеджеру.
    """
    status = order["status"] or "новая заявка"
    route_key = order["auto_status_route"] or choose_auto_route(order["from_country"] or "", order["to_city"] or "")
    route = AUTO_STATUS_ROUTES.get(route_key) or AUTO_STATUS_ROUTES["cis_local"]
    started_at = parse_iso_datetime(order["auto_status_started_at"] or order["created_at"] or now_iso())

    current_index = _route_stage_index(route_key, status)
    next_stage = None
    next_due = None
    if 0 <= current_index < len(route["stages"]) - 1:
        day, next_status, next_comment = route["stages"][current_index + 1]
        next_due = _stage_due_datetime(started_at, day)
        next_stage = {
            "status": next_status,
            "comment": next_comment,
            "due_at": next_due,
        }

    now = datetime.now(timezone.utc)
    normalized_status = normalize(status)

    if normalized_status in {"проблема", "отменён"}:
        level = "red"
        label = "Нужна проверка менеджером"
        client_hint = "По грузу есть вопрос. Лучше связаться с менеджером."
        show_support = True
    elif next_due and now > next_due + timedelta(hours=TRUSTFLOW_GRACE_HOURS):
        level = "yellow"
        label = "Возможна задержка"
        client_hint = "Ожидаемый этап уже должен был обновиться. Менеджеру стоит проверить груз."
        show_support = True
    else:
        level = "green"
        label = "Всё идёт по плану"
        client_hint = "Писать менеджеру не нужно: следующее обновление появится автоматически."
        show_support = False

    if normalized_status in {"готов к выдаче", "доставлен", "передан курьеру"}:
        level = "green"
        label = "Груз на финальном этапе"
        client_hint = "Можно уточнить получение или доставку у менеджера."
        show_support = normalized_status != "доставлен"

    due_text = ""
    if next_stage and next_due:
        due_text = next_due.strftime("%d.%m.%Y %H:%M") if AUTO_STATUS_TIME_UNIT == "minutes" else next_due.strftime("%d.%m.%Y")

    return {
        "level": level,
        "label": label,
        "status_explanation": status_explanation(status),
        "next_stage": next_stage["status"] if next_stage else "финальный этап",
        "next_comment": next_stage["comment"] if next_stage else "Груз находится на завершающем этапе.",
        "next_due_text": due_text,
        "client_hint": client_hint,
        "show_support": show_support,
        "route_name": route["name"],
    }


async def record_track_view(code: str, request: Optional[web.Request] = None) -> None:
    ip = ""
    user_agent = ""
    if request:
        ip = request.headers.get("X-Forwarded-For", request.remote or "")
        user_agent = request.headers.get("User-Agent", "")[:300]
    async with get_db() as db:
        await db.execute(
            "INSERT INTO track_views (tracking_code, ip, user_agent, created_at) VALUES (?, ?, ?, ?)",
            (code.upper(), ip, user_agent, now_iso()),
        )
        await db.commit()


async def track_view_stats(code: Optional[str] = None) -> dict:
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    async with get_db() as db:
        if code:
            total = await db_fetchone(db, "SELECT COUNT(*) AS cnt FROM track_views WHERE UPPER(tracking_code)=UPPER(?)", (code,))
            today_row = await db_fetchone(db, "SELECT COUNT(*) AS cnt FROM track_views WHERE UPPER(tracking_code)=UPPER(?) AND created_at LIKE ?", (code, f"{today}%",))
        else:
            total = await db_fetchone(db, "SELECT COUNT(*) AS cnt FROM track_views")
            today_row = await db_fetchone(db, "SELECT COUNT(*) AS cnt FROM track_views WHERE created_at LIKE ?", (f"{today}%",))
    return {"total": total["cnt"] if total else 0, "today": today_row["cnt"] if today_row else 0}


def promise_route_window(order: aiosqlite.Row) -> tuple[str, str, str]:
    route_key = order["auto_status_route"] or choose_auto_route(order["from_country"] or "", order["to_city"] or "")
    if route_key == "china_cis":
        return "12–16 дней", "после оплаты и принятия на склад", "Не обещать точную дату прибытия до формирования партии."
    if route_key == "turkey_cis":
        return "7–12 дней", "после оплаты и принятия на склад", "Предупредить, что сроки зависят от партии и таможни."
    if route_key == "uae_cis":
        return "7–12 дней", "после оплаты и принятия на склад", "Предупредить про возможную проверку товара."
    return "2–5 дней", "после оплаты и принятия на склад", "Не обещать доставку день-в-день без подтверждения склада."


def row_has_open_items(rows: list[aiosqlite.Row]) -> bool:
    for r in rows:
        try:
            if str(r["status"]).lower() != "closed":
                return True
        except Exception:
            return True
    return False


def promise_os_profile(order: aiosqlite.Row, history: list[aiosqlite.Row], photos: list[aiosqlite.Row], tickets: list[aiosqlite.Row], views: dict) -> dict:
    score = 100
    risk_reasons = []
    proof_points = []

    paid = float(order["paid_amount"] or 0)
    weight = float(order["weight"] or 0)
    status = normalize(order["status"] or "")

    if paid <= 0:
        score -= 22
        risk_reasons.append("оплата ещё не подтверждена")
    else:
        proof_points.append("оплата подтверждена")

    if weight <= 0:
        score -= 12
        risk_reasons.append("не указан фактический вес")
    else:
        proof_points.append("вес указан")

    if not photos:
        score -= 10
        risk_reasons.append("нет фото со склада")
    else:
        proof_points.append("фото груза добавлено")

    if not int(order["auto_status_enabled"] or 0):
        score -= 10
        risk_reasons.append("автостатусы ещё не включены")
    else:
        proof_points.append("автостатусы включены")

    updated_at = parse_iso_datetime(order["updated_at"] or order["created_at"] or now_iso())
    stale_hours = (datetime.now(timezone.utc) - updated_at).total_seconds() / 3600
    if status not in {"доставлен", "отменён"} and stale_hours > PROMISE_STALE_HOURS:
        score -= 18
        risk_reasons.append(f"нет обновления больше {PROMISE_STALE_HOURS} часов")

    if status in {"проблема", "отменён"}:
        score -= 35
        risk_reasons.append("груз помечен как проблемный")

    if row_has_open_items(tickets):
        score -= 12
        risk_reasons.append("есть открытое обращение клиента")

    if views.get("today", 0) >= PROMISE_HIGH_TRACK_VIEWS:
        score -= 8
        risk_reasons.append("клиент часто проверяет трекер сегодня")

    score = max(0, min(100, score))

    anxiety = 0
    anxiety += min(40, int(views.get("today", 0)) * 8)
    anxiety += 25 if stale_hours > PROMISE_STALE_HOURS else 0
    anxiety += 15 if not photos else 0
    anxiety += 15 if paid <= 0 else 0
    anxiety += 20 if status in {"проблема", "отменён"} else 0
    anxiety = max(0, min(100, anxiety))

    if score >= 80:
        gate = "✅ Можно обещать"
        level = "green"
        client_label = "Всё по плану"
    elif score >= 55:
        gate = "⚠️ Обещать осторожно"
        level = "yellow"
        client_label = "Есть риск задержки или вопроса"
    else:
        gate = "🔴 Нужна проверка"
        level = "red"
        client_label = "Нужна проверка менеджером"

    safe_window, promise_start, warning = promise_route_window(order)
    if not risk_reasons:
        risk_reasons.append("критичных рисков не найдено")
    if not proof_points:
        proof_points.append("доказательства ещё собираются")

    return {
        "score": score,
        "anxiety": anxiety,
        "gate": gate,
        "level": level,
        "client_label": client_label,
        "safe_window": safe_window,
        "promise_start": promise_start,
        "warning": warning,
        "risk_reasons": risk_reasons[:6],
        "proof_points": proof_points[:6],
        "stale_hours": round(stale_hours, 1),
    }


def promise_os_html(order: aiosqlite.Row, profile: dict) -> str:
    risks = "".join(f"<li>{html.escape(str(x))}</li>" for x in profile["risk_reasons"])
    proofs = "".join(f"<li>{html.escape(str(x))}</li>" for x in profile["proof_points"])
    emoji = "🟢" if profile["level"] == "green" else "🟡" if profile["level"] == "yellow" else "🔴"
    return f"""
      <div class="promise {html.escape(profile['level'])}">
        <div class="promise-head">
          <div>
            <div class="promise-kicker">CargoPromise OS</div>
            <div class="promise-title">{emoji} {html.escape(profile['client_label'])}</div>
          </div>
          <div class="score">{profile['score']}<span>/100</span></div>
        </div>
        <div class="grid" style="margin-top:14px;">
          <div class="item"><div class="label">Безопасное обещание</div><div class="value">{html.escape(profile['safe_window'])}</div></div>
          <div class="item"><div class="label">Срок считать</div><div class="value">{html.escape(profile['promise_start'])}</div></div>
          <div class="item"><div class="label">Anxiety Score</div><div class="value">{profile['anxiety']}/100</div></div>
          <div class="item"><div class="label">Risk Gate</div><div class="value">{html.escape(profile['gate'])}</div></div>
        </div>
        <div class="promise-two">
          <div>
            <b>Риски обещания</b>
            <ul>{risks}</ul>
          </div>
          <div>
            <b>Доказательства по грузу</b>
            <ul>{proofs}</ul>
          </div>
        </div>
        <div class="hint"><b>Что нельзя обещать:</b> {html.escape(profile['warning'])}</div>
      </div>
    """


async def promise_os_leak_map(limit: int = 10) -> dict:
    radar = await smartflow_problem_radar(limit)
    views = await track_view_stats()
    async with get_db() as db:
        risky = await db_fetchall(
            db,
            """
            SELECT * FROM orders
            WHERE status NOT IN ('доставлен', 'отменён')
            ORDER BY updated_at DESC
            LIMIT ?
            """,
            (limit,),
        )
        debt_row = await db_fetchone(
            db,
            "SELECT COALESCE(SUM(COALESCE(price,0)-COALESCE(paid_amount,0)),0) AS debt FROM orders WHERE COALESCE(price,0)>COALESCE(paid_amount,0)",
        )
    top_profiles = []
    for order in risky:
        data = await smart_cargo_card_data(order["id"])
        v = await track_view_stats(order["tracking_code"])
        p = promise_os_profile(order, data.get("history", []), data.get("photos", []), data.get("tickets", []), v)
        if p["score"] < 80 or p["anxiety"] >= 40:
            top_profiles.append((order, p))
    top_profiles.sort(key=lambda item: (item[1]["anxiety"], -item[1]["score"]), reverse=True)
    return {
        "radar": radar,
        "views": views,
        "debt": float(debt_row["debt"] or 0) if debt_row else 0,
        "top_profiles": top_profiles[:limit],
    }


def format_promise_os_report(data: dict) -> str:
    radar = data["radar"]
    top = data["top_profiles"]
    lines = [
        "<b>🧭 CargoPromise OS</b>",
        "",
        "Карта обещаний, рисков и потерь по грузам.",
        "",
        f"Самопроверок трекера сегодня: <b>{data['views']['today']}</b>",
        f"Долг / риск неоплаты: <b>{data['debt']:g} {safe(CURRENCY)}</b>",
        f"Грузов без обновления: <b>{len(radar['stale'])}</b>",
        f"Грузов без веса: <b>{len(radar['no_weight'])}</b>",
        f"Открытых жалоб: <b>{len(radar['open_complaints'])}</b>",
        f"Открытых тикетов: <b>{len(radar['open_tickets'])}</b>",
        "",
        "<b>Грузы, где может начаться вопрос/спор:</b>",
    ]
    if not top:
        lines.append("Критичных рисков пока нет.")
    else:
        for order, p in top[:8]:
            lines.append(f"— <b>{safe(order['tracking_code'])}</b>: TrustScore {p['score']}/100, Anxiety {p['anxiety']}/100 · {safe(p['gate'])}")
    lines.append("")
    lines.append("Смысл: не ждать жалоб, а видеть риск до того, как клиент напишет «где мой груз?»")
    return "\n".join(lines)


async def enable_auto_status_for_order(order_id: int, actor_id: int = 0) -> Optional[aiosqlite.Row]:
    """Включает бесплатные автостатусы одной кнопкой.

    Маршрут выбирается автоматически по направлению груза, поэтому админу
    не нужно вводить команды /autoroute и выбирать шаблон вручную.
    """
    order = await get_order(order_id)
    if not order:
        return None
    if not await order_payment_confirmed(order):
        return None
    route_key = choose_auto_route(order["from_country"] or "", order["to_city"] or "")
    async with get_db() as db:
        await db.execute(
            """
            UPDATE orders
            SET auto_status_enabled=1,
                auto_status_route=?,
                auto_status_started_at=?,
                auto_status_last_at=NULL,
                updated_at=?
            WHERE id=?
            """,
            (route_key, now_iso(), now_iso(), order_id),
        )
        await db.execute(
            "INSERT INTO status_history (order_id, status, comment, created_by, created_at) VALUES (?, ?, ?, ?, ?)",
            (order_id, order["status"] or "новая заявка", f"Автостатусы включены. Маршрут: {auto_route_label(route_key)}", actor_id, now_iso()),
        )
        await db.commit()
    return await get_order(order_id)


async def enable_auto_status_for_active_orders(actor_id: int = 0, limit: int = 500) -> int:
    """Включает автостатусы сразу для всех активных грузов.

    Это нужно для реальной карго-работы: админ не включает статусы каждому
    клиенту отдельно. Бот сам подбирает маршрут по направлению каждого груза.
    """
    async with get_db() as db:
        rows = await db_fetchall(
            db,
            """
            SELECT id FROM orders
            WHERE COALESCE(auto_status_enabled, 0)=0
              AND (COALESCE(paid_amount, 0) > 0 OR payment_status IN ('оплачено', 'частично'))
              AND status NOT IN ('доставлен', 'отменён', 'проблема', 'передан курьеру')
            ORDER BY id DESC
            LIMIT ?
            """,
            (limit,),
        )
    count = 0
    for row in rows:
        updated = await enable_auto_status_for_order(row["id"], actor_id)
        if updated:
            count += 1
    return count


async def list_orders_without_auto_status(limit: int = 10) -> list[aiosqlite.Row]:
    async with get_db() as db:
        return await db_fetchall(
            db,
            """
            SELECT * FROM orders
            WHERE COALESCE(auto_status_enabled, 0)=0
              AND (COALESCE(paid_amount, 0) > 0 OR payment_status IN ('оплачено', 'частично'))
              AND status NOT IN ('доставлен', 'отменён', 'проблема', 'передан курьеру')
            ORDER BY id DESC
            LIMIT ?
            """,
            (limit,),
        )



async def get_custom_statuses() -> list[aiosqlite.Row]:
    async with get_db() as db:
        return await db_fetchall(db, "SELECT * FROM custom_statuses ORDER BY id ASC")


async def get_all_statuses() -> list[str]:
    rows = await get_custom_statuses()
    statuses = list(STATUSES)
    for row in rows:
        name = (row["name"] or "").strip()
        if name and name not in statuses:
            statuses.append(name)
    return statuses


async def add_custom_status(name: str) -> bool:
    name = normalize(name)
    if not name:
        return False
    async with get_db() as db:
        await db.execute("INSERT OR IGNORE INTO custom_statuses (name, created_at) VALUES (?, ?)", (name, now_iso()))
        await db.commit()
    return True


async def delete_custom_status(status_id: int) -> None:
    async with get_db() as db:
        await db.execute("DELETE FROM custom_statuses WHERE id=?", (status_id,))
        await db.commit()


async def create_batch(name: str, route_key: str, created_by: int) -> aiosqlite.Row:
    route_key = route_key if route_key in AUTO_STATUS_ROUTES else "cis_local"
    async with get_db() as db:
        cur = await db.execute(
            "INSERT INTO batches (name, route_key, status, created_by, created_at, updated_at) VALUES (?, ?, 'новая заявка', ?, ?, ?)",
            (name.strip(), route_key, created_by, now_iso(), now_iso()),
        )
        batch_id = cur.lastrowid
        await db.commit()
    return await get_batch(batch_id)


async def get_batch(batch_id: int) -> Optional[aiosqlite.Row]:
    async with get_db() as db:
        return await db_fetchone(db, "SELECT * FROM batches WHERE id=?", (batch_id,))


async def list_batches(limit: int = 10) -> list[aiosqlite.Row]:
    async with get_db() as db:
        return await db_fetchall(db, "SELECT * FROM batches ORDER BY id DESC LIMIT ?", (limit,))


async def add_orders_to_batch(batch_id: int, codes_text: str) -> tuple[int, int]:
    codes = re.findall(r"CG\d{6}\d{4,}", (codes_text or "").upper())
    if not codes:
        codes = [x.strip().upper() for x in re.split(r"[,\n\s]+", codes_text or "") if x.strip().upper().startswith("CG")]
    ok = 0
    fail = 0
    async with get_db() as db:
        for code in dict.fromkeys(codes):
            order = await db_fetchone(db, "SELECT id FROM orders WHERE UPPER(tracking_code)=UPPER(?)", (code,))
            if not order:
                fail += 1
                continue
            await db.execute("INSERT OR IGNORE INTO batch_orders (batch_id, order_id, created_at) VALUES (?, ?, ?)", (batch_id, order["id"], now_iso()))
            ok += 1
        await db.commit()
    return ok, fail


async def get_batch_orders(batch_id: int) -> list[aiosqlite.Row]:
    async with get_db() as db:
        return await db_fetchall(
            db,
            """
            SELECT o.* FROM orders o
            JOIN batch_orders bo ON bo.order_id=o.id
            WHERE bo.batch_id=?
            ORDER BY o.id DESC
            """,
            (batch_id,),
        )


async def update_batch_status(batch_id: int, status: str, actor_id: int) -> tuple[Optional[aiosqlite.Row], list[aiosqlite.Row]]:
    batch = await get_batch(batch_id)
    if not batch:
        return None, []
    orders = await get_batch_orders(batch_id)
    updated_orders = []
    async with get_db() as db:
        await db.execute("UPDATE batches SET status=?, updated_at=? WHERE id=?", (status, now_iso(), batch_id))
        await db.commit()
    for order in orders:
        updated = await update_order_status(order["id"], status, actor_id, f"Статус изменён по партии #{batch_id}")
        if updated:
            updated_orders.append(updated)
    return await get_batch(batch_id), updated_orders


def format_batch(batch: aiosqlite.Row, count: int = 0) -> str:
    return (
        f"<b>📦 Партия #{batch['id']}</b>\n"
        f"Название: {safe(batch['name'])}\n"
        f"Маршрут: {safe(auto_route_label(batch['route_key'] or 'cis_local'))}\n"
        f"Статус: <b>{safe(batch['status'])}</b>\n"
        f"Грузов: {count}\n"
        f"Создана: {safe((batch['created_at'] or '')[:16].replace('T', ' '))}"
    )


async def list_clients(limit: int = 10, query: str = "") -> list[aiosqlite.Row]:
    async with get_db() as db:
        if query:
            like = f"%{query.lower()}%"
            return await db_fetchall(
                db,
                """
                SELECT u.telegram_id, u.username, u.full_name, u.role, COUNT(o.id) AS orders_count,
                       COALESCE(SUM(o.price),0) AS revenue,
                       COALESCE(SUM(o.price - COALESCE(o.paid_amount,0)),0) AS debt
                FROM users u
                LEFT JOIN orders o ON o.user_id=u.telegram_id
                WHERE u.role IN ('client','partner')
                  AND (LOWER(COALESCE(u.username,'')) LIKE ? OR LOWER(COALESCE(u.full_name,'')) LIKE ? OR CAST(u.telegram_id AS TEXT) LIKE ?)
                GROUP BY u.telegram_id
                ORDER BY orders_count DESC, u.updated_at DESC
                LIMIT ?
                """,
                (like, like, like, limit),
            )
        return await db_fetchall(
            db,
            """
            SELECT u.telegram_id, u.username, u.full_name, u.role, COUNT(o.id) AS orders_count,
                   COALESCE(SUM(o.price),0) AS revenue,
                   COALESCE(SUM(o.price - COALESCE(o.paid_amount,0)),0) AS debt
            FROM users u
            LEFT JOIN orders o ON o.user_id=u.telegram_id
            WHERE u.role IN ('client','partner')
            GROUP BY u.telegram_id
            ORDER BY orders_count DESC, u.updated_at DESC
            LIMIT ?
            """,
            (limit,),
        )


def format_client_row(row: aiosqlite.Row) -> str:
    name = row['full_name'] or ''
    username = f"@{row['username']}" if row['username'] else ''
    debt = max(float(row['debt'] or 0), 0)
    return f"<b>{safe(name or username or row['telegram_id'])}</b> · ID <code>{row['telegram_id']}</code>\nЗаказов: {row['orders_count']} · Долг: {debt} {safe(CURRENCY)}"

async def smart_cargo_card_data(order_id: int) -> dict:
    """Данные для умной карточки груза: история, оплаты, фото, тикеты."""
    async with get_db() as db:
        order = await db_fetchone(db, "SELECT * FROM orders WHERE id=?", (order_id,))
        if not order:
            return {}
        history = await db_fetchall(db, "SELECT * FROM status_history WHERE order_id=? ORDER BY id ASC", (order_id,))
        photos = await db_fetchall(db, "SELECT * FROM cargo_photos WHERE order_id=? ORDER BY id DESC LIMIT 5", (order_id,))
        payments = await db_fetchall(db, "SELECT * FROM payments WHERE order_id=? ORDER BY id DESC LIMIT 10", (order_id,))
        tickets = await db_fetchall(
            db,
            "SELECT * FROM support_tickets WHERE tracking_code=? ORDER BY id DESC LIMIT 5",
            (order["tracking_code"],),
        )
    return {"order": order, "history": history, "photos": photos, "payments": payments, "tickets": tickets}


async def smartflow_problem_radar(limit: int = 20) -> dict:
    """Показывает проблемные места без ручного поиска по CRM.

    Это главная уникальная логика: система сама подсвечивает, где нужны действия.
    """
    stale_cutoff = (datetime.now(timezone.utc) - timedelta(hours=ISSUE_STALE_HOURS)).isoformat(timespec="seconds")
    async with get_db() as db:
        stale = await db_fetchall(
            db,
            """
            SELECT * FROM orders
            WHERE status NOT IN ('доставлен', 'отменён')
              AND COALESCE(updated_at, created_at) < ?
            ORDER BY updated_at ASC
            LIMIT ?
            """,
            (stale_cutoff, limit),
        )
        no_weight = await db_fetchall(
            db,
            """
            SELECT * FROM orders
            WHERE status NOT IN ('доставлен', 'отменён')
              AND COALESCE(weight, 0) <= 0
            ORDER BY id DESC
            LIMIT ?
            """,
            (limit,),
        )
        debtors = await db_fetchall(
            db,
            """
            SELECT *, (COALESCE(price, 0) - COALESCE(paid_amount, 0)) AS debt
            FROM orders
            WHERE COALESCE(price, 0) > COALESCE(paid_amount, 0)
            ORDER BY debt DESC
            LIMIT ?
            """,
            (limit,),
        )
        open_complaints = await db_fetchall(
            db,
            "SELECT * FROM complaints WHERE status='open' ORDER BY id DESC LIMIT ?",
            (limit,),
        )
        open_tickets = await db_fetchall(
            db,
            "SELECT * FROM support_tickets WHERE status!='closed' ORDER BY updated_at DESC LIMIT ?",
            (limit,),
        )
        active_total = await db_fetchone(
            db,
            "SELECT COUNT(*) AS cnt FROM orders WHERE status NOT IN ('доставлен', 'отменён')",
        )
    return {
        "stale": stale,
        "no_weight": no_weight,
        "debtors": debtors,
        "open_complaints": open_complaints,
        "open_tickets": open_tickets,
        "active_total": active_total["cnt"] if active_total else 0,
    }


def _problem_codes(rows: list[aiosqlite.Row], max_items: int = 5) -> str:
    if not rows:
        return "нет"
    codes = []
    for r in rows[:max_items]:
        code = r["tracking_code"] if "tracking_code" in r.keys() else f"#{r['id']}"
        codes.append(str(code))
    extra = f" +{len(rows) - max_items}" if len(rows) > max_items else ""
    return ", ".join(codes) + extra


def format_problem_radar(radar: dict) -> str:
    return (
        "<b>🚨 SmartFlow: проблемные грузы</b>\n\n"
        f"Активных грузов: <b>{radar.get('active_total', 0)}</b>\n\n"
        f"⏳ Давно без обновления: <b>{len(radar['stale'])}</b>\n"
        f"{safe(_problem_codes(radar['stale']))}\n\n"
        f"⚖️ Без веса: <b>{len(radar['no_weight'])}</b>\n"
        f"{safe(_problem_codes(radar['no_weight']))}\n\n"
        f"💸 С долгом: <b>{len(radar['debtors'])}</b>\n"
        f"{safe(_problem_codes(radar['debtors']))}\n\n"
        f"⚠️ Открытых жалоб: <b>{len(radar['open_complaints'])}</b>\n"
        f"🛠️ Открытых обращений: <b>{len(radar['open_tickets'])}</b>\n\n"
        "Это не просто список заявок: система сама показывает, где менеджеру нужно действие."
    )


async def owner_smartflow_report_text() -> str:
    finance = await finance_report()
    radar = await smartflow_problem_radar(10)
    views = await track_view_stats()
    totals = finance["totals"]
    today = finance["today"]
    return (
        "<b>📊 Ежедневный SmartFlow-отчёт владельцу</b>\n\n"
        f"Сегодня заявок: <b>{today['today_orders']}</b>\n"
        f"Сегодня оборот: <b>{today['today_revenue']} {safe(CURRENCY)}</b>\n"
        f"Сегодня маржа: <b>{today['today_margin']} {safe(CURRENCY)}</b>\n"
        f"Клиенты сами проверили грузы сегодня: <b>{views['today']}</b>\n\n"
        f"Всего заявок: <b>{totals['total_orders']}</b>\n"
        f"Всего оборот: <b>{totals['revenue']} {safe(CURRENCY)}</b>\n"
        f"Всего маржа: <b>{totals['margin']} {safe(CURRENCY)}</b>\n\n"
        f"🚨 Без обновления: <b>{len(radar['stale'])}</b>\n"
        f"⚖️ Без веса: <b>{len(radar['no_weight'])}</b>\n"
        f"💸 С долгом: <b>{len(radar['debtors'])}</b>\n"
        f"⚠️ Жалобы: <b>{len(radar['open_complaints'])}</b>\n"
        f"🛠️ Тикеты: <b>{len(radar['open_tickets'])}</b>\n\n"
        "Владелец видит картину сам, без ручного поиска по чатам и таблицам."
    )

async def disable_auto_status_for_order(order_id: int, actor_id: int = 0) -> Optional[aiosqlite.Row]:
    order = await get_order(order_id)
    if not order:
        return None
    async with get_db() as db:
        await db.execute(
            "UPDATE orders SET auto_status_enabled=0, updated_at=? WHERE id=?",
            (now_iso(), order_id),
        )
        await db.execute(
            "INSERT INTO status_history (order_id, status, comment, created_by, created_at) VALUES (?, ?, ?, ?, ?)",
            (order_id, order["status"] or "новая заявка", "Автостатусы выключены", actor_id, now_iso()),
        )
        await db.commit()
    return await get_order(order_id)


def parse_iso_datetime(value: str) -> datetime:
    if not value:
        return datetime.now(timezone.utc)
    try:
        dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except Exception:
        return datetime.now(timezone.utc)


def get_auto_target_status(order: aiosqlite.Row) -> Optional[tuple[str, str, str]]:
    route_key = order["auto_status_route"] or choose_auto_route(order["from_country"] or "", order["to_city"] or "")
    route = AUTO_STATUS_ROUTES.get(route_key) or AUTO_STATUS_ROUTES["cis_local"]
    started_at = parse_iso_datetime(order["auto_status_started_at"] or order["created_at"] or now_iso())
    elapsed_seconds = (datetime.now(timezone.utc) - started_at).total_seconds()
    # В реальной работе этапы считаются днями. Для демонстрации можно поставить
    # AUTO_STATUS_TIME_UNIT=minutes, тогда те же значения станут минутами.
    elapsed_units = elapsed_seconds / 60 if AUTO_STATUS_TIME_UNIT == "minutes" else elapsed_seconds / 86400
    target = None
    target_index = -1
    for idx, (day, status, comment) in enumerate(route["stages"]):
        if elapsed_units >= day:
            target = (status, comment, route["name"])
            target_index = idx
    if not target:
        return None
    current_status = order["status"] or "новая заявка"
    if current_status in TERMINAL_STATUSES:
        return None
    stage_statuses = [s for _, s, _ in route["stages"]]
    current_index = stage_statuses.index(current_status) if current_status in stage_statuses else -1
    if target_index <= current_index or target[0] == current_status:
        return None
    return target


async def create_order(
    user_id: int,
    order_type: str,
    from_country: str,
    to_city: str,
    cargo_type: str,
    weight: float,
    volume: float,
    description: str,
    customer_name: str,
    phone: str,
    partner_id: Optional[int] = None,
) -> aiosqlite.Row:
    async with get_db() as db:
        cur = await db.execute(
            """
            INSERT INTO orders (
                user_id, order_type, from_country, to_city, cargo_type,
                weight, volume, description, customer_name, phone,
                status, partner_id, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'новая заявка', ?, ?, ?)
            """,
            (
                user_id,
                order_type,
                from_country,
                to_city,
                cargo_type,
                weight,
                volume,
                description,
                customer_name,
                phone,
                partner_id,
                now_iso(),
                now_iso(),
            ),
        )
        order_id = cur.lastrowid
        code = f"CG{datetime.now().strftime('%y%m%d')}{order_id:05d}"
        route_key = choose_auto_route(from_country, to_city)
        await db.execute(
            "UPDATE orders SET tracking_code=?, auto_status_route=?, auto_status_enabled=0, auto_status_started_at=NULL WHERE id=?",
            (code, route_key, order_id),
        )
        await db.execute(
            "INSERT INTO status_history (order_id, status, comment, created_by, created_at) VALUES (?, ?, ?, ?, ?)",
            (
                order_id,
                "новая заявка",
                "Заявка создана клиентом. Автостатусы ждут оплаты и включения админом.",
                user_id,
                now_iso(),
            ),
        )
        await db.commit()
        return await db_fetchone(db, "SELECT * FROM orders WHERE id=?", (order_id,))


async def update_order_status(
    order_id: int,
    status: str,
    created_by: int,
    comment: str = "",
) -> Optional[aiosqlite.Row]:
    async with get_db() as db:
        await db.execute(
            "UPDATE orders SET status=?, updated_at=? WHERE id=?",
            (status, now_iso(), order_id),
        )
        await db.execute(
            "INSERT INTO status_history (order_id, status, comment, created_by, created_at) VALUES (?, ?, ?, ?, ?)",
            (order_id, status, comment, created_by, now_iso()),
        )
        await db.commit()
        return await db_fetchone(db, "SELECT * FROM orders WHERE id=?", (order_id,))


async def set_order_price(code: str, price: float, cost: float) -> Optional[aiosqlite.Row]:
    margin = price - cost
    async with get_db() as db:
        existing = await db_fetchone(db, "SELECT paid_amount FROM orders WHERE UPPER(tracking_code)=UPPER(?)", (code,))
        if not existing:
            return None
        paid = float(existing["paid_amount"] or 0)
        payment_status = "оплачено" if price > 0 and paid >= price else "частично" if paid > 0 else "не оплачено"
        await db.execute(
            "UPDATE orders SET price=?, cost=?, margin=?, payment_status=?, updated_at=? WHERE UPPER(tracking_code)=UPPER(?)",
            (price, cost, margin, payment_status, now_iso(), code),
        )
        await db.commit()
        return await db_fetchone(db, "SELECT * FROM orders WHERE UPPER(tracking_code)=UPPER(?)", (code,))


async def set_order_weight(code: str, weight: float, actor_id: int) -> Optional[aiosqlite.Row]:
    async with get_db() as db:
        row = await db_fetchone(db, "SELECT id FROM orders WHERE UPPER(tracking_code)=UPPER(?)", (code,))
        if not row:
            return None
        await db.execute("UPDATE orders SET weight=?, updated_at=? WHERE id=?", (weight, now_iso(), row["id"]))
        await db.execute(
            "INSERT INTO status_history (order_id, status, comment, created_by, created_at) VALUES (?, ?, ?, ?, ?)",
            (row["id"], "вес обновлён", f"Вес: {weight} кг", actor_id, now_iso()),
        )
        await db.commit()
        return await db_fetchone(db, "SELECT * FROM orders WHERE id=?", (row["id"],))


async def get_rate_from_db(from_country: str) -> float:
    key = normalize(from_country)
    async with get_db() as db:
        row = await db_fetchone(db, "SELECT rate FROM tariffs WHERE country_key=?", (key,))
    if row:
        return float(row["rate"])
    return float(DEFAULT_RATES.get(key, 3.5))


async def estimate_delivery_price_db(from_country: str, weight: float, volume: float = 0) -> tuple[float, float]:
    rate = await get_rate_from_db(from_country)
    billable_weight = max(float(weight or 0), float(volume or 0) * 167 if volume else 0)
    if billable_weight <= 0:
        billable_weight = float(weight or 0)
    return round(rate, 2), round(rate * billable_weight, 2)


async def set_tariff(country_name: str, rate: float) -> None:
    key = normalize(country_name)
    async with get_db() as db:
        await db.execute(
            """
            INSERT INTO tariffs (country_key, country_name, rate, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(country_key) DO UPDATE SET country_name=excluded.country_name, rate=excluded.rate, updated_at=excluded.updated_at
            """,
            (key, country_name.strip(), rate, now_iso(), now_iso()),
        )
        await db.commit()


async def list_tariffs() -> list[aiosqlite.Row]:
    async with get_db() as db:
        return await db_fetchall(db, "SELECT * FROM tariffs ORDER BY country_name")


async def get_status_history(order_id: int) -> list[aiosqlite.Row]:
    async with get_db() as db:
        return await db_fetchall(
            db,
            "SELECT * FROM status_history WHERE order_id=? ORDER BY id ASC",
            (order_id,),
        )


async def add_cargo_photo(order_id: int, file_id: str, comment: str, created_by: int) -> None:
    async with get_db() as db:
        await db.execute(
            "INSERT INTO cargo_photos (order_id, file_id, comment, created_by, created_at) VALUES (?, ?, ?, ?, ?)",
            (order_id, file_id, comment, created_by, now_iso()),
        )
        await db.commit()


async def list_cargo_photos(order_id: int) -> list[aiosqlite.Row]:
    async with get_db() as db:
        return await db_fetchall(
            db,
            "SELECT * FROM cargo_photos WHERE order_id=? ORDER BY id DESC LIMIT 10",
            (order_id,),
        )


async def add_payment(code: str, amount: float, comment: str, created_by: int) -> Optional[aiosqlite.Row]:
    async with get_db() as db:
        order = await db_fetchone(db, "SELECT * FROM orders WHERE UPPER(tracking_code)=UPPER(?)", (code,))
        if not order:
            return None
        paid = float(order["paid_amount"] or 0) + amount
        price = float(order["price"] or 0)
        payment_status = "оплачено" if price > 0 and paid >= price else "частично" if paid > 0 else "не оплачено"
        await db.execute(
            "INSERT INTO payments (order_id, amount, comment, created_by, created_at) VALUES (?, ?, ?, ?, ?)",
            (order["id"], amount, comment, created_by, now_iso()),
        )
        await db.execute(
            "UPDATE orders SET paid_amount=?, payment_status=?, updated_at=? WHERE id=?",
            (paid, payment_status, now_iso(), order["id"]),
        )
        await db.commit()
        return await db_fetchone(db, "SELECT * FROM orders WHERE id=?", (order["id"],))


def demo_payment_amount_for_order(order: aiosqlite.Row) -> float:
    price = float(order["price"] or 0)
    paid = float(order["paid_amount"] or 0)
    debt = max(price - paid, 0)
    if debt <= 0:
        return float(DEMO_PAYMENT_AMOUNT)
    return round(debt, 2)


async def ensure_demo_price_if_needed(order: aiosqlite.Row) -> aiosqlite.Row:
    """Для демо ставим тестовую сумму, если цена ещё не указана менеджером."""
    price = float(order["price"] or 0)
    if price > 0:
        return order
    cost = round(float(DEMO_PAYMENT_AMOUNT) * 0.7, 2)
    updated = await set_order_price(order["tracking_code"], float(DEMO_PAYMENT_AMOUNT), cost)
    return updated or order


async def create_demo_payment_request(code: str, provider: str = "kaspi") -> Optional[dict]:
    order = await get_order_by_code(code)
    if not order:
        return None
    order = await ensure_demo_price_if_needed(order)
    provider_key = provider if provider in DEMO_PAYMENT_PROVIDERS else "kaspi"
    amount = demo_payment_amount_for_order(order)
    payment_id = f"DP{datetime.now().strftime('%y%m%d%H%M%S')}{order['id']}"
    async with get_db() as db:
        await db.execute(
            """
            INSERT INTO demo_payment_requests (payment_id, order_id, tracking_code, amount, status, provider, created_at)
            VALUES (?, ?, ?, ?, 'created', ?, ?)
            """,
            (payment_id, order["id"], order["tracking_code"], amount, provider_key, now_iso()),
        )
        await db.commit()
    return {
        "payment_id": payment_id,
        "order": await get_order(order["id"]),
        "amount": amount,
        "provider": provider_key,
        "provider_info": provider_info(provider_key),
        "checkout_url": f"{PUBLIC_BASE_URL}/demo-pay/{order['tracking_code']}/checkout/{provider_key}?payment_id={payment_id}",
    }


async def complete_demo_payment(code: str, provider: str = "kaspi", payment_id: str = "", actor_id: int = 0) -> Optional[dict]:
    order = await get_order_by_code(code)
    if not order:
        return None
    order = await ensure_demo_price_if_needed(order)
    provider_key = provider if provider in DEMO_PAYMENT_PROVIDERS else "kaspi"

    async with get_db() as db:
        existing_payment = None
        if payment_id:
            existing_payment = await db_fetchone(
                db,
                "SELECT status, amount FROM demo_payment_requests WHERE payment_id=?",
                (payment_id,),
            )

    # Защита от двойного клика: если payment_id уже paid, второй раз оплату не добавляем.
    if existing_payment and existing_payment["status"] == "paid":
        current = await get_order_by_code(code)
        return {
            "order": current,
            "amount": float(existing_payment["amount"] or 0),
            "provider": provider_key,
            "payment_id": payment_id,
            "already_paid": True,
        }

    amount = demo_payment_amount_for_order(order)
    updated = await add_payment(
        order["tracking_code"],
        amount,
        f"Виртуальная демо-оплата: {provider_info(provider_key)['label']} · {payment_id or 'без payment_id'}",
        actor_id,
    )
    if not updated:
        return None

    async with get_db() as db:
        if payment_id:
            await db.execute(
                "UPDATE demo_payment_requests SET status='paid', paid_at=? WHERE payment_id=?",
                (now_iso(), payment_id),
            )
        else:
            payment_id = f"DP{datetime.now().strftime('%y%m%d%H%M%S')}{order['id']}"
            await db.execute(
                """
                INSERT INTO demo_payment_requests (payment_id, order_id, tracking_code, amount, status, provider, created_at, paid_at)
                VALUES (?, ?, ?, ?, 'paid', ?, ?, ?)
                """,
                (payment_id, order["id"], order["tracking_code"], amount, provider_key, now_iso(), now_iso()),
            )
        await db.commit()

    return {
        "order": updated,
        "amount": amount,
        "provider": provider_key,
        "payment_id": payment_id,
        "already_paid": False,
    }


async def list_debts(limit: int = 30) -> list[aiosqlite.Row]:
    async with get_db() as db:
        return await db_fetchall(
            db,
            """
            SELECT *, (COALESCE(price, 0) - COALESCE(paid_amount, 0)) AS debt
            FROM orders
            WHERE COALESCE(price, 0) > COALESCE(paid_amount, 0)
            ORDER BY updated_at DESC LIMIT ?
            """,
            (limit,),
        )


async def assign_courier(code: str, courier_id: int, address: str, actor_id: int) -> Optional[aiosqlite.Row]:
    async with get_db() as db:
        row = await db_fetchone(db, "SELECT id FROM orders WHERE UPPER(tracking_code)=UPPER(?)", (code,))
        if not row:
            return None
        await db.execute(
            "UPDATE orders SET courier_id=?, delivery_address=?, status=?, updated_at=? WHERE id=?",
            (courier_id, address, "передан курьеру", now_iso(), row["id"]),
        )
        await db.execute(
            "INSERT INTO status_history (order_id, status, comment, created_by, created_at) VALUES (?, ?, ?, ?, ?)",
            (row["id"], "передан курьеру", f"Курьер: {courier_id}. Адрес: {address}", actor_id, now_iso()),
        )
        await db.commit()
        return await db_fetchone(db, "SELECT * FROM orders WHERE id=?", (row["id"],))


async def courier_orders(courier_id: int) -> list[aiosqlite.Row]:
    async with get_db() as db:
        return await db_fetchall(
            db,
            "SELECT * FROM orders WHERE courier_id=? AND status!='доставлен' ORDER BY updated_at DESC LIMIT 30",
            (courier_id,),
        )


async def list_orders(limit: int = 10, status: Optional[str] = None, user_id: Optional[int] = None) -> list[aiosqlite.Row]:
    query = "SELECT * FROM orders"
    params = []
    conditions = []
    if status:
        conditions.append("status=?")
        params.append(status)
    if user_id:
        conditions.append("user_id=?")
        params.append(user_id)
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    query += " ORDER BY id DESC LIMIT ?"
    params.append(limit)
    async with get_db() as db:
        return await db_fetchall(db, query, tuple(params))


async def orders_on_warehouse() -> list[aiosqlite.Row]:
    async with get_db() as db:
        return await db_fetchall(db, 
            """
            SELECT * FROM orders
            WHERE status IN ('принят на склад', 'ожидает отправки')
            ORDER BY id DESC LIMIT 20
            """
        )


async def create_complaint(user_id: int, code: str, text: str, urgency: str) -> aiosqlite.Row:
    order = await get_order_by_code(code) if code else None
    order_id = order["id"] if order else None
    async with get_db() as db:
        cur = await db.execute(
            """
            INSERT INTO complaints (user_id, order_id, tracking_code, text, urgency, status, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, 'open', ?, ?)
            """,
            (user_id, order_id, code, text, urgency, now_iso(), now_iso()),
        )
        cid = cur.lastrowid
        await db.commit()
        return await db_fetchone(db, "SELECT * FROM complaints WHERE id=?", (cid,))


async def list_open_complaints(limit: int = 10) -> list[aiosqlite.Row]:
    async with get_db() as db:
        return await db_fetchall(db, 
            "SELECT * FROM complaints WHERE status='open' ORDER BY id DESC LIMIT ?",
            (limit,),
        )


async def close_complaint(complaint_id: int) -> None:
    async with get_db() as db:
        await db.execute(
            "UPDATE complaints SET status='closed', updated_at=? WHERE id=?",
            (now_iso(), complaint_id),
        )
        await db.commit()


async def reply_complaint(complaint_id: int, reply: str) -> Optional[aiosqlite.Row]:
    async with get_db() as db:
        await db.execute(
            "UPDATE complaints SET admin_reply=?, status='closed', updated_at=? WHERE id=?",
            (reply, now_iso(), complaint_id),
        )
        await db.commit()
        return await db_fetchone(db, "SELECT * FROM complaints WHERE id=?", (complaint_id,))



def support_topic_label(topic: str) -> str:
    labels = {
        "cargo": "Вопрос по грузу",
        "payment": "Оплата / долг",
        "damage": "Фото / повреждение",
        "bot": "Ошибка в боте",
        "other": "Другое",
    }
    return labels.get(topic or "", topic or "Другое")


async def create_support_ticket(user_id: int, code: str, topic: str, text: str) -> aiosqlite.Row:
    async with get_db() as db:
        cur = await db.execute(
            """
            INSERT INTO support_tickets (user_id, tracking_code, topic, status, priority, last_message, created_at, updated_at)
            VALUES (?, ?, ?, 'open', 'normal', ?, ?, ?)
            """,
            (user_id, code, topic, text, now_iso(), now_iso()),
        )
        ticket_id = cur.lastrowid
        await db.execute(
            """
            INSERT INTO support_messages (ticket_id, sender_id, sender_role, text, created_at)
            VALUES (?, ?, 'client', ?, ?)
            """,
            (ticket_id, user_id, text, now_iso()),
        )
        await db.commit()
        return await db_fetchone(db, "SELECT * FROM support_tickets WHERE id=?", (ticket_id,))


async def get_support_ticket(ticket_id: int) -> Optional[aiosqlite.Row]:
    async with get_db() as db:
        return await db_fetchone(db, "SELECT * FROM support_tickets WHERE id=?", (ticket_id,))


async def list_open_support_tickets(limit: int = 15) -> list[aiosqlite.Row]:
    async with get_db() as db:
        return await db_fetchall(
            db,
            "SELECT * FROM support_tickets WHERE status!='closed' ORDER BY updated_at DESC LIMIT ?",
            (limit,),
        )


async def list_user_support_tickets(user_id: int, limit: int = 10) -> list[aiosqlite.Row]:
    async with get_db() as db:
        return await db_fetchall(
            db,
            "SELECT * FROM support_tickets WHERE user_id=? ORDER BY updated_at DESC LIMIT ?",
            (user_id, limit),
        )


async def list_support_messages(ticket_id: int, limit: int = 10) -> list[aiosqlite.Row]:
    async with get_db() as db:
        return await db_fetchall(
            db,
            "SELECT * FROM support_messages WHERE ticket_id=? ORDER BY id DESC LIMIT ?",
            (ticket_id, limit),
        )


async def add_support_message(ticket_id: int, sender_id: int, sender_role: str, text: str, new_status: Optional[str] = None) -> Optional[aiosqlite.Row]:
    ticket = await get_support_ticket(ticket_id)
    if not ticket:
        return None
    status = new_status or ("answered" if sender_role == "admin" else "open")
    async with get_db() as db:
        await db.execute(
            """
            INSERT INTO support_messages (ticket_id, sender_id, sender_role, text, created_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (ticket_id, sender_id, sender_role, text, now_iso()),
        )
        await db.execute(
            "UPDATE support_tickets SET last_message=?, status=?, updated_at=? WHERE id=?",
            (text, status, now_iso(), ticket_id),
        )
        await db.commit()
        return await db_fetchone(db, "SELECT * FROM support_tickets WHERE id=?", (ticket_id,))


async def close_support_ticket(ticket_id: int) -> Optional[aiosqlite.Row]:
    async with get_db() as db:
        await db.execute(
            "UPDATE support_tickets SET status='closed', updated_at=?, closed_at=? WHERE id=?",
            (now_iso(), now_iso(), ticket_id),
        )
        await db.commit()
        return await db_fetchone(db, "SELECT * FROM support_tickets WHERE id=?", (ticket_id,))


def format_support_ticket(ticket: aiosqlite.Row) -> str:
    return (
        f"<b>🛠️ Обращение №{ticket['id']}</b>\n"
        f"Тема: {safe(support_topic_label(ticket['topic']))}\n"
        f"Груз: {safe(ticket['tracking_code'] or 'не указан')}\n"
        f"Клиент ID: <code>{ticket['user_id']}</code>\n"
        f"Статус: <b>{safe(ticket['status'])}</b>\n"
        f"Обновлено: {safe((ticket['updated_at'] or '')[:16].replace('T', ' '))}\n\n"
        f"Последнее сообщение:\n{safe(ticket['last_message'])}"
    )


async def finance_report() -> dict:
    async with get_db() as db:
        totals = await db_fetchone(db, 
            """
            SELECT COUNT(*) AS total_orders,
                   COALESCE(SUM(price), 0) AS revenue,
                   COALESCE(SUM(cost), 0) AS cost,
                   COALESCE(SUM(margin), 0) AS margin
            FROM orders
            """
        )
        today = datetime.now().strftime("%Y-%m-%d")
        today_row = await db_fetchone(db, 
            """
            SELECT COUNT(*) AS today_orders,
                   COALESCE(SUM(price), 0) AS today_revenue,
                   COALESCE(SUM(cost), 0) AS today_cost,
                   COALESCE(SUM(margin), 0) AS today_margin
            FROM orders
            WHERE created_at LIKE ?
            """,
            (f"%{today}%",),
        )
        by_status = await db_fetchall(db, 
            "SELECT status, COUNT(*) AS cnt FROM orders GROUP BY status ORDER BY cnt DESC"
        )
        return {"totals": totals, "today": today_row, "by_status": by_status}


async def partner_report(user_id: int) -> dict:
    async with get_db() as db:
        leads = await db_fetchone(db, 
            "SELECT COUNT(*) AS cnt FROM partner_leads WHERE partner_id=?",
            (user_id,),
        )
        orders = await db_fetchone(db, 
            """
            SELECT COUNT(*) AS cnt, COALESCE(SUM(price), 0) AS revenue
            FROM orders WHERE partner_id=?
            """,
            (user_id,),
        )
        partners = await db_fetchall(db, 
            "SELECT partner_id, COUNT(*) AS cnt FROM partner_leads GROUP BY partner_id ORDER BY cnt DESC LIMIT 10"
        )
        return {"leads": leads, "orders": orders, "partners": partners}


async def notify_admins(bot: Bot, text: str, reply_markup: Optional[InlineKeyboardMarkup] = None) -> None:
    for admin_id in ADMIN_IDS:
        try:
            await bot.send_message(admin_id, text, reply_markup=reply_markup)
        except Exception as e:
            logger.warning("Failed to notify admin %s: %s", admin_id, e)


def admin_payment_notify_text(order: aiosqlite.Row, provider: str, amount: float, payment_id: str = "") -> str:
    info = provider_info(provider)
    return (
        "<b>💳 Новая оплата</b>\n\n"
        f"Груз: <b>{safe(order['tracking_code'])}</b>\n"
        f"Банк: <b>{safe(info['label'])}</b>\n"
        f"Сумма: <b>{amount:g} {safe(CURRENCY)}</b>\n"
        f"Статус оплаты: <b>{safe(order['payment_status'] or 'оплачено')}</b>\n"
        f"Всего оплачено: <b>{safe(str(order['paid_amount'] or 0))} {safe(CURRENCY)}</b>\n"
        f"Payment ID: <code>{safe(payment_id or 'demo')}</code>\n\n"
        "Клиент оплатил груз. Теперь можно включить автостатусы одной кнопкой."
    )


async def notify_user(bot: Bot, user_id: int, text: str, reply_markup: Optional[InlineKeyboardMarkup] = None) -> None:
    try:
        await bot.send_message(user_id, text, reply_markup=reply_markup)
    except Exception as e:
        logger.warning("Failed to notify user %s: %s", user_id, e)


def detect_intent(text: str) -> Optional[str]:
    t = normalize(text)
    if any(w in t for w in ["где груз", "где посыл", "статус", "трек", "когда прид", "мой груз"]):
        return "tracking"
    if any(w in t for w in ["сколько", "цена", "стоимость", "тариф", "рассчитать", "за кг"]):
        return "calculator"
    if any(w in t for w in ["выкуп", "купить из китая", "1688", "таобао", "pinduoduo", "ссылка на товар"]):
        return "buyout"
    if any(w in t for w in ["опт", "партия", "короб", "контейнер", "большой груз"]):
        return "wholesale"
    if any(w in t for w in ["поддерж", "техпод", "оператор", "менеджер", "помощ", "саппорт"]):
        return "support"
    if any(w in t for w in ["жалоба", "проблем", "задерж", "повреж", "потерял"]):
        return "complaint"
    if any(w in t for w in ["документ", "тамож", "инвойс", "декларац"]):
        return "customs"
    return None


def estimate_delivery_price(from_country: str, weight: float, volume: float = 0) -> tuple[float, float]:
    key = normalize(from_country)
    rate = DEFAULT_RATES.get(key, 3.5)
    billable_weight = max(float(weight or 0), float(volume or 0) * 167 if volume else 0)
    if billable_weight <= 0:
        billable_weight = float(weight or 0)
    return round(rate, 2), round(rate * billable_weight, 2)


def format_order(order: aiosqlite.Row) -> str:
    return (
        f"<b>📦 Груз {safe(order['tracking_code'])}</b>\n"
        f"Тип заявки: {safe(order['order_type'])}\n"
        f"Маршрут: {safe(order['from_country'])} → {safe(order['to_city'])}\n"
        f"Товар: {safe(order['cargo_type'])}\n"
        f"Вес: {safe(order['weight'])} кг\n"
        f"Объём: {safe(order['volume'])} м³\n"
        f"Статус: <b>{safe(order['status'])}</b>\n"
        f"Автостатусы: {'вкл' if int(order['auto_status_enabled'] or 0) else 'выкл'} | {safe(auto_route_label(order['auto_status_route'] or choose_auto_route(order['from_country'] or '', order['to_city'] or '')))}\n"
        f"Цена клиенту: {safe(order['price'])} {safe(CURRENCY)}\n"
        f"Оплачено: {safe(order['paid_amount'])} {safe(CURRENCY)} | {safe(order['payment_status'])}\n"
        f"Курьер: {safe(order['courier_id']) if order['courier_id'] else 'не назначен'}\n"
        f"Адрес доставки: {safe(order['delivery_address']) if order['delivery_address'] else 'не указан'}\n"
        f"Описание: {safe(order['description'])}\n"
        f"Клиент: {safe(order['customer_name'])}, {safe(order['phone'])}"
    )


def format_short_order(order: aiosqlite.Row) -> str:
    return (
        f"{safe(order['tracking_code'])} | {safe(order['status'])} | "
        f"{safe(order['from_country'])} → {safe(order['to_city'])} | {safe(order['weight'])} кг"
    )


async def show_order_to_admin(bot: Bot, order: aiosqlite.Row) -> None:
    text = "<b>🆕 Новая заявка</b>\n\n" + format_order(order)
    await notify_admins(bot, text, new_order_actions_keyboard(order["id"]))


def format_history(rows: list[aiosqlite.Row]) -> str:
    if not rows:
        return "История пока пустая."
    lines = []
    for r in rows:
        dt = safe((r["created_at"] or "")[:16].replace("T", " "))
        comment = f" — {safe(r['comment'])}" if r["comment"] else ""
        lines.append(f"{dt}: <b>{safe(r['status'])}</b>{comment}")
    return "\n".join(lines)


def make_qr_bytes(text: str) -> bytes:
    img = qrcode.make(text)
    bio = BytesIO()
    img.save(bio, format="PNG")
    bio.seek(0)
    return bio.read()


def _register_pdf_font() -> str:
    for fp in [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
    ]:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("CargoFont", fp))
                return "CargoFont"
            except Exception:
                pass
    return "Helvetica"


def generate_receipt_pdf(order: aiosqlite.Row) -> Path:
    font = _register_pdf_font()
    path = RECEIPT_DIR / f"receipt_{order['tracking_code']}.pdf"
    c = canvas.Canvas(str(path), pagesize=A4)
    width, height = A4
    y = height - 60
    c.setFont(font, 16)
    c.drawString(50, y, f"{COMPANY_NAME} — квитанция")
    y -= 40
    c.setFont(font, 11)
    lines = [
        f"Номер груза: {order['tracking_code']}",
        f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        f"Клиент: {order['customer_name']} / {order['phone']}",
        f"Маршрут: {order['from_country']} → {order['to_city']}",
        f"Товар: {order['cargo_type']}",
        f"Вес: {order['weight']} кг",
        f"Статус: {order['status']}",
        f"Цена: {order['price']} {CURRENCY}",
        f"Оплачено: {order['paid_amount']} {CURRENCY}",
        f"Статус оплаты: {order['payment_status']}",
        f"Адрес доставки: {order['delivery_address'] or 'не указан'}",
        "",
        "Квитанция сформирована автоматически в Telegram-системе CargoAI CRM.",
    ]
    for line in lines:
        c.drawString(50, y, str(line))
        y -= 22
    c.showPage()
    c.save()
    return path


async def export_orders_to_excel() -> Path:
    path = EXPORT_DIR / f"orders_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "orders"
    headers = [
        "tracking_code", "status", "order_type", "from_country", "to_city", "cargo_type",
        "weight", "volume", "price", "cost", "margin", "paid_amount", "payment_status",
        "customer_name", "phone", "courier_id", "delivery_address", "created_at", "updated_at", "description"
    ]
    ws.append(headers)
    async with get_db() as db:
        rows = await db_fetchall(db, "SELECT * FROM orders ORDER BY id DESC")
    for o in rows:
        ws.append([o[h] if h in o.keys() else "" for h in headers])
    widths = {"A": 18, "B": 18, "C": 14, "D": 14, "E": 14, "F": 18, "N": 18, "O": 18, "Q": 26, "T": 45}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    wb.save(path)
    return path


async def import_statuses_from_excel(path: Path, actor_id: int, bot: Bot) -> tuple[int, int]:
    wb = load_workbook(path)
    ws = wb.active
    headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    def idx(name: str) -> Optional[int]:
        return headers.index(name) if name in headers else None
    i_code, i_status, i_comment = idx("tracking_code"), idx("status"), idx("comment")
    i_price, i_cost, i_paid = idx("price"), idx("cost"), idx("paid_amount")
    if i_code is None:
        return 0, 0
    ok = 0
    fail = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(row[i_code] or "").strip().upper()
        if not code:
            continue
        try:
            order = await get_order_by_code(code)
            if not order:
                fail += 1
                continue
            if i_status is not None and row[i_status]:
                status = str(row[i_status]).strip()
                comment = str(row[i_comment] or "") if i_comment is not None else "Импорт из Excel"
                updated = await update_order_status(order["id"], status, actor_id, comment)
                await notify_user(bot, updated["user_id"], f"📦 Статус груза <b>{safe(code)}</b> обновлён: <b>{safe(status)}</b>")
            if i_price is not None and row[i_price] is not None:
                price = float(row[i_price] or 0)
                cost = float(row[i_cost] or 0) if i_cost is not None else float(order["cost"] or 0)
                await set_order_price(code, price, cost)
            if i_paid is not None and row[i_paid] is not None:
                current = await get_order_by_code(code)
                target_paid = float(row[i_paid] or 0)
                delta = target_paid - float(current["paid_amount"] or 0)
                if abs(delta) > 0.0001:
                    await add_payment(code, delta, "Импорт оплаты из Excel", actor_id)
            ok += 1
        except Exception as e:
            logger.exception("Excel import failed for %s: %s", code, e)
            fail += 1
    return ok, fail


# =========================
# FSM STATES
# =========================
class CargoForm(StatesGroup):
    from_country = State()
    to_city = State()
    weight = State()
    volume = State()
    cargo_type = State()
    description = State()
    customer_name = State()
    phone = State()


class CalcForm(StatesGroup):
    from_country = State()
    weight = State()
    volume = State()


class TrackForm(StatesGroup):
    code = State()


class BuyoutForm(StatesGroup):
    link = State()
    details = State()
    customer_name = State()
    phone = State()


class WholesaleForm(StatesGroup):
    from_country = State()
    to_city = State()
    goods = State()
    boxes = State()
    weight = State()
    documents = State()
    customer_name = State()
    phone = State()


class ComplaintForm(StatesGroup):
    code = State()
    text = State()
    urgency = State()


class SupportForm(StatesGroup):
    code = State()
    topic = State()
    text = State()


class ClientSupportReplyForm(StatesGroup):
    ticket_id = State()
    text = State()


class AdminSupportReplyForm(StatesGroup):
    ticket_id = State()
    text = State()


class StatusForm(StatesGroup):
    code = State()
    status = State()
    comment = State()


class AutoStatusForm(StatesGroup):
    code = State()


class WeightForm(StatesGroup):
    code = State()
    weight = State()


class BroadcastForm(StatesGroup):
    audience = State()
    text = State()


class HistoryForm(StatesGroup):
    code = State()


class PhotoAddForm(StatesGroup):
    code = State()
    photo = State()


class PhotoViewForm(StatesGroup):
    code = State()


class QrForm(StatesGroup):
    code = State()


class ReceiptForm(StatesGroup):
    code = State()


class PaymentForm(StatesGroup):
    code = State()
    amount = State()
    comment = State()


class ExcelImportForm(StatesGroup):
    file = State()


class TariffForm(StatesGroup):
    country = State()
    rate = State()


class CourierAssignForm(StatesGroup):
    code = State()
    courier_id = State()
    address = State()


class CourierDeliveredForm(StatesGroup):
    code = State()


class BatchCreateForm(StatesGroup):
    name = State()
    route = State()
    codes = State()


class BatchAddForm(StatesGroup):
    batch_id = State()
    codes = State()


class BatchStatusForm(StatesGroup):
    batch_id = State()


class StatusSettingsForm(StatesGroup):
    name = State()


class ClientSearchForm(StatesGroup):
    query = State()


# =========================
# COMMANDS / START
# =========================
@router.message(CommandStart())
async def cmd_start(message: Message, state: FSMContext, command: CommandObject = None):
    await state.clear()
    ref_partner_id = None
    args = command.args if command else None
    if args and args.startswith("partner_"):
        raw = args.replace("partner_", "", 1)
        if raw.isdigit():
            ref_partner_id = int(raw)
    await upsert_user(message, ref_partner_id)

    user_id = message.from_user.id if message.from_user else 0
    name = message.from_user.first_name if message.from_user else ""
    menu_text, menu_keyboard = await menu_for_user(user_id, name)

    if args and args.startswith("track_"):
        code = args.replace("track_", "", 1).strip().upper()
        order = await get_order_by_code(code)
        if order:
            rows = await get_status_history(order["id"])
            await message.answer(
                format_order(order) + "\n\n<b>🕓 История:</b>\n" + format_history(rows),
                reply_markup=track_button(order["tracking_code"]),
            )
            await message.answer(menu_text, reply_markup=menu_keyboard)
            return

    await message.answer(menu_text, reply_markup=menu_keyboard)


@router.message(Command("admin"))
async def cmd_admin(message: Message, state: FSMContext):
    await state.clear()
    await upsert_user(message)
    if not message.from_user or not is_admin(message.from_user.id):
        await message.answer("⛔ У вас нет доступа к админ‑меню.")
        return
    await message.answer(admin_menu_text(), reply_markup=admin_keyboard())


@router.message(Command("warehouse"))
async def cmd_warehouse(message: Message, state: FSMContext):
    await state.clear()
    await upsert_user(message)
    if not message.from_user or not is_warehouse(message.from_user.id):
        await message.answer("⛔ У вас нет доступа к складскому меню.")
        return
    await message.answer(warehouse_menu_text(), reply_markup=warehouse_keyboard())


@router.message(Command("help"))
async def cmd_help(message: Message):
    if message.from_user and is_admin(message.from_user.id):
        await message.answer(
            "<b>Помощь CargoPilot Pro</b>\n\n"
            "/start — клиентское меню\n"
            "/admin — админ-меню\n"
            "/warehouse — меню склада\n"
            "/courier — меню курьера\n"
            "/routes — маршруты автостатусов\n"
            "/support — открытые обращения\n"
            "/help_admin — все команды админа"
        )
    else:
        await message.answer(
            "<b>Помощь</b>\n\n"
            "Нажмите /start, чтобы открыть меню.\n"
            "Вы можете оформить доставку, рассчитать стоимость, проверить статус груза, создать обращение в техподдержку или посмотреть свои заказы."
        )


@router.message(Command("help_admin"))
async def cmd_help_admin(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    await message.answer(
        "<b>Команды админа</b>\n\n"
        "/admin — открыть админ‑меню\n"
        "/setstatus CG26050300001 в пути комментарий — сменить статус\n"
        "/setprice CG26050300001 100 70 — цена клиенту, себестоимость\n"
        "/replycomplaint 1 текст ответа — ответить на жалобу\n"
        "/support — список обращений техподдержки\n"
        "/replyticket 1 текст ответа — ответить в обращение\n"
        "/closeticket 1 — закрыть обращение\n"
        "/makepartner 123456789 — сделать пользователя партнёром\n"
        "/role 123456789 warehouse — выдать роль warehouse/client/partner/courier/admin\n"
        "/history CG... — история статусов\n"
        "/photos CG... — фото груза\n"
        "/qr CG... — QR-код груза\n"
        "/pay CG... 10000 комментарий — внести оплату\n"
        "/debts — список долгов\n"
        "/export_orders — выгрузка Excel\n"
        "/settariff Китай 3.5 — изменить тариф\n"
        "/tariffs — список тарифов\n"
        "/receipt CG... — PDF-квитанция\n"
        "/assigncourier CG... 123456789 адрес — назначить курьера\n"
        "🤖 Автостатусы — включить по кнопке без команд\n"
        "📦 Все грузы → карточка груза → 🤖 Включить автостатусы\n"
        "📦 Партии — вести грузы партиями и менять статус всем сразу\n"
        "👥 Клиенты — база клиентов и поиск\n"
        "⚙️ Статусы — добавить свои статусы через кнопки\n"
    )


@router.message(Command("setstatus"))
async def cmd_setstatus(message: Message, bot: Bot):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    parts = (message.text or "").split(maxsplit=2)
    if len(parts) < 3:
        await message.answer("Формат: /setstatus CG26050300001 в пути")
        return
    code = parts[1].upper()
    rest = parts[2]
    status = None
    for s in sorted(STATUSES, key=len, reverse=True):
        if rest.lower().startswith(s.lower()):
            status = s
            comment = rest[len(s):].strip()
            break
    if not status:
        status = rest.strip()
        comment = ""
    order = await get_order_by_code(code)
    if not order:
        await message.answer("Груз не найден.")
        return
    updated = await update_order_status(order["id"], status, message.from_user.id, comment)
    await message.answer("✅ Статус обновлён.\n\n" + format_order(updated))
    await notify_user(
        bot,
        updated["user_id"],
        status_notify_text(updated["tracking_code"], status, comment),
        reply_markup=track_button(updated["tracking_code"]),
    )


@router.message(Command("routes"))
async def cmd_routes(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    lines = ["<b>🤖 Шаблоны автостатусов</b>"]
    for key, route in AUTO_STATUS_ROUTES.items():
        steps = ", ".join(f"день {d}: {s}" for d, s, _ in route["stages"][:4])
        lines.append(f"<code>{safe(key)}</code> — {safe(route['name'])}\n{safe(steps)} ...")
    lines.append("\nТеперь всё работает без команд: откройте <b>/admin</b>, нажмите <b>🤖 Автостатусы</b> и выберите действие кнопкой. Бот сам определит маршрут по направлению груза.")
    await message.answer("\n\n".join(lines), reply_markup=admin_keyboard())


@router.message(Command("autostatus"))
async def cmd_autostatus(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    parts = (message.text or "").split(maxsplit=2)
    if len(parts) < 3 or parts[2].lower() not in {"on", "off", "вкл", "выкл"}:
        await message.answer("Формат: /autostatus CG26050300001 on или /autostatus CG26050300001 off")
        return
    code = parts[1].upper()
    order = await get_order_by_code(code)
    if not order:
        await message.answer("Груз не найден.")
        return
    if parts[2].lower() in {"on", "вкл"}:
        if not await order_payment_confirmed(order):
            await message.answer(
                f"⚠️ Автостатусы для <b>{safe(code)}</b> пока нельзя включить.\n\n"
                "Сначала клиент должен оплатить груз, после оплаты админ получит уведомление и сможет включить автостатусы.",
                reply_markup=admin_keyboard(),
            )
            return
        updated = await enable_auto_status_for_order(order["id"], message.from_user.id)
        route_name = auto_route_label(updated["auto_status_route"] or choose_auto_route(updated["from_country"] or "", updated["to_city"] or ""))
        await message.answer(
            f"✅ Автостатусы включены для <b>{safe(code)}</b>.\n"
            f"Маршрут выбран автоматически: <b>{safe(route_name)}</b>.\n\n"
            "Теперь бот будет сам менять статусы по расписанию маршрута.",
            reply_markup=admin_keyboard(),
        )
    else:
        await disable_auto_status_for_order(order["id"], message.from_user.id)
        await message.answer(f"⏸ Автостатусы для <b>{safe(code)}</b> выключены.", reply_markup=admin_keyboard())


@router.message(Command("autostatus_all"))
async def cmd_autostatus_all(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2 or parts[1].lower() not in {"on", "вкл"}:
        await message.answer("Формат: /autostatus_all on")
        return
    count = await enable_auto_status_for_active_orders(message.from_user.id)
    await message.answer(
        f"✅ Автостатусы включены для активных грузов: {count}.\n\n"
        "Автостатусы включаются только для оплаченных активных грузов.",
        reply_markup=admin_keyboard(),
    )


@router.message(Command("autoroute"))
async def cmd_autoroute(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    parts = (message.text or "").split(maxsplit=2)
    if len(parts) < 3 or parts[2] not in AUTO_STATUS_ROUTES:
        await message.answer("Формат: /autoroute CG26050300001 china_cis\n\nСписок: /routes")
        return
    code = parts[1].upper()
    route_key = parts[2]
    async with get_db() as db:
        row = await db_fetchone(db, "SELECT id FROM orders WHERE UPPER(tracking_code)=UPPER(?)", (code,))
        if not row:
            await message.answer("Груз не найден.")
            return
        await db.execute(
            "UPDATE orders SET auto_status_route=?, auto_status_enabled=1, auto_status_started_at=?, updated_at=? WHERE id=?",
            (route_key, now_iso(), now_iso(), row["id"]),
        )
        await db.commit()
    await message.answer(f"✅ Маршрут автостатусов для {safe(code)}: {safe(auto_route_label(route_key))}")


@router.message(F.text == "🤖 Автостатусы")
async def admin_auto_status_menu(message: Message, state: FSMContext):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    await state.clear()
    orders = await list_orders(limit=10)
    text = (
        "<b>🤖 Автостатусы без команд</b>\n\n"
        "Админу ничего не нужно писать вручную. Нажмите кнопку ниже — бот сам включит автостатусы, "
        "сам определит маршрут по направлению груза и будет менять этапы по расписанию.\n\n"
        "<b>Как работает:</b>\n"
        "1) клиент оставляет заявку;\n"
        "2) бот выбирает маршрут: Китай/Турция/ОАЭ/СНГ;\n"
        "3) статус меняется автоматически;\n"
        "4) клиент получает уведомления.\n\n"
        "Выберите действие:"
    )
    await message.answer(text, reply_markup=auto_status_panel_keyboard(orders))


@router.message(AutoStatusForm.code)
async def auto_status_code(message: Message, state: FSMContext):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    raw = (message.text or "").strip()
    if normalize(raw) in {"все", "all", "всё"}:
        count = await enable_auto_status_for_active_orders(message.from_user.id)
        await state.clear()
        await message.answer(
            f"✅ Автостатусы включены для активных грузов: {count}.\n\n"
            "Бот сам выбрал маршруты по направлениям и будет уведомлять клиентов при смене статусов.",
            reply_markup=admin_keyboard(),
        )
        return
    code = parse_tracking_code(raw) or raw.upper()
    order = await get_order_by_code(code)
    if not order:
        await message.answer("Груз не найден. Нажмите кнопку «🤖 Автостатусы» ещё раз или введите другой номер CG...")
        return
    updated = await enable_auto_status_for_order(order["id"], message.from_user.id)
    await state.clear()
    route_name = auto_route_label(updated["auto_status_route"] or choose_auto_route(updated["from_country"] or "", updated["to_city"] or ""))
    await message.answer(
        f"✅ Автостатусы включены для <b>{safe(updated['tracking_code'])}</b>.\n"
        f"Маршрут выбран автоматически: <b>{safe(route_name)}</b>.\n\n"
        "Бот сам будет менять статусы по расписанию маршрута и уведомлять клиента.\n"
        "Если груз задержится, статус можно изменить вручную в любой момент.",
        reply_markup=admin_keyboard(),
    )


@router.message(Command("setprice"))
async def cmd_setprice(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    parts = (message.text or "").split()
    if len(parts) != 4:
        await message.answer("Формат: /setprice CG26050300001 100 70")
        return
    code = parts[1].upper()
    try:
        price = float(parts[2].replace(",", "."))
        cost = float(parts[3].replace(",", "."))
    except ValueError:
        await message.answer("Цена и себестоимость должны быть числами.")
        return
    order = await set_order_price(code, price, cost)
    if not order:
        await message.answer("Груз не найден.")
        return
    await message.answer(
        f"✅ Финансы обновлены.\n"
        f"Груз: {safe(code)}\n"
        f"Цена: {price} {safe(CURRENCY)}\n"
        f"Себестоимость: {cost} {safe(CURRENCY)}\n"
        f"Маржа: {price - cost} {safe(CURRENCY)}"
    )


@router.message(Command("replycomplaint"))
async def cmd_reply_complaint(message: Message, bot: Bot):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    parts = (message.text or "").split(maxsplit=2)
    if len(parts) < 3 or not parts[1].isdigit():
        await message.answer("Формат: /replycomplaint 1 текст ответа")
        return
    complaint = await reply_complaint(int(parts[1]), parts[2])
    if not complaint:
        await message.answer("Жалоба не найдена.")
        return
    await message.answer("✅ Ответ отправлен клиенту, жалоба закрыта.")
    await notify_user(
        bot,
        complaint["user_id"],
        f"💬 Ответ по вашей жалобе №{complaint['id']}:\n\n{safe(parts[2])}",
    )



@router.message(Command("support"))
async def cmd_support(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    tickets = await list_open_support_tickets(20)
    if not tickets:
        await message.answer("Открытых обращений в техподдержку нет.", reply_markup=admin_keyboard())
        return
    await message.answer("<b>🛠️ Открытые обращения техподдержки</b>", reply_markup=admin_keyboard())
    for t in tickets:
        await message.answer(format_support_ticket(t) + f"\n\nОтветить: /replyticket {t['id']} ваш текст", reply_markup=support_ticket_actions_keyboard(t["id"]))


@router.message(Command("replyticket"))
async def cmd_reply_ticket(message: Message, bot: Bot):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    parts = (message.text or "").split(maxsplit=2)
    if len(parts) < 3 or not parts[1].isdigit():
        await message.answer("Формат: /replyticket 1 текст ответа")
        return
    ticket = await add_support_message(int(parts[1]), message.from_user.id, "admin", parts[2], "answered")
    if not ticket:
        await message.answer("Обращение не найдено.")
        return
    await message.answer(f"✅ Ответ отправлен по обращению №{ticket['id']}.")
    await notify_user(
        bot,
        ticket["user_id"],
        f"🛠️ Ответ техподдержки по обращению №{ticket['id']}:\n\n{safe(parts[2])}\n\nЕсли вопрос не решён, нажмите «🛠️ Техподдержка» → «📋 Мои обращения» и ответьте в этот тикет.",
    )


@router.message(Command("closeticket"))
async def cmd_close_ticket(message: Message, bot: Bot):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    parts = (message.text or "").split()
    if len(parts) != 2 or not parts[1].isdigit():
        await message.answer("Формат: /closeticket 1")
        return
    ticket = await close_support_ticket(int(parts[1]))
    if not ticket:
        await message.answer("Обращение не найдено.")
        return
    await message.answer(f"✅ Обращение №{ticket['id']} закрыто.")
    await notify_user(bot, ticket["user_id"], f"✅ Ваше обращение №{ticket['id']} закрыто. Спасибо за обращение.")


@router.message(Command("makepartner"))
async def cmd_makepartner(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    parts = (message.text or "").split()
    if len(parts) != 2 or not parts[1].lstrip("-").isdigit():
        await message.answer("Формат: /makepartner 123456789")
        return
    user_id = int(parts[1])
    async with get_db() as db:
        await db.execute(
            "INSERT OR IGNORE INTO users (telegram_id, username, full_name, role, partner_code, created_at, updated_at) VALUES (?, '', '', 'partner', ?, ?, ?)",
            (user_id, partner_code(user_id), now_iso(), now_iso()),
        )
        await db.execute("UPDATE users SET role='partner', updated_at=? WHERE telegram_id=?", (now_iso(), user_id))
        await db.commit()
    await message.answer(f"✅ Пользователь {user_id} теперь партнёр.")


@router.message(Command("role"))
async def cmd_role(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    parts = (message.text or "").split()
    if len(parts) != 3 or not parts[1].lstrip("-").isdigit() or parts[2] not in {"client", "partner", "warehouse", "courier", "admin"}:
        await message.answer("Формат: /role 123456789 warehouse|courier|client|partner|admin")
        return
    user_id = int(parts[1])
    role = parts[2]
    async with get_db() as db:
        await db.execute(
            "INSERT OR IGNORE INTO users (telegram_id, username, full_name, role, partner_code, created_at, updated_at) VALUES (?, '', '', ?, ?, ?, ?)",
            (user_id, role, partner_code(user_id), now_iso(), now_iso()),
        )
        await db.execute("UPDATE users SET role=?, updated_at=? WHERE telegram_id=?", (role, now_iso(), user_id))
        await db.commit()
    await message.answer(f"✅ Роль пользователя {user_id}: {role}")


# =========================
# CLIENT FLOWS
# =========================
@router.message(Command("history"))
async def cmd_history(message: Message):
    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2:
        await message.answer("Формат: /history CG26050300001")
        return
    order = await get_order_by_code(parts[1].strip().upper())
    if not order:
        await message.answer("Груз не найден.")
        return
    if message.from_user and not is_admin(message.from_user.id) and order["user_id"] != message.from_user.id:
        await message.answer("⛔ Историю можно смотреть только по своему грузу.")
        return
    rows = await get_status_history(order["id"])
    await message.answer(f"<b>🕓 История {safe(order['tracking_code'])}</b>\n\n" + format_history(rows))


@router.message(Command("photos"))
async def cmd_photos(message: Message):
    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2:
        await message.answer("Формат: /photos CG26050300001")
        return
    order = await get_order_by_code(parts[1].strip().upper())
    if not order:
        await message.answer("Груз не найден.")
        return
    if message.from_user and not is_admin(message.from_user.id) and order["user_id"] != message.from_user.id:
        await message.answer("⛔ Фото можно смотреть только по своему грузу.")
        return
    photos = await list_cargo_photos(order["id"])
    if not photos:
        await message.answer("По этому грузу пока нет фото.")
        return
    for ph in photos:
        await message.answer_photo(ph["file_id"], caption=f"📷 {safe(order['tracking_code'])}\n{safe(ph['comment'])}\n{safe((ph['created_at'] or '')[:16].replace('T', ' '))}")


@router.message(Command("qr"))
async def cmd_qr(message: Message, bot: Bot):
    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2:
        await message.answer("Формат: /qr CG26050300001")
        return
    order = await get_order_by_code(parts[1].strip().upper())
    if not order:
        await message.answer("Груз не найден.")
        return
    if message.from_user and not is_admin(message.from_user.id) and order["user_id"] != message.from_user.id:
        await message.answer("⛔ QR можно получить только по своему грузу.")
        return
    me = await bot.get_me()
    payload = track_url(order['tracking_code'])
    qr_bytes = make_qr_bytes(payload)
    await message.answer_photo(BufferedInputFile(qr_bytes, filename=f"{order['tracking_code']}.png"), caption=f"🔳 QR для груза {safe(order['tracking_code'])}")


@router.message(Command("pay"))
async def cmd_pay(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    parts = (message.text or "").split(maxsplit=3)
    if len(parts) < 3:
        await message.answer("Формат: /pay CG26050300001 10000 комментарий")
        return
    code = parts[1].upper()
    try:
        amount = float(parts[2].replace(",", "."))
    except ValueError:
        await message.answer("Сумма должна быть числом.")
        return
    comment = parts[3] if len(parts) >= 4 else "Оплата"
    order = await add_payment(code, amount, comment, message.from_user.id)
    if not order:
        await message.answer("Груз не найден.")
        return
    debt = max(float(order["price"] or 0) - float(order["paid_amount"] or 0), 0)
    await message.answer(
        f"✅ Оплата внесена.\n\n"
        f"Груз: {safe(order['tracking_code'])}\n"
        f"Цена: {order['price']} {safe(CURRENCY)}\n"
        f"Оплачено: {order['paid_amount']} {safe(CURRENCY)}\n"
        f"Долг: {debt} {safe(CURRENCY)}\n"
        f"Статус оплаты: {safe(order['payment_status'])}"
    )


@router.message(Command("debts"))
async def cmd_debts(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    rows = await list_debts()
    if not rows:
        await message.answer("Долгов нет.")
        return
    lines = []
    for o in rows:
        lines.append(f"{safe(o['tracking_code'])} | долг: {o['debt']} {safe(CURRENCY)} | клиент: {safe(o['customer_name'])} {safe(o['phone'])}")
    await message.answer("<b>💸 Неоплаченные заказы</b>\n\n" + "\n".join(lines), reply_markup=debts_keyboard(rows))


@router.message(Command("export_orders"))
async def cmd_export_orders(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    path = await export_orders_to_excel()
    await message.answer_document(FSInputFile(path), caption="📤 Excel-выгрузка заказов")


@router.message(Command("settariff"))
async def cmd_settariff(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    parts = (message.text or "").split(maxsplit=2)
    if len(parts) < 3:
        await message.answer("Формат: /settariff Китай 3.5")
        return
    try:
        rate = float(parts[2].replace(",", "."))
    except ValueError:
        await message.answer("Тариф должен быть числом.")
        return
    await set_tariff(parts[1], rate)
    await message.answer(f"✅ Тариф обновлён: {safe(parts[1])} — {rate} {safe(CURRENCY)}/кг")


@router.message(Command("tariffs"))
async def cmd_tariffs(message: Message):
    rows = await list_tariffs()
    if not rows:
        await message.answer("Тарифов пока нет.")
        return
    lines = [f"— {safe(r['country_name'])}: {r['rate']} {safe(CURRENCY)}/кг" for r in rows]
    await message.answer("<b>⚙️ Тарифы</b>\n\n" + "\n".join(lines))


@router.message(Command("receipt"))
async def cmd_receipt(message: Message):
    if not message.from_user:
        return
    parts = (message.text or "").split(maxsplit=1)
    if len(parts) < 2:
        await message.answer("Формат: /receipt CG26050300001")
        return
    order = await get_order_by_code(parts[1].strip().upper())
    if not order:
        await message.answer("Груз не найден.")
        return
    if not is_admin(message.from_user.id) and order["user_id"] != message.from_user.id:
        await message.answer("⛔ Можно получить квитанцию только по своему грузу.")
        return
    path = generate_receipt_pdf(order)
    await message.answer_document(FSInputFile(path), caption=f"📄 Квитанция {safe(order['tracking_code'])}")


@router.message(Command("assigncourier"))
async def cmd_assign_courier(message: Message, bot: Bot):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    parts = (message.text or "").split(maxsplit=3)
    if len(parts) < 4 or not parts[2].lstrip("-").isdigit():
        await message.answer("Формат: /assigncourier CG26050300001 123456789 адрес доставки")
        return
    code = parts[1].upper()
    courier_id = int(parts[2])
    address = parts[3]
    order = await assign_courier(code, courier_id, address, message.from_user.id)
    if not order:
        await message.answer("Груз не найден.")
        return
    async with get_db() as db:
        await db.execute(
            "INSERT OR IGNORE INTO users (telegram_id, username, full_name, role, partner_code, created_at, updated_at) VALUES (?, '', '', 'courier', ?, ?, ?)",
            (courier_id, partner_code(courier_id), now_iso(), now_iso()),
        )
        await db.execute("UPDATE users SET role='courier', updated_at=? WHERE telegram_id=?", (now_iso(), courier_id))
        await db.commit()
    await message.answer("✅ Курьер назначен.\n\n" + format_order(order))
    await notify_user(bot, order["user_id"], f"🚚 Ваш груз <b>{safe(order['tracking_code'])}</b> передан курьеру. Адрес: {safe(address)}")
    await notify_user(bot, courier_id, f"🚚 Вам назначена доставка: {safe(order['tracking_code'])}\nАдрес: {safe(address)}")


@router.message(Command("courier"))
async def cmd_courier(message: Message, state: FSMContext):
    await state.clear()
    if not message.from_user or not (is_courier_static(message.from_user.id) or await has_role(message.from_user.id, "courier")):
        await message.answer("⛔ У вас нет доступа к курьерскому меню.")
        return
    await message.answer(courier_menu_text(), reply_markup=courier_keyboard())


@router.message(F.text == "👤 Клиентское меню")
async def show_client_menu(message: Message, state: FSMContext):
    await state.clear()
    name = message.from_user.first_name if message.from_user else ""
    await message.answer(client_menu_text(name), reply_markup=client_keyboard())


@router.message(F.text == "📦 Отправить груз")
@router.message(F.text == "📦 Оформить доставку")
@router.message(F.text == "📦 Оставить заявку")
async def cargo_start(message: Message, state: FSMContext):
    await state.clear()
    await state.set_state(CargoForm.from_country)
    await message.answer("Оформим заявку на доставку. Откуда груз? Например: Китай, Турция, Дубай, Россия")


@router.message(CargoForm.from_country)
async def cargo_from_country(message: Message, state: FSMContext):
    await state.update_data(from_country=message.text)
    await state.set_state(CargoForm.to_city)
    await message.answer("В какой город доставить? Например: Алматы, Астана, Шымкент")


@router.message(CargoForm.to_city)
async def cargo_to_city(message: Message, state: FSMContext):
    await state.update_data(to_city=message.text)
    await state.set_state(CargoForm.weight)
    await message.answer("Примерный вес в кг? Например: 12.5")


@router.message(CargoForm.weight)
async def cargo_weight(message: Message, state: FSMContext):
    try:
        weight = float((message.text or "").replace(",", "."))
        if weight <= 0:
            raise ValueError
    except ValueError:
        await message.answer("Введите вес числом. Например: 12.5")
        return
    await state.update_data(weight=weight)
    await state.set_state(CargoForm.volume)
    await message.answer("Объём в м³, если знаете. Если не знаете — напишите 0")


@router.message(CargoForm.volume)
async def cargo_volume(message: Message, state: FSMContext):
    try:
        volume = float((message.text or "0").replace(",", "."))
        if volume < 0:
            raise ValueError
    except ValueError:
        await message.answer("Введите объём числом. Если не знаете — 0")
        return
    await state.update_data(volume=volume)
    await state.set_state(CargoForm.cargo_type)
    await message.answer("Что за товар? Например: одежда, техника, косметика, запчасти")


@router.message(CargoForm.cargo_type)
async def cargo_type(message: Message, state: FSMContext):
    await state.update_data(cargo_type=message.text)
    await state.set_state(CargoForm.description)
    await message.answer("Добавьте комментарий: количество, особенности, нужна ли упаковка/забор товара.")


@router.message(CargoForm.description)
async def cargo_description(message: Message, state: FSMContext):
    await state.update_data(description=message.text)
    await state.set_state(CargoForm.customer_name)
    await message.answer("Ваше имя?")


@router.message(CargoForm.customer_name)
async def cargo_name(message: Message, state: FSMContext):
    await state.update_data(customer_name=message.text)
    await state.set_state(CargoForm.phone)
    await message.answer("Ваш телефон или WhatsApp/Telegram для связи?")


@router.message(CargoForm.phone)
async def cargo_phone(message: Message, state: FSMContext, bot: Bot):
    if not message.from_user:
        return
    data = await state.get_data()
    await upsert_user(message)
    user = await get_user(message.from_user.id)
    partner_id = user["ref_partner_id"] if user else None
    order = await create_order(
        user_id=message.from_user.id,
        order_type="cargo",
        from_country=data["from_country"],
        to_city=data["to_city"],
        cargo_type=data["cargo_type"],
        weight=float(data["weight"]),
        volume=float(data["volume"]),
        description=data["description"],
        customer_name=data["customer_name"],
        phone=message.text or "",
        partner_id=partner_id,
    )
    rate, estimate = await estimate_delivery_price_db(data["from_country"], float(data["weight"]), float(data["volume"]))
    await state.clear()
    await message.answer(
        f"✅ Заявка создана.\n\n"
        f"Ваш номер груза: <b>{safe(order['tracking_code'])}</b>\n"
        f"Предварительная стоимость: около <b>{estimate} {safe(CURRENCY)}</b>\n"
        "Нажмите кнопку ниже, чтобы отследить груз.",
        reply_markup=track_button(order['tracking_code']),
    )
    await show_order_to_admin(bot, order)


@router.message(F.text == "🧮 Рассчитать доставку")
async def calc_start(message: Message, state: FSMContext):
    await state.clear()
    await state.set_state(CalcForm.from_country)
    await message.answer("Откуда груз? Китай, Турция, Дубай или Россия?")


@router.message(CalcForm.from_country)
async def calc_country(message: Message, state: FSMContext):
    await state.update_data(from_country=message.text)
    await state.set_state(CalcForm.weight)
    await message.answer("Вес в кг?")


@router.message(CalcForm.weight)
async def calc_weight(message: Message, state: FSMContext):
    try:
        weight = float((message.text or "").replace(",", "."))
        if weight <= 0:
            raise ValueError
    except ValueError:
        await message.answer("Введите вес числом. Например: 10")
        return
    await state.update_data(weight=weight)
    await state.set_state(CalcForm.volume)
    await message.answer("Объём в м³, если знаете. Если не знаете — напишите 0")


@router.message(CalcForm.volume)
async def calc_finish(message: Message, state: FSMContext):
    try:
        volume = float((message.text or "0").replace(",", "."))
    except ValueError:
        await message.answer("Введите объём числом. Если не знаете — 0")
        return
    data = await state.get_data()
    rate, estimate = await estimate_delivery_price_db(data["from_country"], float(data["weight"]), volume)
    await state.clear()
    await message.answer(
        f"🧮 <b>Предварительный расчёт</b>\n\n"
        f"Страна: {safe(data['from_country'])}\n"
        f"Вес: {safe(data['weight'])} кг\n"
        f"Объём: {safe(volume)} м³\n"
        f"Тариф: {rate} {safe(CURRENCY)}/кг\n"
        f"Примерная стоимость: <b>{estimate} {safe(CURRENCY)}</b>\n\n"
        "Это предварительный расчёт. Финальная цена зависит от фактического веса, категории товара и условий доставки.",
        reply_markup=client_keyboard(),
    )


@router.message(F.text == "🔎 Где мой груз?")
async def track_start(message: Message, state: FSMContext):
    await state.clear()
    await state.set_state(TrackForm.code)
    await message.answer("Введите номер груза. Например: CG26050300001")


@router.message(TrackForm.code)
async def track_finish(message: Message, state: FSMContext):
    code = parse_tracking_code(message.text or "") or (message.text or "").strip().upper()
    order = await get_order_by_code(code)
    await state.clear()
    if not order:
        await message.answer("Груз не найден. Проверьте номер или напишите менеджеру.", reply_markup=client_keyboard())
        return
    debt = max(float(order["price"] or 0) - float(order["paid_amount"] or 0), 0)
    await message.answer(format_order(order), reply_markup=track_pay_button(order['tracking_code'], debt))


@router.message(F.text == "💳 Оплатить доставку")
async def payment_start(message: Message, state: FSMContext):
    await state.clear()
    await state.set_state(PaymentForm.code)
    await message.answer(
        "Введите номер груза для оплаты. Например: CG26050300001\n\n"
        "Оплата открывается отдельно от страницы отслеживания."
    )


@router.message(PaymentForm.code)
async def payment_finish(message: Message, state: FSMContext):
    code = parse_tracking_code(message.text or "") or (message.text or "").strip().upper()
    await state.clear()
    order = await get_order_by_code(code)
    if not order:
        await message.answer("Груз не найден. Проверьте номер или напишите менеджеру.", reply_markup=client_keyboard())
        return
    order = await ensure_demo_price_if_needed(order)
    amount = demo_payment_amount_for_order(order)
    await message.answer(
        f"<b>💳 Оплата доставки</b>\n\n"
        f"Груз: <b>{safe(order['tracking_code'])}</b>\n"
        f"Сумма к оплате: <b>{amount:g} {safe(CURRENCY)}</b>\n\n"
        "Для демо доступны виртуальные варианты: Kaspi, Halyk, ЦентрКредит и Freedom. "
        "Реальные деньги не списываются.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="💳 Выбрать способ оплаты", url=payment_page_url(order["tracking_code"]))],
            [InlineKeyboardButton(text="🔎 Отследить груз", url=track_url(order["tracking_code"]))],
        ]),
    )


@router.message(F.text == "📋 Мои заказы")
async def my_orders(message: Message):
    if not message.from_user:
        return
    orders = await list_orders(limit=10, user_id=message.from_user.id)
    if not orders:
        await message.answer("У вас пока нет заказов.", reply_markup=client_keyboard())
        return
    text = "<b>📋 Ваши последние заказы</b>\n\n" + "\n".join(format_short_order(o) for o in orders)
    await message.answer(text, reply_markup=client_keyboard())


@router.message(F.text == "🛒 Выкуп товара")
async def buyout_start(message: Message, state: FSMContext):
    await state.clear()
    await state.set_state(BuyoutForm.link)
    await message.answer("Отправьте ссылку на товар или описание товара, который нужно выкупить.")


@router.message(BuyoutForm.link)
async def buyout_link(message: Message, state: FSMContext):
    await state.update_data(link=message.text)
    await state.set_state(BuyoutForm.details)
    await message.answer("Укажите детали: размер, цвет, количество, бюджет, комментарий.")


@router.message(BuyoutForm.details)
async def buyout_details(message: Message, state: FSMContext):
    await state.update_data(details=message.text)
    await state.set_state(BuyoutForm.customer_name)
    await message.answer("Ваше имя?")


@router.message(BuyoutForm.customer_name)
async def buyout_name(message: Message, state: FSMContext):
    await state.update_data(customer_name=message.text)
    await state.set_state(BuyoutForm.phone)
    await message.answer("Телефон или Telegram для связи?")


@router.message(BuyoutForm.phone)
async def buyout_finish(message: Message, state: FSMContext, bot: Bot):
    if not message.from_user:
        return
    data = await state.get_data()
    await upsert_user(message)
    user = await get_user(message.from_user.id)
    partner_id = user["ref_partner_id"] if user else None
    description = f"Ссылка/товар: {data['link']}\nДетали: {data['details']}"
    order = await create_order(
        user_id=message.from_user.id,
        order_type="buyout",
        from_country="Китай/маркетплейс",
        to_city="уточнить",
        cargo_type="выкуп товара",
        weight=0,
        volume=0,
        description=description,
        customer_name=data["customer_name"],
        phone=message.text or "",
        partner_id=partner_id,
    )
    await state.clear()
    await message.answer(
        f"✅ Заявка на выкуп создана.\n\n"
        f"Номер: <b>{safe(order['tracking_code'])}</b>\n"
        "Менеджер проверит товар, стоимость и доставку.\n"
        "Нажмите кнопку ниже, чтобы отследить заявку.",
        reply_markup=track_button(order['tracking_code']),
    )
    await show_order_to_admin(bot, order)


@router.message(F.text == "🏢 Оптовая доставка")
async def wholesale_start(message: Message, state: FSMContext):
    await state.clear()
    await state.set_state(WholesaleForm.from_country)
    await message.answer("Откуда партия? Например: Китай, Турция, Дубай")


@router.message(WholesaleForm.from_country)
async def wholesale_country(message: Message, state: FSMContext):
    await state.update_data(from_country=message.text)
    await state.set_state(WholesaleForm.to_city)
    await message.answer("Куда доставить партию?")


@router.message(WholesaleForm.to_city)
async def wholesale_city(message: Message, state: FSMContext):
    await state.update_data(to_city=message.text)
    await state.set_state(WholesaleForm.goods)
    await message.answer("Что за товар и какая партия?")


@router.message(WholesaleForm.goods)
async def wholesale_goods(message: Message, state: FSMContext):
    await state.update_data(goods=message.text)
    await state.set_state(WholesaleForm.boxes)
    await message.answer("Сколько коробок/мест примерно?")


@router.message(WholesaleForm.boxes)
async def wholesale_boxes(message: Message, state: FSMContext):
    await state.update_data(boxes=message.text)
    await state.set_state(WholesaleForm.weight)
    await message.answer("Общий вес примерно в кг?")


@router.message(WholesaleForm.weight)
async def wholesale_weight(message: Message, state: FSMContext):
    try:
        weight = float((message.text or "").replace(",", "."))
    except ValueError:
        await message.answer("Введите вес числом. Например: 250")
        return
    await state.update_data(weight=weight)
    await state.set_state(WholesaleForm.documents)
    await message.answer("Нужны документы/таможенное сопровождение? Напишите что есть: инвойс, упаковочный лист, договор и т.д.")


@router.message(WholesaleForm.documents)
async def wholesale_docs(message: Message, state: FSMContext):
    await state.update_data(documents=message.text)
    await state.set_state(WholesaleForm.customer_name)
    await message.answer("Ваше имя?")


@router.message(WholesaleForm.customer_name)
async def wholesale_name(message: Message, state: FSMContext):
    await state.update_data(customer_name=message.text)
    await state.set_state(WholesaleForm.phone)
    await message.answer("Телефон или Telegram для связи?")


@router.message(WholesaleForm.phone)
async def wholesale_finish(message: Message, state: FSMContext, bot: Bot):
    if not message.from_user:
        return
    data = await state.get_data()
    await upsert_user(message)
    user = await get_user(message.from_user.id)
    partner_id = user["ref_partner_id"] if user else None
    description = (
        f"Товар: {data['goods']}\n"
        f"Коробки/места: {data['boxes']}\n"
        f"Документы: {data['documents']}"
    )
    order = await create_order(
        user_id=message.from_user.id,
        order_type="wholesale",
        from_country=data["from_country"],
        to_city=data["to_city"],
        cargo_type="оптовая партия",
        weight=float(data["weight"]),
        volume=0,
        description=description,
        customer_name=data["customer_name"],
        phone=message.text or "",
        partner_id=partner_id,
    )
    await state.clear()
    await message.answer(
        f"✅ Заявка на оптовую доставку создана.\n\n"
        f"Номер: <b>{safe(order['tracking_code'])}</b>\n"
        "Менеджер проверит маршрут, документы и тариф.\n"
        "Нажмите кнопку ниже, чтобы отследить заявку.",
        reply_markup=track_button(order['tracking_code']),
    )
    await show_order_to_admin(bot, order)


@router.message(F.text == "⚠️ Жалоба / проблема")
async def complaint_start(message: Message, state: FSMContext):
    await state.clear()
    await state.set_state(ComplaintForm.code)
    await message.answer("Введите номер груза. Если номера нет — напишите 0")


@router.message(ComplaintForm.code)
async def complaint_code(message: Message, state: FSMContext):
    code = "" if (message.text or "").strip() == "0" else (parse_tracking_code(message.text or "") or (message.text or "").strip().upper())
    await state.update_data(code=code)
    await state.set_state(ComplaintForm.text)
    await message.answer("Опишите проблему коротко и понятно.")


@router.message(ComplaintForm.text)
async def complaint_text(message: Message, state: FSMContext):
    await state.update_data(text=message.text)
    await state.set_state(ComplaintForm.urgency)
    await message.answer("Срочность: низкая / средняя / высокая")


@router.message(ComplaintForm.urgency)
async def complaint_finish(message: Message, state: FSMContext, bot: Bot):
    if not message.from_user:
        return
    data = await state.get_data()
    complaint = await create_complaint(message.from_user.id, data.get("code", ""), data.get("text", ""), message.text or "средняя")
    await state.clear()
    await message.answer(
        f"✅ Жалоба №{complaint['id']} создана. Менеджер увидит её в админ‑панели.",
        reply_markup=client_keyboard(),
    )
    await notify_admins(
        bot,
        f"<b>⚠️ Новая жалоба №{complaint['id']}</b>\n"
        f"Груз: {safe(complaint['tracking_code'])}\n"
        f"Срочность: {safe(complaint['urgency'])}\n"
        f"Клиент ID: {complaint['user_id']}\n\n"
        f"{safe(complaint['text'])}",
        complaint_actions_keyboard(complaint["id"]),
    )



@router.message(F.text == "🛠️ Техподдержка")
async def support_start(message: Message, state: FSMContext):
    await state.clear()
    if message.from_user and is_admin(message.from_user.id):
        tickets = await list_open_support_tickets(20)
        if not tickets:
            await message.answer("Открытых обращений в техподдержку нет.", reply_markup=admin_keyboard())
            return
        await message.answer("<b>🛠️ Открытые обращения техподдержки</b>", reply_markup=admin_keyboard())
        for t in tickets:
            await message.answer(format_support_ticket(t) + f"\n\nОтветить: /replyticket {t['id']} ваш текст", reply_markup=support_ticket_actions_keyboard(t["id"]))
        return
    await message.answer(
        "<b>🛠️ Техподдержка</b>\n\n"
        "Здесь можно создать обращение, если нужен ответ менеджера: вопрос по грузу, оплате, фото, повреждению или работе бота.\n\n"
        "Выберите действие:",
        reply_markup=support_client_menu_keyboard(),
    )


@router.callback_query(F.data == "support_new")
async def cb_support_new(call: CallbackQuery, state: FSMContext):
    await state.clear()
    await state.set_state(SupportForm.code)
    await call.message.answer("Введите номер груза <code>CG...</code>. Если вопрос общий — напишите <b>0</b>.")
    await call.answer()


@router.message(SupportForm.code)
async def support_code(message: Message, state: FSMContext):
    code = "" if (message.text or "").strip() == "0" else (parse_tracking_code(message.text or "") or (message.text or "").strip().upper())
    await state.update_data(code=code)
    await state.set_state(SupportForm.topic)
    await message.answer("Выберите тему обращения:", reply_markup=support_topic_keyboard())


@router.callback_query(F.data.startswith("support_topic:"))
async def cb_support_topic(call: CallbackQuery, state: FSMContext):
    topic = call.data.split(":", 1)[1]
    await state.update_data(topic=topic)
    await state.set_state(SupportForm.text)
    await call.message.answer(
        f"Тема: <b>{safe(support_topic_label(topic))}</b>\n\n"
        "Опишите вопрос одним сообщением. Укажите детали, чтобы менеджер быстрее понял ситуацию."
    )
    await call.answer()


@router.message(SupportForm.text)
async def support_finish(message: Message, state: FSMContext, bot: Bot):
    if not message.from_user:
        return
    data = await state.get_data()
    await upsert_user(message)
    ticket = await create_support_ticket(
        message.from_user.id,
        data.get("code", ""),
        data.get("topic", "other"),
        message.text or "",
    )
    await state.clear()
    await message.answer(
        f"✅ Обращение №{ticket['id']} создано. Менеджер получил уведомление.\n\n"
        "Чтобы посмотреть свои обращения или дописать сообщение, нажмите «🛠️ Техподдержка» → «📋 Мои обращения».",
        reply_markup=client_keyboard(),
    )
    await notify_admins(
        bot,
        "<b>🛠️ Новое обращение в техподдержку</b>\n\n" + format_support_ticket(ticket),
        support_ticket_actions_keyboard(ticket["id"]),
    )


@router.callback_query(F.data == "support_my")
async def cb_support_my(call: CallbackQuery):
    tickets = await list_user_support_tickets(call.from_user.id, 10)
    if not tickets:
        await call.message.answer("У вас пока нет обращений в техподдержку.")
        await call.answer()
        return
    await call.message.answer("<b>📋 Ваши обращения</b>")
    for t in tickets:
        await call.message.answer(format_support_ticket(t), reply_markup=support_ticket_actions_keyboard(t["id"], include_client_reply=True))
    await call.answer()


@router.callback_query(F.data.startswith("support_client_reply:"))
async def cb_support_client_reply(call: CallbackQuery, state: FSMContext):
    ticket_id = int(call.data.split(":", 1)[1])
    ticket = await get_support_ticket(ticket_id)
    if not ticket or ticket["user_id"] != call.from_user.id:
        await call.answer("Это не ваше обращение", show_alert=True)
        return
    if ticket["status"] == "closed":
        await call.answer("Обращение уже закрыто", show_alert=True)
        return
    await state.clear()
    await state.update_data(ticket_id=ticket_id)
    await state.set_state(ClientSupportReplyForm.text)
    await call.message.answer(f"Напишите ответ в обращение №{ticket_id} одним сообщением.")
    await call.answer()


@router.message(ClientSupportReplyForm.text)
async def support_client_reply_finish(message: Message, state: FSMContext, bot: Bot):
    if not message.from_user:
        return
    data = await state.get_data()
    ticket_id = int(data["ticket_id"])
    ticket = await get_support_ticket(ticket_id)
    if not ticket or ticket["user_id"] != message.from_user.id:
        await state.clear()
        await message.answer("Обращение не найдено.", reply_markup=client_keyboard())
        return
    updated = await add_support_message(ticket_id, message.from_user.id, "client", message.text or "", "open")
    await state.clear()
    await message.answer(f"✅ Сообщение добавлено в обращение №{ticket_id}.", reply_markup=client_keyboard())
    await notify_admins(
        bot,
        "<b>🛠️ Новое сообщение в обращении</b>\n\n" + format_support_ticket(updated),
        support_ticket_actions_keyboard(ticket_id),
    )


@router.message(F.text == "❓ FAQ")
async def faq(message: Message):
    await message.answer(
        "<b>❓ FAQ</b>\n\n"
        "<b>Как узнать статус?</b> Нажмите «🔎 Где мой груз?» и введите номер CG...\n\n"
        "<b>Цена финальная?</b> Калькулятор даёт предварительный расчёт. Финальная цена зависит от фактического веса, объёма и категории товара.\n\n"
        "<b>Какие документы нужны?</b> Обычно нужны описание товара, количество, стоимость, инвойс/ссылка/упаковочный лист, если есть.\n\n"
        "<b>Что делать при проблеме?</b> Нажмите «⚠️ Жалоба / проблема», укажите номер груза и описание.",
        reply_markup=client_keyboard(),
    )


@router.message(F.text == "🤝 Партнёрка")
async def partner_menu(message: Message, bot: Bot):
    if not message.from_user:
        return
    await upsert_user(message)
    me = await bot.get_me()
    code = message.from_user.id
    link = f"https://t.me/{me.username}?start=partner_{code}"
    report = await partner_report(message.from_user.id)
    orders = report["orders"]
    commission = round((orders["revenue"] or 0) * DEFAULT_COMMISSION_PERCENT / 100, 2)
    await message.answer(
        f"<b>🤝 Партнёрская система</b>\n\n"
        f"Ваша ссылка:\n{safe(link)}\n\n"
        f"Приведено клиентов: {report['leads']['cnt']}\n"
        f"Заказов от ваших клиентов: {orders['cnt']}\n"
        f"Оборот: {orders['revenue']} {safe(CURRENCY)}\n"
        f"Примерная комиссия {DEFAULT_COMMISSION_PERCENT}%: {commission} {safe(CURRENCY)}",
        reply_markup=client_keyboard(),
    )


@router.message(F.text == "👤 Мой кабинет")
async def client_cabinet(message: Message):
    if not message.from_user:
        return
    await upsert_user(message)
    orders = await list_orders(limit=50, user_id=message.from_user.id)
    active = [o for o in orders if o["status"] != "доставлен"]
    debt = sum(max(float(o["price"] or 0) - float(o["paid_amount"] or 0), 0) for o in orders)
    await message.answer(
        f"<b>👤 Мой кабинет</b>\n\n"
        f"ID: {message.from_user.id}\n"
        f"Активных грузов: {len(active)}\n"
        f"Всего заказов: {len(orders)}\n"
        f"Долг: {debt} {safe(CURRENCY)}\n\n"
        "Для статуса нажмите «🔎 Где мой груз?». Для фото — «📷 Фото груза».",
        reply_markup=client_keyboard(),
    )


@router.message(F.text == "🕓 История статусов")
async def history_start(message: Message, state: FSMContext):
    await state.clear()
    await state.set_state(HistoryForm.code)
    await message.answer("Введите номер груза CG...")


@router.message(HistoryForm.code)
async def history_finish(message: Message, state: FSMContext):
    if not message.from_user:
        return
    code = parse_tracking_code(message.text or "") or (message.text or "").strip().upper()
    order = await get_order_by_code(code)
    await state.clear()
    if not order:
        await message.answer("Груз не найден.", reply_markup=client_keyboard())
        return
    if not is_admin(message.from_user.id) and order["user_id"] != message.from_user.id:
        await message.answer("⛔ Историю можно смотреть только по своему грузу.", reply_markup=client_keyboard())
        return
    rows = await get_status_history(order["id"])
    await message.answer(f"<b>🕓 История {safe(order['tracking_code'])}</b>\n\n" + format_history(rows), reply_markup=client_keyboard())


@router.message(F.text == "📷 Фото груза")
async def photo_view_start(message: Message, state: FSMContext):
    await state.clear()
    await state.set_state(PhotoViewForm.code)
    await message.answer("Введите номер груза CG...")


@router.message(PhotoViewForm.code)
async def photo_view_finish(message: Message, state: FSMContext):
    if not message.from_user:
        return
    code = parse_tracking_code(message.text or "") or (message.text or "").strip().upper()
    order = await get_order_by_code(code)
    await state.clear()
    if not order:
        await message.answer("Груз не найден.", reply_markup=client_keyboard())
        return
    if not is_admin(message.from_user.id) and order["user_id"] != message.from_user.id:
        await message.answer("⛔ Фото можно смотреть только по своему грузу.", reply_markup=client_keyboard())
        return
    photos = await list_cargo_photos(order["id"])
    if not photos:
        await message.answer("По этому грузу пока нет фото.", reply_markup=client_keyboard())
        return
    for ph in photos:
        await message.answer_photo(ph["file_id"], caption=f"📷 {safe(order['tracking_code'])}\n{safe(ph['comment'])}")


@router.message(F.text == "🔳 QR груза")
async def qr_start(message: Message, state: FSMContext):
    await state.clear()
    await state.set_state(QrForm.code)
    await message.answer("Введите номер груза CG...")


@router.message(QrForm.code)
async def qr_finish(message: Message, state: FSMContext, bot: Bot):
    if not message.from_user:
        return
    code = parse_tracking_code(message.text or "") or (message.text or "").strip().upper()
    order = await get_order_by_code(code)
    await state.clear()
    if not order:
        await message.answer("Груз не найден.", reply_markup=client_keyboard())
        return
    if not is_admin(message.from_user.id) and order["user_id"] != message.from_user.id:
        await message.answer("⛔ QR можно получить только по своему грузу.", reply_markup=client_keyboard())
        return
    me = await bot.get_me()
    payload = track_url(order['tracking_code'])
    await message.answer_photo(BufferedInputFile(make_qr_bytes(payload), filename=f"{order['tracking_code']}.png"), caption=f"🔳 QR для {safe(order['tracking_code'])}", reply_markup=client_keyboard())


@router.message(F.text == "💸 Долги/оплаты")
@router.message(F.text == "💳 Оплаты/долги")
async def admin_debts_button(message: Message):
    await cmd_debts(message)


@router.message(F.text == "📤 Excel экспорт")
async def admin_export_button(message: Message):
    await cmd_export_orders(message)


@router.message(F.text == "📥 Обновить из Excel")
async def excel_import_start(message: Message, state: FSMContext):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    await state.clear()
    await state.set_state(ExcelImportForm.file)
    await message.answer(
        "Отправьте .xlsx файл с колонками:\n"
        "tracking_code, status, comment, price, cost, paid_amount\n\n"
        "Обязательная колонка только tracking_code. Остальные можно не заполнять."
    )


@router.message(ExcelImportForm.file, F.document)
async def excel_import_finish(message: Message, state: FSMContext, bot: Bot):
    if not message.from_user or not is_admin(message.from_user.id) or not message.document:
        return
    if not (message.document.file_name or "").lower().endswith(".xlsx"):
        await message.answer("Нужен файл .xlsx")
        return
    path = EXPORT_DIR / f"import_{message.from_user.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    await bot.download(message.document, destination=path)
    ok, fail = await import_statuses_from_excel(path, message.from_user.id, bot)
    await state.clear()
    await message.answer(f"✅ Импорт завершён. Обновлено: {ok}, ошибок: {fail}", reply_markup=admin_keyboard())


@router.message(F.text == "⚙️ Тарифы")
async def tariffs_menu(message: Message, state: FSMContext):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    await state.clear()
    rows = await list_tariffs()
    lines = [f"— {safe(r['country_name'])}: {r['rate']} {safe(CURRENCY)}/кг" for r in rows]
    await message.answer(
        "<b>⚙️ Тарифы</b>\n\n" + ("\n".join(lines) if lines else "Тарифов пока нет.") +
        "\n\nВыберите действие кнопкой ниже.",
        reply_markup=tariffs_panel_keyboard(),
    )


@router.message(TariffForm.country)
async def tariff_country(message: Message, state: FSMContext):
    await state.update_data(country=message.text)
    await state.set_state(TariffForm.rate)
    await message.answer("Введите тариф за 1 кг. Например: 3.5")


@router.message(TariffForm.rate)
async def tariff_rate(message: Message, state: FSMContext):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    try:
        rate = float((message.text or "").replace(",", "."))
    except ValueError:
        await message.answer("Введите число. Например: 3.5")
        return
    data = await state.get_data()
    await set_tariff(data["country"], rate)
    await state.clear()
    await message.answer(f"✅ Тариф сохранён: {safe(data['country'])} — {rate} {safe(CURRENCY)}/кг", reply_markup=admin_keyboard())


@router.message(F.text == "👥 Роли")
async def roles_info(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    await message.answer(
        "<b>👥 Роли сотрудников</b>\n\n"
        "Выдать роль:\n"
        "/role 123456789 warehouse\n"
        "/role 123456789 courier\n"
        "/role 123456789 partner\n"
        "/role 123456789 client\n\n"
        "Склад: /warehouse\nКурьер: /courier",
        reply_markup=admin_keyboard(),
    )


@router.message(F.text == "📄 PDF квитанция")
async def receipt_start(message: Message, state: FSMContext):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    await state.clear()
    await state.set_state(ReceiptForm.code)
    await message.answer("Введите номер груза CG...")


@router.message(ReceiptForm.code)
async def receipt_finish(message: Message, state: FSMContext):
    code = parse_tracking_code(message.text or "") or (message.text or "").strip().upper()
    order = await get_order_by_code(code)
    await state.clear()
    if not order:
        await message.answer("Груз не найден.", reply_markup=admin_keyboard())
        return
    path = generate_receipt_pdf(order)
    await message.answer_document(FSInputFile(path), caption=f"📄 Квитанция {safe(order['tracking_code'])}", reply_markup=admin_keyboard())


@router.message(F.text == "🚚 Курьеры")
async def courier_assign_start(message: Message, state: FSMContext):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    await state.clear()
    await state.set_state(CourierAssignForm.code)
    await message.answer("Введите номер груза, который нужно передать курьеру.")


@router.message(CourierAssignForm.code)
async def courier_assign_code(message: Message, state: FSMContext):
    code = parse_tracking_code(message.text or "") or (message.text or "").strip().upper()
    order = await get_order_by_code(code)
    if not order:
        await message.answer("Груз не найден. Введите номер ещё раз.")
        return
    await state.update_data(code=code)
    await state.set_state(CourierAssignForm.courier_id)
    await message.answer("Введите Telegram ID курьера.")


@router.message(CourierAssignForm.courier_id)
async def courier_assign_id(message: Message, state: FSMContext):
    if not (message.text or "").lstrip("-").isdigit():
        await message.answer("ID должен быть числом.")
        return
    await state.update_data(courier_id=int(message.text))
    await state.set_state(CourierAssignForm.address)
    await message.answer("Введите адрес доставки или комментарий для курьера.")


@router.message(CourierAssignForm.address)
async def courier_assign_finish(message: Message, state: FSMContext, bot: Bot):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    data = await state.get_data()
    order = await assign_courier(data["code"], int(data["courier_id"]), message.text or "", message.from_user.id)
    await state.clear()
    if not order:
        await message.answer("Груз не найден.", reply_markup=admin_keyboard())
        return
    async with get_db() as db:
        cid = int(data["courier_id"])
        await db.execute(
            "INSERT OR IGNORE INTO users (telegram_id, username, full_name, role, partner_code, created_at, updated_at) VALUES (?, '', '', 'courier', ?, ?, ?)",
            (cid, partner_code(cid), now_iso(), now_iso()),
        )
        await db.execute("UPDATE users SET role='courier', updated_at=? WHERE telegram_id=?", (now_iso(), cid))
        await db.commit()
    await message.answer("✅ Курьер назначен.\n\n" + format_order(order), reply_markup=admin_keyboard())
    await notify_user(bot, order["user_id"], f"🚚 Ваш груз <b>{safe(order['tracking_code'])}</b> передан курьеру.")
    await notify_user(bot, int(data["courier_id"]), f"🚚 Вам назначена доставка: {safe(order['tracking_code'])}\nАдрес: {safe(message.text)}")


@router.message(F.text == "🚚 Курьерское меню")
async def show_courier_menu(message: Message, state: FSMContext):
    await cmd_courier(message, state)


@router.message(F.text == "🚚 Мои доставки")
async def courier_my_deliveries(message: Message):
    if not message.from_user or not (is_courier_static(message.from_user.id) or await has_role(message.from_user.id, "courier")):
        return
    rows = await courier_orders(message.from_user.id)
    if not rows:
        await message.answer("У вас нет активных доставок.", reply_markup=courier_keyboard())
        return
    text = "<b>🚚 Мои доставки</b>\n\n" + "\n".join(
        f"{safe(o['tracking_code'])} | {safe(o['delivery_address'])} | {safe(o['phone'])}" for o in rows
    )
    await message.answer(text, reply_markup=courier_keyboard())


@router.message(F.text == "✅ Отметить доставлено")
async def courier_delivered_start(message: Message, state: FSMContext):
    if not message.from_user or not (is_courier_static(message.from_user.id) or await has_role(message.from_user.id, "courier")):
        return
    await state.clear()
    await state.set_state(CourierDeliveredForm.code)
    await message.answer("Введите номер груза CG...")


@router.message(CourierDeliveredForm.code)
async def courier_delivered_finish(message: Message, state: FSMContext, bot: Bot):
    if not message.from_user:
        return
    code = parse_tracking_code(message.text or "") or (message.text or "").strip().upper()
    order = await get_order_by_code(code)
    await state.clear()
    if not order or int(order["courier_id"] or 0) != message.from_user.id:
        await message.answer("Груз не найден среди ваших доставок.", reply_markup=courier_keyboard())
        return
    updated = await update_order_status(order["id"], "доставлен", message.from_user.id, "Курьер отметил доставку")
    await message.answer("✅ Доставка отмечена.\n\n" + format_order(updated), reply_markup=courier_keyboard())
    await notify_user(bot, updated["user_id"], f"✅ Ваш груз <b>{safe(updated['tracking_code'])}</b> доставлен.")


# =========================
# ADMIN / WAREHOUSE MENUS
# =========================
@router.message(F.text == "🏭 Меню склада")
async def show_warehouse_menu(message: Message, state: FSMContext):
    await state.clear()
    if not message.from_user or not is_warehouse(message.from_user.id):
        await message.answer("⛔ Нет доступа.")
        return
    await message.answer(warehouse_menu_text(), reply_markup=warehouse_keyboard())


@router.message(F.text == "📥 Новые заявки")
async def admin_new_orders(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    orders = await list_orders(limit=10, status="новая заявка")
    if not orders:
        await message.answer("Новых заявок нет.", reply_markup=admin_keyboard())
        return
    for order in orders:
        await message.answer(format_order(order), reply_markup=order_actions_keyboard(order["id"]))


@router.message(F.text == "📦 Все грузы")
async def admin_all_orders(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    orders = await list_orders(limit=15)
    if not orders:
        await message.answer("Грузов пока нет.", reply_markup=admin_keyboard())
        return
    text = "<b>📦 Последние грузы</b>\n\n" + "\n".join(format_short_order(o) for o in orders)
    await message.answer(text, reply_markup=admin_keyboard())


@router.message(F.text == "🔎 Найти груз")
async def admin_find_start(message: Message, state: FSMContext):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    await state.clear()
    await state.set_state(TrackForm.code)
    await message.answer("Введите номер груза CG...")


@router.message(F.text == "🔁 Изменить статус")
async def status_start(message: Message, state: FSMContext):
    if not message.from_user or not (is_admin(message.from_user.id) or is_warehouse(message.from_user.id)):
        return
    await state.clear()
    await state.set_state(StatusForm.code)
    await message.answer("Введите номер груза CG...")


@router.message(StatusForm.code)
async def status_code(message: Message, state: FSMContext):
    code = parse_tracking_code(message.text or "") or (message.text or "").strip().upper()
    order = await get_order_by_code(code)
    if not order:
        await message.answer("Груз не найден. Введите номер ещё раз или /start для отмены.")
        return
    await state.update_data(order_id=order["id"], code=code)
    await state.set_state(StatusForm.status)
    statuses = await get_all_statuses()
    await message.answer("Выберите статус:", reply_markup=status_keyboard(order["id"], statuses))


@router.message(StatusForm.status)
async def status_manual_status(message: Message, state: FSMContext):
    await state.update_data(status=message.text)
    await state.set_state(StatusForm.comment)
    await message.answer("Комментарий к статусу. Если не нужен — напишите 0")


@router.message(StatusForm.comment)
async def status_manual_finish(message: Message, state: FSMContext, bot: Bot):
    if not message.from_user:
        return
    data = await state.get_data()
    comment = "" if (message.text or "").strip() == "0" else (message.text or "")
    updated = await update_order_status(data["order_id"], data["status"], message.from_user.id, comment)
    await state.clear()
    await message.answer("✅ Статус обновлён.\n\n" + format_order(updated), reply_markup=admin_keyboard() if is_admin(message.from_user.id) else warehouse_keyboard())
    await notify_user(
        bot,
        updated["user_id"],
        status_notify_text(updated["tracking_code"], updated["status"], comment),
        reply_markup=track_button(updated["tracking_code"]),
    )


@router.message(F.text == "💬 Жалобы")
async def admin_complaints(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    complaints = await list_open_complaints(10)
    if not complaints:
        await message.answer("Открытых жалоб нет.", reply_markup=admin_keyboard())
        return
    for c in complaints:
        await message.answer(
            f"<b>⚠️ Жалоба №{c['id']}</b>\n"
            f"Груз: {safe(c['tracking_code'])}\n"
            f"Клиент ID: {c['user_id']}\n"
            f"Срочность: {safe(c['urgency'])}\n"
            f"Статус: {safe(c['status'])}\n\n"
            f"{safe(c['text'])}\n\n"
            f"Ответить: /replycomplaint {c['id']} ваш текст",
            reply_markup=complaint_actions_keyboard(c["id"]),
        )



@router.message(F.text == "🛠️ Техподдержка")
async def admin_support_menu(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    tickets = await list_open_support_tickets(20)
    if not tickets:
        await message.answer("Открытых обращений в техподдержку нет.", reply_markup=admin_keyboard())
        return
    await message.answer("<b>🛠️ Открытые обращения техподдержки</b>", reply_markup=admin_keyboard())
    for t in tickets:
        await message.answer(format_support_ticket(t) + f"\n\nОтветить: /replyticket {t['id']} ваш текст", reply_markup=support_ticket_actions_keyboard(t["id"]))


@router.message(F.text == "🚨 Проблемные грузы")
async def admin_problem_radar(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    radar = await smartflow_problem_radar(20)
    await message.answer(format_problem_radar(radar), reply_markup=admin_keyboard())


@router.message(F.text == "📊 SmartFlow отчёт")
async def admin_smartflow_report(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    await message.answer(await owner_smartflow_report_text(), reply_markup=admin_keyboard())


@router.message(F.text == "🧭 Promise OS")
@router.message(F.text == "📊 SmartFlow отчёт")
async def admin_promise_os(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    data = await promise_os_leak_map(12)
    await message.answer(format_promise_os_report(data), reply_markup=admin_keyboard())


@router.message(F.text == "💰 Финансы")
@router.message(F.text == "📊 Отчёт за день")
async def admin_finance(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    report = await finance_report()
    totals = report["totals"]
    today = report["today"]
    status_lines = "\n".join(f"— {safe(r['status'])}: {r['cnt']}" for r in report["by_status"])
    await message.answer(
        f"<b>📊 Отчёт</b>\n\n"
        f"<b>Сегодня:</b>\n"
        f"Заявок: {today['today_orders']}\n"
        f"Оборот: {today['today_revenue']} {safe(CURRENCY)}\n"
        f"Себестоимость: {today['today_cost']} {safe(CURRENCY)}\n"
        f"Маржа: {today['today_margin']} {safe(CURRENCY)}\n\n"
        f"<b>Всего:</b>\n"
        f"Заявок: {totals['total_orders']}\n"
        f"Оборот: {totals['revenue']} {safe(CURRENCY)}\n"
        f"Себестоимость: {totals['cost']} {safe(CURRENCY)}\n"
        f"Маржа: {totals['margin']} {safe(CURRENCY)}\n\n"
        f"<b>По статусам:</b>\n{status_lines or 'нет данных'}\n\n"
        "Чтобы внести финансы по грузу: /setprice CG26050300001 100 70",
        reply_markup=admin_keyboard(),
    )


@router.message(F.text == "🤝 Партнёры")
async def admin_partners(message: Message):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    report = await partner_report(message.from_user.id)
    top = report["partners"]
    if not top:
        await message.answer("Партнёрских лидов пока нет.", reply_markup=admin_keyboard())
        return
    lines = [f"Партнёр {r['partner_id']}: {r['cnt']} клиентов" for r in top]
    await message.answer("<b>🤝 Партнёры</b>\n\n" + "\n".join(lines), reply_markup=admin_keyboard())


@router.message(F.text == "📢 Рассылка")
async def broadcast_start(message: Message, state: FSMContext):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    await state.clear()
    await message.answer(
        "<b>📢 Рассылка</b>\n\nВыберите аудиторию кнопкой ниже. Потом отправьте текст сообщения.",
        reply_markup=broadcast_audience_keyboard(),
    )


@router.callback_query(F.data.startswith("broadcast_audience:"))
async def cb_broadcast_audience(call: CallbackQuery, state: FSMContext):
    if not call.from_user or not is_admin(call.from_user.id):
        await call.answer("Нет доступа", show_alert=True)
        return
    audience = call.data.split(":", 1)[1]
    await state.clear()
    await state.update_data(audience=audience)
    await state.set_state(BroadcastForm.text)
    labels = {"all": "всем клиентам", "active": "клиентам с активными грузами", "debtors": "должникам", "partners": "партнёрам"}
    await call.message.answer(f"Аудитория: <b>{safe(labels.get(audience, audience))}</b>.\n\nТеперь отправьте текст рассылки одним сообщением. Для отмены: /admin")
    await call.answer()


@router.message(BroadcastForm.text)
async def broadcast_finish(message: Message, state: FSMContext, bot: Bot):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    text = message.text or ""
    data = await state.get_data()
    audience = data.get("audience", "all")
    async with get_db() as db:
        if audience == "active":
            users = await db_fetchall(db, "SELECT DISTINCT user_id AS telegram_id FROM orders WHERE status NOT IN ('доставлен','отменён')")
        elif audience == "debtors":
            users = await db_fetchall(db, "SELECT DISTINCT user_id AS telegram_id FROM orders WHERE price > COALESCE(paid_amount,0)")
        elif audience == "partners":
            users = await db_fetchall(db, "SELECT telegram_id FROM users WHERE role='partner'")
        else:
            users = await db_fetchall(db, "SELECT telegram_id FROM users WHERE role IN ('client', 'partner')")
    sent = 0
    failed = 0
    for user in users:
        try:
            await bot.send_message(user["telegram_id"], text)
            sent += 1
            await asyncio.sleep(0.04)
        except Exception:
            failed += 1
    await state.clear()
    await message.answer(f"✅ Рассылка завершена. Отправлено: {sent}, ошибок: {failed}", reply_markup=admin_keyboard())



@router.message(F.text == "📦 Партии")
async def batches_menu(message: Message, state: FSMContext):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    await state.clear()
    batches = await list_batches(8)
    text = (
        "<b>📦 Партии грузов</b>\n\n"
        "Партии нужны, чтобы не менять статус каждому грузу отдельно. "
        "Создайте партию, добавьте в неё номера CG..., потом меняйте статус всей партии одной кнопкой."
    )
    await message.answer(text, reply_markup=batch_panel_keyboard(batches))


@router.callback_query(F.data == "batch_create")
async def cb_batch_create(call: CallbackQuery, state: FSMContext):
    if not call.from_user or not is_admin(call.from_user.id):
        await call.answer("Нет доступа", show_alert=True)
        return
    await state.clear()
    await state.set_state(BatchCreateForm.name)
    await call.message.answer("Введите название партии. Например: Китай 05.05 или Турция май")
    await call.answer()


@router.message(BatchCreateForm.name)
async def batch_create_name(message: Message, state: FSMContext):
    await state.update_data(name=message.text)
    await state.set_state(BatchCreateForm.route)
    await message.answer(
        "Введите направление партии:\n"
        "1 — Китай → СНГ\n2 — Турция → СНГ\n3 — Дубай/ОАЭ → СНГ\n4 — СНГ → СНГ\n\n"
        "Можно также написать: china, turkey, uae, cis"
    )


def parse_route_choice(text: str) -> str:
    t = normalize(text)
    if t in {"1", "china", "китай", "china_cis"}:
        return "china_cis"
    if t in {"2", "turkey", "турция", "turkey_cis"}:
        return "turkey_cis"
    if t in {"3", "uae", "дубай", "оаэ", "uae_cis"}:
        return "uae_cis"
    return "cis_local"


@router.message(BatchCreateForm.route)
async def batch_create_route(message: Message, state: FSMContext):
    if not message.from_user:
        return
    data = await state.get_data()
    route_key = parse_route_choice(message.text or "")
    batch = await create_batch(data["name"], route_key, message.from_user.id)
    await state.update_data(batch_id=batch["id"])
    await state.set_state(BatchCreateForm.codes)
    await message.answer(
        f"✅ Партия создана: <b>#{batch['id']} {safe(batch['name'])}</b>\n"
        f"Маршрут: {safe(auto_route_label(route_key))}\n\n"
        "Теперь отправьте номера грузов CG... через пробел, запятую или с новой строки.\n"
        "Если хотите добавить позже — напишите 0."
    )


@router.message(BatchCreateForm.codes)
async def batch_create_codes(message: Message, state: FSMContext):
    data = await state.get_data()
    batch_id = int(data["batch_id"])
    if (message.text or "").strip() == "0":
        await state.clear()
        batch = await get_batch(batch_id)
        await message.answer(format_batch(batch, 0), reply_markup=batch_actions_keyboard(batch_id))
        return
    ok, fail = await add_orders_to_batch(batch_id, message.text or "")
    orders = await get_batch_orders(batch_id)
    batch = await get_batch(batch_id)
    await state.clear()
    await message.answer(
        f"✅ В партию добавлено: {ok}. Не найдено: {fail}.\n\n" + format_batch(batch, len(orders)),
        reply_markup=batch_actions_keyboard(batch_id),
    )


@router.callback_query(F.data == "batch_list")
async def cb_batch_list(call: CallbackQuery):
    if not call.from_user or not is_admin(call.from_user.id):
        await call.answer("Нет доступа", show_alert=True)
        return
    batches = await list_batches(10)
    if not batches:
        await call.message.answer("Партии пока не созданы.")
    else:
        await call.message.answer("<b>📋 Активные партии</b>", reply_markup=batch_panel_keyboard(batches))
    await call.answer()


@router.callback_query(F.data.startswith("batch_open:"))
async def cb_batch_open(call: CallbackQuery):
    if not call.from_user or not is_admin(call.from_user.id):
        await call.answer("Нет доступа", show_alert=True)
        return
    batch_id = int(call.data.split(":", 1)[1])
    batch = await get_batch(batch_id)
    if not batch:
        await call.answer("Партия не найдена", show_alert=True)
        return
    orders = await get_batch_orders(batch_id)
    await call.message.answer(format_batch(batch, len(orders)), reply_markup=batch_actions_keyboard(batch_id))
    await call.answer()


@router.callback_query(F.data == "batch_add_orders")
async def cb_batch_add_choose(call: CallbackQuery, state: FSMContext):
    if not call.from_user or not is_admin(call.from_user.id):
        await call.answer("Нет доступа", show_alert=True)
        return
    await state.clear()
    await state.set_state(BatchAddForm.batch_id)
    await call.message.answer("Введите номер партии. Например: 1")
    await call.answer()


@router.callback_query(F.data.startswith("batch_add_to:"))
async def cb_batch_add_to(call: CallbackQuery, state: FSMContext):
    if not call.from_user or not is_admin(call.from_user.id):
        await call.answer("Нет доступа", show_alert=True)
        return
    batch_id = int(call.data.split(":", 1)[1])
    await state.clear()
    await state.update_data(batch_id=batch_id)
    await state.set_state(BatchAddForm.codes)
    await call.message.answer("Отправьте номера грузов CG... через пробел, запятую или с новой строки.")
    await call.answer()


@router.message(BatchAddForm.batch_id)
async def batch_add_id(message: Message, state: FSMContext):
    if not (message.text or "").strip().isdigit():
        await message.answer("Введите только номер партии. Например: 1")
        return
    batch = await get_batch(int(message.text.strip()))
    if not batch:
        await message.answer("Партия не найдена. Введите другой номер.")
        return
    await state.update_data(batch_id=batch["id"])
    await state.set_state(BatchAddForm.codes)
    await message.answer("Теперь отправьте номера грузов CG... через пробел, запятую или с новой строки.")


@router.message(BatchAddForm.codes)
async def batch_add_codes(message: Message, state: FSMContext):
    data = await state.get_data()
    batch_id = int(data["batch_id"])
    ok, fail = await add_orders_to_batch(batch_id, message.text or "")
    orders = await get_batch_orders(batch_id)
    batch = await get_batch(batch_id)
    await state.clear()
    await message.answer(
        f"✅ Добавлено в партию: {ok}. Не найдено: {fail}.\n\n" + format_batch(batch, len(orders)),
        reply_markup=batch_actions_keyboard(batch_id),
    )


@router.callback_query(F.data == "batch_change_status")
async def cb_batch_status_choose(call: CallbackQuery, state: FSMContext):
    if not call.from_user or not is_admin(call.from_user.id):
        await call.answer("Нет доступа", show_alert=True)
        return
    await state.clear()
    await state.set_state(BatchStatusForm.batch_id)
    await call.message.answer("Введите номер партии, которой нужно сменить статус.")
    await call.answer()


@router.callback_query(F.data.startswith("batch_status_for:"))
async def cb_batch_status_for(call: CallbackQuery):
    if not call.from_user or not is_admin(call.from_user.id):
        await call.answer("Нет доступа", show_alert=True)
        return
    batch_id = int(call.data.split(":", 1)[1])
    statuses = await get_all_statuses()
    await call.message.answer("Выберите новый статус для всей партии:", reply_markup=batch_status_keyboard(batch_id, statuses))
    await call.answer()


@router.message(BatchStatusForm.batch_id)
async def batch_status_id(message: Message, state: FSMContext):
    if not (message.text or "").strip().isdigit():
        await message.answer("Введите только номер партии. Например: 1")
        return
    batch_id = int(message.text.strip())
    batch = await get_batch(batch_id)
    if not batch:
        await message.answer("Партия не найдена. Введите другой номер.")
        return
    await state.clear()
    statuses = await get_all_statuses()
    await message.answer("Выберите новый статус для всей партии:", reply_markup=batch_status_keyboard(batch_id, statuses))


@router.callback_query(F.data.startswith("batch_set_status:"))
async def cb_batch_set_status(call: CallbackQuery, bot: Bot):
    if not call.from_user or not is_admin(call.from_user.id):
        await call.answer("Нет доступа", show_alert=True)
        return
    _, batch_id_raw, status = call.data.split(":", 2)
    batch_id = int(batch_id_raw)
    batch, orders = await update_batch_status(batch_id, status, call.from_user.id)
    if not batch:
        await call.answer("Партия не найдена", show_alert=True)
        return
    for order in orders:
        await notify_user(
            bot,
            order["user_id"],
            status_notify_text(order["tracking_code"], status, "Статус обновлён по партии."),
            reply_markup=track_button(order["tracking_code"]),
        )
        await asyncio.sleep(0.04)
    await call.message.answer(f"✅ Статус партии #{batch_id} изменён на <b>{safe(status)}</b>.\nУведомлено грузов: {len(orders)}")
    await call.answer("Готово")


@router.callback_query(F.data.startswith("batch_orders:"))
async def cb_batch_orders(call: CallbackQuery):
    if not call.from_user or not is_admin(call.from_user.id):
        await call.answer("Нет доступа", show_alert=True)
        return
    batch_id = int(call.data.split(":", 1)[1])
    orders = await get_batch_orders(batch_id)
    if not orders:
        await call.message.answer("В этой партии пока нет грузов.")
    else:
        await call.message.answer("<b>📋 Грузы в партии</b>\n\n" + "\n".join(format_short_order(o) for o in orders[:30]))
    await call.answer()


@router.message(F.text == "👥 Клиенты")
async def clients_menu(message: Message, state: FSMContext):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    await state.clear()
    rows = await list_clients(10)
    text = "<b>👥 База клиентов</b>\n\n" + ("\n\n".join(format_client_row(r) for r in rows) if rows else "Клиентов пока нет.")
    await state.set_state(ClientSearchForm.query)
    await message.answer(text + "\n\nЧтобы найти клиента, отправьте имя, username или Telegram ID. Для выхода: /admin")


@router.message(ClientSearchForm.query)
async def clients_search(message: Message, state: FSMContext):
    query = (message.text or "").strip()
    rows = await list_clients(10, query=query)
    if not rows:
        await message.answer("Клиенты не найдены. Попробуйте другой запрос или /admin для выхода.")
        return
    await message.answer("<b>🔎 Найденные клиенты</b>\n\n" + "\n\n".join(format_client_row(r) for r in rows))


@router.message(F.text == "⚙️ Статусы")
async def statuses_settings_menu(message: Message, state: FSMContext):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    await state.clear()
    custom = await get_custom_statuses()
    all_statuses = await get_all_statuses()
    await message.answer(
        "<b>⚙️ Статусы</b>\n\n"
        "Здесь можно добавить свои статусы под компанию. Они появятся в кнопках смены статуса.\n\n"
        + "\n".join(f"— {safe(st)}" for st in all_statuses),
        reply_markup=statuses_panel_keyboard(custom),
    )


@router.callback_query(F.data == "status_add")
async def cb_status_add(call: CallbackQuery, state: FSMContext):
    if not call.from_user or not is_admin(call.from_user.id):
        await call.answer("Нет доступа", show_alert=True)
        return
    await state.clear()
    await state.set_state(StatusSettingsForm.name)
    await call.message.answer("Напишите новый статус. Например: принят в Гуанчжоу")
    await call.answer()


@router.message(StatusSettingsForm.name)
async def status_add_finish(message: Message, state: FSMContext):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    await add_custom_status(message.text or "")
    await state.clear()
    await message.answer("✅ Статус добавлен. Он появится в меню смены статуса.", reply_markup=admin_keyboard())


@router.callback_query(F.data == "status_list")
async def cb_status_list(call: CallbackQuery):
    if not call.from_user or not is_admin(call.from_user.id):
        await call.answer("Нет доступа", show_alert=True)
        return
    statuses = await get_all_statuses()
    await call.message.answer("<b>📋 Все статусы</b>\n\n" + "\n".join(f"— {safe(st)}" for st in statuses))
    await call.answer()


@router.callback_query(F.data.startswith("status_delete:"))
async def cb_status_delete(call: CallbackQuery):
    if not call.from_user or not is_admin(call.from_user.id):
        await call.answer("Нет доступа", show_alert=True)
        return
    status_id = int(call.data.split(":", 1)[1])
    await delete_custom_status(status_id)
    custom = await get_custom_statuses()
    await call.message.answer("✅ Статус удалён.", reply_markup=statuses_panel_keyboard(custom))
    await call.answer()


@router.callback_query(F.data == "tariff_add")
async def cb_tariff_add(call: CallbackQuery, state: FSMContext):
    if not call.from_user or not is_admin(call.from_user.id):
        await call.answer("Нет доступа", show_alert=True)
        return
    await state.clear()
    await state.set_state(TariffForm.country)
    await call.message.answer("Напишите страну/направление для тарифа. Например: Китай")
    await call.answer()


@router.callback_query(F.data == "tariff_list")
async def cb_tariff_list(call: CallbackQuery):
    rows = await list_tariffs()
    lines = [f"— {safe(r['country_name'])}: {r['rate']} {safe(CURRENCY)}/кг" for r in rows]
    await call.message.answer("<b>⚙️ Тарифы</b>\n\n" + ("\n".join(lines) if lines else "Тарифов пока нет."))
    await call.answer()


@router.callback_query(F.data.startswith("pay_remind:"))
async def cb_pay_remind(call: CallbackQuery, bot: Bot):
    if not call.from_user or not is_admin(call.from_user.id):
        await call.answer("Нет доступа", show_alert=True)
        return
    order_id = int(call.data.split(":", 1)[1])
    order = await get_order(order_id)
    if not order:
        await call.answer("Груз не найден", show_alert=True)
        return
    debt = max(float(order["price"] or 0) - float(order["paid_amount"] or 0), 0)
    if debt <= 0:
        await call.answer("Долга нет", show_alert=True)
        return
    await notify_user(
        bot,
        order["user_id"],
        f"💰 Напоминание об оплате по грузу <b>{safe(order['tracking_code'])}</b>.\n"
        f"К оплате осталось: <b>{debt} {safe(CURRENCY)}</b>.\n"
        "Если вы уже оплатили, отправьте подтверждение менеджеру или в техподдержку.",
    )
    await call.answer("Напоминание отправлено")
    await call.message.answer(f"✅ Напоминание отправлено клиенту по грузу {safe(order['tracking_code'])}.")


@router.callback_query(F.data.startswith("support_tpl:"))
async def cb_support_template(call: CallbackQuery, bot: Bot):
    if not call.from_user or not is_admin(call.from_user.id):
        await call.answer("Нет доступа", show_alert=True)
        return
    _, ticket_id_raw, key = call.data.split(":", 2)
    ticket_id = int(ticket_id_raw)
    text = SUPPORT_REPLY_TEMPLATES.get(key)
    if not text:
        await call.answer("Шаблон не найден", show_alert=True)
        return
    ticket = await add_support_message(ticket_id, call.from_user.id, "admin", text, "answered")
    if not ticket:
        await call.answer("Обращение не найдено", show_alert=True)
        return
    await notify_user(bot, ticket["user_id"], f"🛠️ Ответ техподдержки по обращению №{ticket_id}:\n\n{safe(text)}")
    await call.answer("Ответ отправлен")
    await call.message.answer(f"✅ Быстрый ответ отправлен по обращению №{ticket_id}.")


@router.message(F.text == "📸 Добавить фото")
async def warehouse_photo_start(message: Message, state: FSMContext):
    if not message.from_user or not is_warehouse(message.from_user.id):
        return
    await state.clear()
    await state.set_state(PhotoAddForm.code)
    await message.answer("Введите номер груза CG...")


@router.message(PhotoAddForm.code)
async def warehouse_photo_code(message: Message, state: FSMContext):
    code = parse_tracking_code(message.text or "") or (message.text or "").strip().upper()
    order = await get_order_by_code(code)
    if not order:
        await message.answer("Груз не найден. Введите номер ещё раз.")
        return
    await state.update_data(code=code, order_id=order["id"])
    await state.set_state(PhotoAddForm.photo)
    await message.answer("Отправьте фото груза. В подписи можно написать комментарий: склад, упаковка, повреждение и т.д.")


@router.message(PhotoAddForm.photo, F.photo)
async def warehouse_photo_finish(message: Message, state: FSMContext, bot: Bot):
    if not message.from_user or not is_warehouse(message.from_user.id) or not message.photo:
        return
    data = await state.get_data()
    order = await get_order(data["order_id"])
    if not order:
        await state.clear()
        await message.answer("Груз не найден.", reply_markup=warehouse_keyboard())
        return
    file_id = message.photo[-1].file_id
    comment = message.caption or "Фото груза"
    await add_cargo_photo(order["id"], file_id, comment, message.from_user.id)
    await state.clear()
    await message.answer("✅ Фото добавлено к грузу.", reply_markup=warehouse_keyboard())
    await notify_user(bot, order["user_id"], f"📷 По вашему грузу <b>{safe(order['tracking_code'])}</b> добавлено новое фото.")


@router.message(PhotoAddForm.photo)
async def warehouse_photo_not_photo(message: Message):
    await message.answer("Отправьте именно фото, не файл. Можно добавить подпись к фото.")


@router.message(F.text == "➕ Принять груз")
@router.message(F.text == "⚖️ Указать вес")
async def warehouse_weight_start(message: Message, state: FSMContext):
    if not message.from_user or not is_warehouse(message.from_user.id):
        return
    await state.clear()
    await state.set_state(WeightForm.code)
    await message.answer("Введите номер груза CG...")


@router.message(WeightForm.code)
async def warehouse_weight_code(message: Message, state: FSMContext):
    code = parse_tracking_code(message.text or "") or (message.text or "").strip().upper()
    order = await get_order_by_code(code)
    if not order:
        await message.answer("Груз не найден. Введите номер ещё раз.")
        return
    await state.update_data(code=code)
    await state.set_state(WeightForm.weight)
    await message.answer("Укажите фактический вес в кг.")


@router.message(WeightForm.weight)
async def warehouse_weight_finish(message: Message, state: FSMContext, bot: Bot):
    if not message.from_user:
        return
    try:
        weight = float((message.text or "").replace(",", "."))
    except ValueError:
        await message.answer("Введите вес числом. Например: 8.4")
        return
    data = await state.get_data()
    updated = await set_order_weight(data["code"], weight, message.from_user.id)
    if not updated:
        await message.answer("Груз не найден.")
        await state.clear()
        return
    updated = await update_order_status(updated["id"], "принят на склад", message.from_user.id, f"Вес: {weight} кг")
    await state.clear()
    await message.answer("✅ Груз принят/вес обновлён.\n\n" + format_order(updated), reply_markup=warehouse_keyboard())
    await notify_user(
        bot,
        updated["user_id"],
        f"📦 Ваш груз <b>{safe(updated['tracking_code'])}</b> принят на склад.\nФактический вес: <b>{weight} кг</b>.",
    )


@router.message(F.text == "📦 Грузы на складе")
async def warehouse_orders(message: Message):
    if not message.from_user or not is_warehouse(message.from_user.id):
        return
    rows = await orders_on_warehouse()
    if not rows:
        await message.answer("На складе пока нет активных грузов.", reply_markup=warehouse_keyboard())
        return
    text = "<b>📦 Грузы на складе</b>\n\n" + "\n".join(format_short_order(o) for o in rows)
    await message.answer(text, reply_markup=warehouse_keyboard())


# =========================
# CALLBACKS
# =========================
@router.callback_query(F.data == "admin_back")
async def cb_admin_back(call: CallbackQuery):
    if not call.from_user or not is_admin(call.from_user.id):
        await call.answer("Нет доступа", show_alert=True)
        return
    await call.message.answer(admin_menu_text(), reply_markup=admin_keyboard())
    await call.answer()


@router.callback_query(F.data == "auto_enable_all")
async def cb_auto_enable_all(call: CallbackQuery):
    if not call.from_user or not is_admin(call.from_user.id):
        await call.answer("Нет доступа", show_alert=True)
        return
    count = await enable_auto_status_for_active_orders(call.from_user.id)
    orders = await list_orders(limit=10)
    await call.message.answer(
        f"✅ Автостатусы включены для активных грузов: <b>{count}</b>.\n\n"
        "Бот сам выбрал маршруты по направлениям и будет менять статусы по расписанию. "
        "Клиенты получат уведомления автоматически.",
        reply_markup=auto_status_panel_keyboard(orders),
    )
    await call.answer("Готово")


@router.callback_query(F.data == "auto_show_disabled")
async def cb_auto_show_disabled(call: CallbackQuery):
    if not call.from_user or not is_admin(call.from_user.id):
        await call.answer("Нет доступа", show_alert=True)
        return
    orders = await list_orders_without_auto_status(limit=10)
    if not orders:
        await call.message.answer(
            "✅ У всех активных грузов уже включены автостатусы.",
            reply_markup=auto_status_panel_keyboard(await list_orders(limit=10)),
        )
    else:
        await call.message.answer(
            "<b>📋 Грузы без автостатусов</b>\n\n"
            "Нажмите на нужный груз — дальше можно включить автостатусы одной кнопкой.",
            reply_markup=auto_status_panel_keyboard(orders),
        )
    await call.answer()


@router.callback_query(F.data.startswith("auto_enable:"))
async def cb_auto_enable(call: CallbackQuery):
    if not call.from_user or not is_admin(call.from_user.id):
        await call.answer("Нет доступа", show_alert=True)
        return
    order_id = int(call.data.split(":", 1)[1])
    order = await get_order(order_id)
    if not order:
        await call.answer("Груз не найден", show_alert=True)
        return
    if not await order_payment_confirmed(order):
        await call.answer("Сначала должна пройти оплата", show_alert=True)
        await call.message.answer(
            f"⚠️ Автостатусы для <b>{safe(order['tracking_code'])}</b> пока не включены.\n\n"
            "Логика такая: клиент оформляет доставку → оплачивает → админу приходит уведомление → только потом админ включает автостатусы.",
            reply_markup=admin_keyboard(),
        )
        return
    updated = await enable_auto_status_for_order(order_id, call.from_user.id)
    if not updated:
        await call.answer("Не удалось включить автостатусы", show_alert=True)
        return
    route_name = auto_route_label(updated["auto_status_route"] or choose_auto_route(updated["from_country"] or "", updated["to_city"] or ""))
    await call.message.answer(
        f"✅ Автостатусы включены для <b>{safe(updated['tracking_code'])}</b>.\n"
        f"Маршрут выбран автоматически: <b>{safe(route_name)}</b>.\n\n"
        "Бот сам будет менять статусы по расписанию маршрута и уведомлять клиента.",
        reply_markup=order_actions_keyboard(order_id),
    )
    await call.answer("Автостатусы включены")


@router.callback_query(F.data.startswith("auto_disable:"))
async def cb_auto_disable(call: CallbackQuery):
    if not call.from_user or not is_admin(call.from_user.id):
        await call.answer("Нет доступа", show_alert=True)
        return
    order_id = int(call.data.split(":", 1)[1])
    updated = await disable_auto_status_for_order(order_id, call.from_user.id)
    if not updated:
        await call.answer("Груз не найден", show_alert=True)
        return
    await call.message.answer(f"⏸ Автостатусы для <b>{safe(updated['tracking_code'])}</b> выключены.")
    await call.answer("Автостатусы выключены")


@router.callback_query(F.data.startswith("open_auto:"))
async def cb_open_auto_status(call: CallbackQuery):
    if not call.from_user or not is_admin(call.from_user.id):
        await call.answer("Нет доступа", show_alert=True)
        return
    order_id = int(call.data.split(":", 1)[1])
    order = await get_order(order_id)
    if not order:
        await call.answer("Груз не найден", show_alert=True)
        return
    route_key = order["auto_status_route"] or choose_auto_route(order["from_country"] or "", order["to_city"] or "")
    await call.message.answer(
        f"<b>🤖 Автостатусы для {safe(order['tracking_code'])}</b>\n\n"
        f"Сейчас: {'включены' if int(order['auto_status_enabled'] or 0) else 'выключены'}\n"
        f"Маршрут: {safe(auto_route_label(route_key))}\n\n"
        "Нажмите кнопку ниже, чтобы бот сам вёл этот груз по статусам.",
        reply_markup=auto_status_keyboard(order_id),
    )
    await call.answer()


@router.callback_query(F.data.startswith("open_status:"))
async def cb_open_status(call: CallbackQuery):
    if not call.from_user or not (is_admin(call.from_user.id) or is_warehouse(call.from_user.id)):
        await call.answer("Нет доступа", show_alert=True)
        return
    order_id = int(call.data.split(":", 1)[1])
    statuses = await get_all_statuses()
    await call.message.answer("Выберите новый статус:", reply_markup=status_keyboard(order_id, statuses))
    await call.answer()


@router.callback_query(F.data.startswith("st:"))
async def cb_set_status(call: CallbackQuery, bot: Bot, state: FSMContext):
    if not call.from_user or not (is_admin(call.from_user.id) or is_warehouse(call.from_user.id)):
        await call.answer("Нет доступа", show_alert=True)
        return
    _, order_id_raw, status = call.data.split(":", 2)
    order_id = int(order_id_raw)
    updated = await update_order_status(order_id, status, call.from_user.id, "Статус изменён через кнопку")
    if not updated:
        await call.answer("Груз не найден", show_alert=True)
        return
    await state.clear()
    await call.message.answer("✅ Статус обновлён.\n\n" + format_order(updated))
    await notify_user(
        bot,
        updated["user_id"],
        f"📦 Статус груза <b>{safe(updated['tracking_code'])}</b> обновлён: <b>{safe(status)}</b>",
    )
    await call.answer("Статус обновлён")



@router.callback_query(F.data.startswith("support_reply:"))
async def cb_support_reply(call: CallbackQuery, state: FSMContext):
    if not call.from_user or not is_admin(call.from_user.id):
        await call.answer("Нет доступа", show_alert=True)
        return
    ticket_id = int(call.data.split(":", 1)[1])
    ticket = await get_support_ticket(ticket_id)
    if not ticket:
        await call.answer("Обращение не найдено", show_alert=True)
        return
    await state.clear()
    await state.update_data(ticket_id=ticket_id)
    await state.set_state(AdminSupportReplyForm.text)
    await call.message.answer(f"Напишите ответ клиенту по обращению №{ticket_id} одним сообщением.")
    await call.answer()


@router.message(AdminSupportReplyForm.text)
async def admin_support_reply_finish(message: Message, state: FSMContext, bot: Bot):
    if not message.from_user or not is_admin(message.from_user.id):
        return
    data = await state.get_data()
    ticket_id = int(data["ticket_id"])
    ticket = await add_support_message(ticket_id, message.from_user.id, "admin", message.text or "", "answered")
    await state.clear()
    if not ticket:
        await message.answer("Обращение не найдено.", reply_markup=admin_keyboard())
        return
    await message.answer(f"✅ Ответ отправлен клиенту по обращению №{ticket_id}.", reply_markup=admin_keyboard())
    await notify_user(
        bot,
        ticket["user_id"],
        f"🛠️ Ответ техподдержки по обращению №{ticket_id}:\n\n{safe(message.text)}\n\nЕсли вопрос не решён, нажмите «🛠️ Техподдержка» → «📋 Мои обращения» и ответьте в этот тикет.",
    )


@router.callback_query(F.data.startswith("support_close:"))
async def cb_support_close(call: CallbackQuery, bot: Bot):
    if not call.from_user or not is_admin(call.from_user.id):
        await call.answer("Нет доступа", show_alert=True)
        return
    ticket_id = int(call.data.split(":", 1)[1])
    ticket = await close_support_ticket(ticket_id)
    if not ticket:
        await call.answer("Обращение не найдено", show_alert=True)
        return
    await notify_user(bot, ticket["user_id"], f"✅ Ваше обращение №{ticket_id} закрыто. Спасибо за обращение.")
    await call.answer("Обращение закрыто")
    try:
        await call.message.edit_text((call.message.text or "") + "\n\n✅ Закрыто")
    except TelegramBadRequest:
        pass


@router.callback_query(F.data.startswith("close_complaint:"))
async def cb_close_complaint(call: CallbackQuery):
    if not call.from_user or not is_admin(call.from_user.id):
        await call.answer("Нет доступа", show_alert=True)
        return
    complaint_id = int(call.data.split(":", 1)[1])
    await close_complaint(complaint_id)
    await call.answer("Жалоба закрыта")
    try:
        await call.message.edit_text((call.message.text or "") + "\n\n✅ Закрыто")
    except TelegramBadRequest:
        pass


# =========================
# FREE AI / INTENT FALLBACK
# =========================
@router.message(F.text)
async def free_ai_router(message: Message, state: FSMContext):
    # This is the free AI-like layer: no paid LLM API, just intent detection + scenarios.
    if not message.from_user:
        return
    await upsert_user(message)
    text = message.text or ""
    code = parse_tracking_code(text)
    if code:
        order = await get_order_by_code(code)
        if order:
            await message.answer(format_order(order), reply_markup=client_keyboard())
            return

    intent = detect_intent(text)
    if intent == "tracking":
        await track_start(message, state)
        return
    if intent == "calculator":
        await calc_start(message, state)
        return
    if intent == "buyout":
        await buyout_start(message, state)
        return
    if intent == "wholesale":
        await wholesale_start(message, state)
        return
    if intent == "support":
        await support_start(message, state)
        return
    if intent == "complaint":
        await complaint_start(message, state)
        return
    if intent == "customs":
        await message.answer(
            "📄 <b>Чек‑лист для документов</b>\n\n"
            "Обычно нужны:\n"
            "— описание товара\n"
            "— количество\n"
            "— примерная стоимость\n"
            "— страна отправки\n"
            "— инвойс или ссылка на товар\n"
            "— упаковочный лист, если есть\n\n"
            "Я не оформляю юридическое решение по таможне, но могу собрать данные для менеджера.",
            reply_markup=client_keyboard(),
        )
        return

    await message.answer(
        "Я могу помочь с доставкой, расчётом, статусом груза, выкупом товара, оптовой партией или жалобой.\n\n"
        "Выберите действие в меню 👇",
        reply_markup=client_keyboard(),
    )



# =========================
# FREE AUTO STATUS ENGINE
# =========================
async def process_auto_statuses(bot: Bot) -> int:
    if not AUTO_STATUS_ENABLED:
        return 0
    async with get_db() as db:
        rows = await db_fetchall(
            db,
            """
            SELECT * FROM orders
            WHERE COALESCE(auto_status_enabled, 0)=1
              AND status NOT IN ('доставлен', 'отменён', 'проблема', 'передан курьеру')
            ORDER BY id ASC
            LIMIT 100
            """,
        )
    changed = 0
    for order in rows:
        try:
            target = get_auto_target_status(order)
            if not target:
                continue
            status, comment, route_name = target
            updated = await update_order_status(order["id"], status, 0, f"Автостатус: {comment}. Маршрут: {route_name}")
            if not updated:
                continue
            async with get_db() as db:
                await db.execute("UPDATE orders SET auto_status_last_at=? WHERE id=?", (now_iso(), order["id"]))
                await db.commit()
            changed += 1
            await notify_user(
                bot,
                updated["user_id"],
                status_notify_text(updated["tracking_code"], status, comment),
                reply_markup=track_button(updated["tracking_code"]),
            )
        except Exception as e:
            logger.exception("Auto status failed for order %s: %s", order["id"], e)
    return changed


async def auto_status_loop(bot: Bot) -> None:
    await asyncio.sleep(15)
    logger.info("Auto status engine: enabled=%s on_new=%s interval=%s unit=%s", AUTO_STATUS_ENABLED, AUTO_STATUS_ON_NEW_ORDERS, AUTO_STATUS_INTERVAL_SECONDS, AUTO_STATUS_TIME_UNIT)
    while True:
        try:
            changed = await process_auto_statuses(bot)
            if changed:
                logger.info("Auto status engine changed %s orders", changed)
        except Exception as e:
            logger.exception("Auto status loop error: %s", e)
        # Для реальной работы обычно 300 сек. Для демо можно поставить 10-30 сек.
        await asyncio.sleep(max(10, AUTO_STATUS_INTERVAL_SECONDS))



# =========================
# PUBLIC WEB TRACKING MODULE
# =========================
def _web_response(html_text: str, status: int = 200) -> web.Response:
    return web.Response(
        text=html_text,
        status=status,
        content_type="text/html",
        charset="utf-8",
        headers={"Cache-Control": "no-store"},
    )


def _status_badge_color(status: str) -> str:
    s = normalize(status)
    if "готов" in s or "достав" in s:
        return "#1fbf75"
    if "проблем" in s or "отмен" in s or "задерж" in s:
        return "#ff4d4f"
    if "тамож" in s:
        return "#ffb020"
    if "пути" in s or "отправ" in s:
        return "#2f80ed"
    return "#6c7a89"


def _track_layout(title: str, content: str) -> str:
    company = html.escape(COMPANY_NAME)
    return f"""<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>{html.escape(title)} — {company}</title>
  <style>
    :root {{
      --bg:#071527;
      --card:#ffffff;
      --muted:#64748b;
      --text:#102033;
      --blue:#1e88e5;
      --cyan:#00bcd4;
      --line:#e8eef6;
    }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; font-family:Arial, sans-serif; background:linear-gradient(135deg,#071527,#0b284a 55%,#073f5d); color:var(--text); }}
    .wrap {{ min-height:100vh; display:flex; align-items:center; justify-content:center; padding:24px; }}
    .shell {{ width:100%; max-width:860px; }}
    .hero {{ color:#fff; margin-bottom:18px; }}
    .hero h1 {{ margin:0 0 8px; font-size:34px; line-height:1.08; }}
    .hero p {{ margin:0; color:#c7d7ea; font-size:16px; }}
    .card {{ background:var(--card); border-radius:24px; padding:24px; box-shadow:0 24px 70px rgba(0,0,0,.28); }}
    .form {{ display:flex; gap:10px; margin-top:18px; }}
    input {{ flex:1; padding:16px 18px; border:1px solid var(--line); border-radius:14px; font-size:16px; outline:none; }}
    input:focus {{ border-color:var(--blue); box-shadow:0 0 0 3px rgba(30,136,229,.12); }}
    button {{ padding:16px 20px; border:0; border-radius:14px; background:linear-gradient(135deg,var(--blue),var(--cyan)); color:#fff; font-weight:700; font-size:16px; cursor:pointer; }}
    .hint {{ color:var(--muted); font-size:14px; margin-top:12px; }}
    .grid {{ display:grid; grid-template-columns:1fr 1fr; gap:14px; margin-top:16px; }}
    .item {{ background:#f8fbff; border:1px solid var(--line); border-radius:16px; padding:14px; }}
    .label {{ color:var(--muted); font-size:12px; text-transform:uppercase; letter-spacing:.04em; margin-bottom:6px; }}
    .value {{ font-size:17px; font-weight:700; color:#0f2744; }}
    .status {{ display:inline-flex; align-items:center; gap:8px; color:white; border-radius:999px; padding:10px 14px; font-weight:700; }}
    .history {{ margin-top:20px; border-top:1px solid var(--line); padding-top:16px; }}
    .smart {{ margin-top:16px; padding:16px; border-radius:18px; background:linear-gradient(135deg,#eef8ff,#f7fffb); border:1px solid #d8edf7; }}
    .smart h3 {{ margin:0 0 8px; }}
    .photos {{ display:flex; flex-wrap:wrap; gap:10px; margin-top:10px; }}
    .photo {{ background:#f1f5f9; border:1px solid var(--line); border-radius:12px; padding:10px; font-size:13px; color:var(--muted); }}
    .support-btn {{ display:inline-block; margin-top:14px; padding:14px 16px; border-radius:14px; background:linear-gradient(135deg,var(--blue),var(--cyan)); color:white; font-weight:700; }}
    .trust {{ margin-top:16px; padding:18px; border-radius:20px; border:1px solid #d8edf7; }}
    .trust.green {{ background:#effcf6; border-color:#b9efd3; }}
    .trust.yellow {{ background:#fff8e7; border-color:#ffe0a3; }}
    .trust.red {{ background:#fff1f1; border-color:#ffc4c4; }}
    .trust-title {{ font-size:21px; font-weight:800; margin-bottom:8px; }}
    .trust-grid {{ display:grid; grid-template-columns:1fr 1fr; gap:12px; margin-top:12px; }}
    .promise {{ margin-top:18px; padding:20px; border-radius:22px; border:1px solid #d8edf7; }}
    .promise.green {{ background:#effcf6; border-color:#b9efd3; }}
    .promise.yellow {{ background:#fff8e7; border-color:#ffe0a3; }}
    .promise.red {{ background:#fff1f1; border-color:#ffc4c4; }}
    .promise-head {{ display:flex; align-items:center; justify-content:space-between; gap:16px; }}
    .promise-kicker {{ color:var(--muted); font-size:12px; text-transform:uppercase; letter-spacing:.05em; font-weight:700; }}
    .promise-title {{ font-size:23px; font-weight:900; margin-top:4px; }}
    .score {{ min-width:92px; min-height:92px; border-radius:50%; display:flex; align-items:center; justify-content:center; flex-direction:column; background:#0f2744; color:#fff; font-size:30px; font-weight:900; }}
    .score span {{ font-size:13px; opacity:.8; }}
    .promise-two {{ display:grid; grid-template-columns:1fr 1fr; gap:14px; margin-top:14px; }}
    .promise-two > div {{ background:rgba(255,255,255,.65); border:1px solid rgba(15,39,68,.08); border-radius:16px; padding:14px; }}
    .promise ul {{ margin:8px 0 0 18px; padding:0; }}
    @media (max-width:640px) {{ .promise-head {{ align-items:flex-start; }} .score {{ min-width:76px; min-height:76px; font-size:24px; }} .promise-two {{ grid-template-columns:1fr; }} }}
    .paybox {{ margin-top:16px; padding:18px; border-radius:20px; background:#f7fbff; border:1px solid #d8edf7; }}
    .pay-actions {{ display:flex; flex-wrap:wrap; gap:10px; margin-top:14px; }}
    .pay-link {{ display:inline-block; padding:14px 16px; border-radius:14px; background:linear-gradient(135deg,var(--blue),var(--cyan)); color:white; font-weight:700; }}
    .pay-secondary {{ display:inline-block; padding:14px 16px; border-radius:14px; background:#eef2f7; color:#0f2744; font-weight:700; }}
    .bank-grid {{ display:grid; grid-template-columns:repeat(2, 1fr); gap:12px; margin-top:16px; }}
    .bank-card {{ display:block; padding:18px; border-radius:18px; color:white; font-weight:800; min-height:100px; box-shadow:0 10px 30px rgba(0,0,0,.14); }}
    .bank-card span {{ display:block; font-weight:400; opacity:.92; margin-top:6px; font-size:14px; }}
    @media (max-width:640px) {{ .bank-grid {{ grid-template-columns:1fr; }} }}
    .hrow {{ display:flex; gap:12px; padding:10px 0; border-bottom:1px solid #f0f3f7; }}
    .dot {{ width:10px; height:10px; border-radius:50%; background:var(--blue); margin-top:5px; flex:0 0 auto; }}
    .hdate {{ color:var(--muted); font-size:13px; min-width:132px; }}
    .hstatus {{ font-weight:700; }}
    .error {{ background:#fff3f3; border:1px solid #ffd4d4; color:#a01616; border-radius:16px; padding:16px; }}
    .footer {{ text-align:center; color:#bed1e5; font-size:12px; margin-top:14px; }}
    a {{ color:var(--blue); text-decoration:none; }}
    @media (max-width:640px) {{
      .wrap {{ padding:16px; align-items:flex-start; }}
      .hero h1 {{ font-size:27px; }}
      .form {{ flex-direction:column; }}
      .grid {{ grid-template-columns:1fr; }}
      .hrow {{ display:block; }}
      .hdate {{ margin-bottom:4px; }}
    }}
  </style>
</head>
<body>
  <div class="wrap">
    <div class="shell">
      <div class="hero">
        <h1>{company}</h1>
        <p>CargoPromise OS: обещания, риски и доказательная история груза</p>
      </div>
      <div class="card">{content}</div>
      <div class="footer">CargoPromise OS · контроль обещаний и рисков в карго</div>
    </div>
  </div>
</body>
</html>"""


def _track_form(code: str = "", error: str = "") -> str:
    err = f'<div class="error">{html.escape(error)}</div>' if error else ""
    return f"""
      <h2 style="margin:0 0 8px;">Отследить груз</h2>
      <div class="hint">Введите номер груза, например <b>CG26050500001</b>.</div>
      {err}
      <form class="form" method="get" action="/track">
        <input name="code" value="{html.escape(code or '')}" placeholder="Номер груза CG..." autocomplete="off" />
        <button type="submit">Проверить</button>
      </form>
      <div class="hint">Если номер не найден, свяжитесь с менеджером карго-компании.</div>
    """


async def landing_page(_: web.Request) -> web.Response:
    return _web_response(_track_layout("Проверка груза", _track_form()))


async def track_page(request: web.Request) -> web.Response:
    code = (request.query.get("code") or "").strip().upper()
    if not code:
        return _web_response(_track_layout("Проверка груза", _track_form()))
    return await render_tracking_result(code, request)


async def track_code_page(request: web.Request) -> web.Response:
    code = (request.match_info.get("code") or "").strip().upper()
    return await render_tracking_result(code, request)


async def render_tracking_result(code: str, request: Optional[web.Request] = None) -> web.Response:
    order = await get_order_by_code(code)
    if not order:
        return _web_response(_track_layout("Груз не найден", _track_form(code, "Груз с таким номером не найден.")), status=404)

    await record_track_view(order["tracking_code"], request)

    data = await smart_cargo_card_data(order["id"])
    history = data.get("history", [])
    photos = data.get("photos", [])
    tickets = data.get("tickets", [])
    views = await track_view_stats(order["tracking_code"])
    trust = trustflow_info(order)
    promise = promise_os_profile(order, history, photos, tickets, views)

    status = order["status"] or "новая заявка"
    badge_color = _status_badge_color(status)
    route = f"{html.escape(str(order['from_country'] or '—'))} → {html.escape(str(order['to_city'] or '—'))}"
    updated = html.escape(((order["updated_at"] or order["created_at"] or "")[:16]).replace("T", " "))

    if history:
        rows = []
        for h in history:
            dt = html.escape(((h["created_at"] or "")[:16]).replace("T", " "))
            st = html.escape(h["status"] or "")
            comment = html.escape(h["comment"] or "")
            rows.append(f'<div class="hrow"><div class="dot"></div><div class="hdate">{dt}</div><div><div class="hstatus">{st}</div><div class="hint">{comment}</div></div></div>')
        history_html = '<div class="history"><h3 style="margin:0 0 8px;">Proof Timeline</h3><div class="hint">Доказательная история по грузу: что произошло и когда.</div>' + "".join(rows) + "</div>"
    else:
        history_html = '<div class="history"><h3 style="margin:0 0 8px;">Proof Timeline</h3><div class="hint">История пока пустая.</div></div>'

    photos_html = ""
    if photos:
        items = []
        for p in photos[:5]:
            dt = html.escape(((p["created_at"] or "")[:16]).replace("T", " "))
            comment = html.escape(p["comment"] or "Фото груза")
            items.append(f'<div class="photo">📷 {comment}<br><span>{dt}</span></div>')
        photos_html = '<div class="smart"><h3>Доказательства склада</h3><div class="photos">' + "".join(items) + "</div></div>"

    support_status = "нет открытых обращений"
    if tickets:
        open_count = len([t for t in tickets if t["status"] != "closed"])
        support_status = f"{open_count} открытых обращений" if open_count else "обращения закрыты"

    support_button = ""
    if trust["show_support"] or promise["score"] < 65:
        support_button = f'<a class="support-btn" href="https://t.me/{html.escape(BOT_USERNAME)}?start=track_{html.escape(order["tracking_code"])}">Проверить у менеджера</a>'

    trust_html = f"""
      <div class="trust {html.escape(trust['level'])}">
        <div class="trust-title">{'🟢' if trust['level']=='green' else '🟡' if trust['level']=='yellow' else '🔴'} {html.escape(trust['label'])}</div>
        <div>{html.escape(trust['status_explanation'])}</div>
        <div class="trust-grid">
          <div class="item"><div class="label">Следующий этап</div><div class="value">{html.escape(trust['next_stage'])}</div></div>
          <div class="item"><div class="label">Ожидаем обновление</div><div class="value">{html.escape(trust['next_due_text'] or 'по обработке')}</div></div>
        </div>
        <div class="hint" style="margin-top:12px;">{html.escape(trust['client_hint'])}</div>
        <div class="pay-actions">{support_button}</div>
      </div>
    """

    promise_html = promise_os_html(order, promise)

    promise_summary = f"""
      <div class="smart">
        <h3>Promise Card</h3>
        <div class="hint">Это не просто трекер. Это карточка обещания: что можно обещать клиенту, где есть риск и какие доказательства уже есть.</div>
        <div class="grid" style="margin-top:12px;">
          <div class="item"><div class="label">Обращения</div><div class="value">{html.escape(support_status)}</div></div>
          <div class="item"><div class="label">Просмотров ссылки</div><div class="value">{views['total']}</div></div>
          <div class="item"><div class="label">Risk Gate</div><div class="value">{html.escape(promise['gate'])}</div></div>
          <div class="item"><div class="label">Dispute Shield</div><div class="value">история зафиксирована</div></div>
        </div>
      </div>
    """

    content = f"""
      <h2 style="margin:0 0 8px;">Груз {html.escape(order['tracking_code'])}</h2>
      <div class="status" style="background:{badge_color};">● {html.escape(status)}</div>
      <div class="grid">
        <div class="item"><div class="label">Маршрут</div><div class="value">{route}</div></div>
        <div class="item"><div class="label">Товар</div><div class="value">{html.escape(str(order['cargo_type'] or '—'))}</div></div>
        <div class="item"><div class="label">Вес</div><div class="value">{html.escape(str(order['weight'] or '—'))} кг</div></div>
        <div class="item"><div class="label">Последнее обновление</div><div class="value">{updated or '—'}</div></div>
      </div>
      {promise_html}
      {trust_html}
      {promise_summary}
      {photos_html}
      {history_html}
      <form class="form" method="get" action="/track" style="margin-top:20px;">
        <input name="code" placeholder="Проверить другой номер CG..." autocomplete="off" />
        <button type="submit">Проверить</button>
      </form>
    """
    return _web_response(_track_layout(f"CargoPromise {order['tracking_code']}", content))


async def api_track_order(request: web.Request) -> web.Response:
    code = (request.match_info.get("code") or "").strip().upper()
    order = await get_order_by_code(code)
    if not order:
        return web.json_response({"ok": False, "error": "not_found"}, status=404)
    data = await smart_cargo_card_data(order["id"])
    history = data.get("history", [])
    photos = data.get("photos", [])
    tickets = data.get("tickets", [])
    trust = trustflow_info(order)
    views = await track_view_stats(order["tracking_code"])
    promise = promise_os_profile(order, history, photos, tickets, views)
    return web.json_response({
        "ok": True,
        "product": "CargoPromise OS",
        "tracking_code": order["tracking_code"],
        "status": order["status"],
        "promise": promise,
        "trustflow": trust,
        "from_country": order["from_country"],
        "to_city": order["to_city"],
        "cargo_type": order["cargo_type"],
        "weight": order["weight"],
        "updated_at": order["updated_at"],
        "history": [{"status": h["status"], "comment": h["comment"], "created_at": h["created_at"]} for h in history],
        "photos_count": len(photos),
        "open_tickets_count": len([t for t in tickets if t["status"] != "closed"]),
        "track_views": views,
    })


def bank_provider_cards(code: str) -> str:
    cards = []
    for key, info in DEMO_PAYMENT_PROVIDERS.items():
        cards.append(
            f'<a class="bank-card" style="background:{html.escape(info["color"])}" href="/demo-pay/{html.escape(code)}/checkout/{html.escape(key)}">'
            f'{html.escape(info["label"])}<span>{html.escape(info["subtitle"])}</span></a>'
        )
    return "".join(cards)


async def demo_payment_page(request: web.Request) -> web.Response:
    if not DEMO_PAYMENT_ENABLED:
        return _web_response(_track_layout("Демо-оплата отключена", '<div class="error">Демо-оплата отключена.</div>'), status=403)
    code = (request.match_info.get("code") or "").strip().upper()
    order = await get_order_by_code(code)
    if not order:
        return _web_response(_track_layout("Груз не найден", _track_form(code, "Груз с таким номером не найден.")), status=404)
    order = await ensure_demo_price_if_needed(order)
    amount = demo_payment_amount_for_order(order)
    content = f"""
      <h2 style="margin:0 0 8px;">💳 Выберите способ оплаты</h2>
      <div class="hint">Демо-режим: реальные деньги не списываются. Это имитация банковской оплаты для показа клиенту.</div>
      <div class="paybox">
        <div class="grid">
          <div class="item"><div class="label">Груз</div><div class="value">{html.escape(order['tracking_code'])}</div></div>
          <div class="item"><div class="label">Сумма</div><div class="value">{amount:g} {html.escape(CURRENCY)}</div></div>
        </div>
        <div class="bank-grid">{bank_provider_cards(order['tracking_code'])}</div>
      </div>
      <div class="hint">В реальном внедрении эти кнопки можно заменить на настоящие Kaspi Pay, Halyk, ЦентрКредит, Freedom или другой эквайринг.</div>
    """
    return _web_response(_track_layout("Выбор оплаты", content))


async def demo_payment_checkout_page(request: web.Request) -> web.Response:
    if not DEMO_PAYMENT_ENABLED:
        return _web_response(_track_layout("Демо-оплата отключена", '<div class="error">Демо-оплата отключена.</div>'), status=403)
    code = (request.match_info.get("code") or "").strip().upper()
    provider = (request.match_info.get("provider") or "kaspi").strip().lower()
    order = await get_order_by_code(code)
    if not order:
        return _web_response(_track_layout("Груз не найден", _track_form(code, "Груз с таким номером не найден.")), status=404)
    order = await ensure_demo_price_if_needed(order)
    payment = await create_demo_payment_request(order["tracking_code"], provider)
    if not payment:
        return _web_response(_track_layout("Ошибка оплаты", '<div class="error">Не удалось создать демо-оплату.</div>'), status=500)
    info = payment["provider_info"]
    amount = payment["amount"]
    payment_id = html.escape(payment["payment_id"])
    content = f"""
      <h2 style="margin:0 0 8px;">💳 {html.escape(info['title'])}</h2>
      <div class="hint">{html.escape(info['subtitle'])}. Демо-режим, реальные деньги не списываются.</div>
      <div class="paybox" style="border-color:{html.escape(info['color'])};">
        <div class="grid">
          <div class="item"><div class="label">Груз</div><div class="value">{html.escape(order['tracking_code'])}</div></div>
          <div class="item"><div class="label">Сумма</div><div class="value">{amount:g} {html.escape(CURRENCY)}</div></div>
          <div class="item"><div class="label">Платёж ID</div><div class="value">{payment_id}</div></div>
          <div class="item"><div class="label">Банк</div><div class="value">{html.escape(info['label'])}</div></div>
        </div>
        <div class="pay-actions">
          <a class="pay-link" style="background:{html.escape(info['color'])};" href="/demo-pay/{html.escape(order['tracking_code'])}/success/{html.escape(payment['provider'])}?payment_id={payment_id}">✅ Оплатить в демо</a>
          <a class="pay-secondary" href="/demo-pay/{html.escape(order['tracking_code'])}">↩️ Выбрать другой банк</a>
          <a class="pay-secondary" href="/track/{html.escape(order['tracking_code'])}">🔎 Вернуться к грузу</a>
        </div>
      </div>
      <div class="hint">В реальной версии сюда подключается официальный платёжный провайдер выбранного банка.</div>
    """
    return _web_response(_track_layout(f"Оплата {info['label']}", content))


async def demo_payment_success_page(request: web.Request) -> web.Response:
    if not DEMO_PAYMENT_ENABLED:
        return _web_response(_track_layout("Демо-оплата отключена", '<div class="error">Демо-оплата отключена.</div>'), status=403)
    code = (request.match_info.get("code") or "").strip().upper()
    provider = (request.match_info.get("provider") or "kaspi").strip().lower()
    payment_id = (request.query.get("payment_id") or "").strip()
    payment_result = await complete_demo_payment(code, provider, payment_id, 0)
    if not payment_result:
        return _web_response(_track_layout("Груз не найден", _track_form(code, "Груз с таким номером не найден.")), status=404)

    updated = payment_result["order"]
    amount = float(payment_result["amount"] or 0)
    provider_key = payment_result["provider"]
    info = provider_info(provider_key)

    # Уведомляем админа один раз, только когда оплата впервые стала paid.
    if not payment_result.get("already_paid"):
        bot = request.app.get("bot")
        if bot:
            await notify_admins(
                bot,
                admin_payment_notify_text(updated, provider_key, amount, payment_result.get("payment_id") or payment_id),
                reply_markup=order_actions_keyboard(updated["id"]),
            )

    content = f"""
      <h2 style="margin:0 0 8px;">✅ Оплата прошла</h2>
      <div class="hint">Это виртуальная демо-оплата через {html.escape(info['label'])}. В реальной версии платёж подтверждает банк.</div>
      <div class="paybox" style="border-color:{html.escape(info['color'])};">
        <div class="grid">
          <div class="item"><div class="label">Груз</div><div class="value">{html.escape(updated['tracking_code'])}</div></div>
          <div class="item"><div class="label">Статус оплаты</div><div class="value">{html.escape(updated['payment_status'] or 'оплачено')}</div></div>
          <div class="item"><div class="label">Оплачено</div><div class="value">{html.escape(str(updated['paid_amount'] or 0))} {html.escape(CURRENCY)}</div></div>
          <div class="item"><div class="label">Банк</div><div class="value">{html.escape(info['label'])}</div></div>
        </div>
        <div class="pay-actions">
          <a class="pay-link" href="/track/{html.escape(updated['tracking_code'])}">🔎 Посмотреть груз</a>
        </div>
      </div>
    """
    return _web_response(_track_layout("Оплата успешна", content))


async def api_demo_payment_create(request: web.Request) -> web.Response:
    code = (request.match_info.get("code") or "").strip().upper()
    provider = (request.match_info.get("provider") or request.query.get("provider") or "kaspi").strip().lower()
    if not DEMO_PAYMENT_ENABLED:
        return web.json_response({"ok": False, "error": "demo_payment_disabled"}, status=403)
    payment = await create_demo_payment_request(code, provider)
    if not payment:
        return web.json_response({"ok": False, "error": "order_not_found"}, status=404)
    return web.json_response({
        "ok": True,
        "mode": "demo",
        "payment_id": payment["payment_id"],
        "tracking_code": payment["order"]["tracking_code"],
        "amount": payment["amount"],
        "currency": CURRENCY,
        "provider": payment["provider"],
        "provider_label": payment["provider_info"]["label"],
        "checkout_url": payment["checkout_url"],
    })


async def api_demo_payment_success(request: web.Request) -> web.Response:
    code = (request.match_info.get("code") or "").strip().upper()
    provider = (request.match_info.get("provider") or request.query.get("provider") or "kaspi").strip().lower()
    payment_id = (request.query.get("payment_id") or "").strip()
    if not DEMO_PAYMENT_ENABLED:
        return web.json_response({"ok": False, "error": "demo_payment_disabled"}, status=403)
    payment_result = await complete_demo_payment(code, provider, payment_id, 0)
    if not payment_result:
        return web.json_response({"ok": False, "error": "order_not_found"}, status=404)

    updated = payment_result["order"]
    amount = float(payment_result["amount"] or 0)
    provider_key = payment_result["provider"]

    if not payment_result.get("already_paid"):
        bot = request.app.get("bot")
        if bot:
            await notify_admins(
                bot,
                admin_payment_notify_text(updated, provider_key, amount, payment_result.get("payment_id") or payment_id),
                reply_markup=order_actions_keyboard(updated["id"]),
            )

    return web.json_response({
        "ok": True,
        "mode": "demo",
        "tracking_code": updated["tracking_code"],
        "payment_status": updated["payment_status"],
        "paid_amount": updated["paid_amount"],
        "currency": CURRENCY,
        "provider": provider_key,
        "provider_label": provider_info(provider_key)["label"],
        "admin_notified": not payment_result.get("already_paid"),
    })


# =========================
# HEALTH SERVER
# =========================
async def health(_: web.Request) -> web.Response:
    return web.json_response({"ok": True, "service": COMPANY_NAME})


async def start_health_server(bot: Bot) -> None:
    # Публичный модуль отслеживания для сайта клиента + health-check для хостинга.
    app = web.Application()
    app["bot"] = bot
    app.router.add_get("/", landing_page)
    app.router.add_get("/track", track_page)
    app.router.add_get("/track/{code}", track_code_page)
    app.router.add_get("/promise/{code}", track_code_page)
    app.router.add_get("/passport/{code}", track_code_page)
    app.router.add_get("/api/track/{code}", api_track_order)
    app.router.add_get("/demo-pay/{code}", demo_payment_page)
    app.router.add_get("/demo-pay/{code}/checkout/{provider}", demo_payment_checkout_page)
    app.router.add_get("/demo-pay/{code}/success/{provider}", demo_payment_success_page)
    app.router.add_get("/api/demo-payment/create/{code}", api_demo_payment_create)
    app.router.add_get("/api/demo-payment/create/{provider}/{code}", api_demo_payment_create)
    app.router.add_get("/api/demo-payment/success/{code}", api_demo_payment_success)
    app.router.add_get("/api/demo-payment/success/{provider}/{code}", api_demo_payment_success)
    app.router.add_get("/health", health)
    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, "0.0.0.0", PORT)
    await site.start()
    logger.info("Health server started on port %s", PORT)


async def owner_report_loop(bot: Bot) -> None:
    """Резервная функция авто-отчёта.

    По умолчанию не запускается. Отчёт админ получает вручную через кнопку
    📊 SmartFlow отчёт, чтобы бот не присылал отчёты сам и не сбивал меню.
    """
    if not OWNER_REPORT_ENABLED:
        return
    await asyncio.sleep(max(3600, OWNER_REPORT_INTERVAL_SECONDS))
    while True:
        try:
            text = await owner_smartflow_report_text()
            await notify_admins(bot, text, reply_markup=admin_keyboard())
        except Exception as e:
            logger.exception("Owner SmartFlow report failed: %s", e)
        await asyncio.sleep(max(3600, OWNER_REPORT_INTERVAL_SECONDS))


async def main() -> None:
    await init_db()
    bot = Bot(BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
    dp = Dispatcher(storage=MemoryStorage())
    dp.include_router(router)
    asyncio.create_task(start_health_server(bot))
    asyncio.create_task(auto_status_loop(bot))
    logger.info("Starting polling. DB=%s", DB_PATH)
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
